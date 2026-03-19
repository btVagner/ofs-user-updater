from flask import render_template, request, redirect, url_for, flash, send_file
from io import BytesIO
from datetime import datetime
import re
import unicodedata
from openpyxl import load_workbook, Workbook

from database.connection import get_connection
from core.auth import login_required, perm_required


def normalize_appt_number(value):
    s = str(value or "").strip()
    if not s:
        return ""
    return re.sub(r"-[^/]+(?=/)", "", s)
def normalize_text_compare(value):
    s = str(value or "").strip().upper()
    if not s:
        return ""
    s_norm = unicodedata.normalize("NFKD", s)
    return "".join(ch for ch in s_norm if not unicodedata.combining(ch))

def normalize_crm_os(numero_os, ocorrencia):
    numero = str(numero_os or "").strip()
    ocorr = str(ocorrencia or "").strip()

    if not numero or not ocorr:
        return ""

    raw = f"{numero}/{ocorr}"
    return normalize_appt_number(raw)


def _cell_to_str(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value).strip()


def load_crm_rows_from_xlsx(file_storage):
    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("O arquivo está vazio.")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    header_map = {h: idx for idx, h in enumerate(headers)}
    required_cols = [
        "NUMERO_OS",
        "OCORRENCIA",
        "STATUS",
        "TIPO_CONTRATO",
        "DATA_AGENDADA",
        "NOME_CLIENTE",
        "CPF",
        "CIDADE",
        "MOTIVO_ABERTURA",
        "SERVICO_ABERTURA",
    ]
    missing = [col for col in required_cols if col not in header_map]
    if missing:
        raise ValueError(
            f"Colunas obrigatórias ausentes no XLSX: {', '.join(missing)}"
        )

    idx_numero_os = header_map["NUMERO_OS"]
    idx_ocorrencia = header_map["OCORRENCIA"]
    idx_status = header_map["STATUS"]
    idx_tipo_contrato = header_map["TIPO_CONTRATO"]
    idx_data_agendada = header_map["DATA_AGENDADA"]
    idx_nome_cliente = header_map["NOME_CLIENTE"]
    idx_cpf = header_map["CPF"]
    idx_cidade = header_map["CIDADE"]
    idx_motivo_abertura = header_map["MOTIVO_ABERTURA"]
    idx_servico_abertura = header_map["SERVICO_ABERTURA"]

    records = []

    for row in rows[1:]:
        numero_os = row[idx_numero_os] if idx_numero_os < len(row) else None
        ocorrencia = row[idx_ocorrencia] if idx_ocorrencia < len(row) else None
        status = row[idx_status] if idx_status < len(row) else None
        tipo_contrato = row[idx_tipo_contrato] if idx_tipo_contrato < len(row) else None
        data_agendada = row[idx_data_agendada] if idx_data_agendada < len(row) else None
        nome_cliente = row[idx_nome_cliente] if idx_nome_cliente < len(row) else None
        cpf = row[idx_cpf] if idx_cpf < len(row) else None
        cidade = row[idx_cidade] if idx_cidade < len(row) else None

        status_str = _cell_to_str(status)
        tipo_contrato_str = _cell_to_str(tipo_contrato)

        status_cmp = normalize_text_compare(status_str)
        tipo_contrato_cmp = normalize_text_compare(tipo_contrato_str)
        motivo_abertura = row[idx_motivo_abertura] if idx_motivo_abertura < len(row) else None
        servico_abertura = row[idx_servico_abertura] if idx_servico_abertura < len(row) else None

        if status_cmp != "AGENDADA":
            continue

        if tipo_contrato_cmp != "REDE PROPRIA":
            continue
        numero_os_str = _cell_to_str(numero_os)
        ocorrencia_str = _cell_to_str(ocorrencia)

        if not numero_os_str or not ocorrencia_str:
            continue

        appt_number = f"{numero_os_str}/{ocorrencia_str}"
        appt_number_norm = normalize_crm_os(numero_os, ocorrencia)

        if not appt_number_norm:
            continue

        records.append((
            numero_os_str,
            ocorrencia_str,
            appt_number,
            appt_number_norm,
            status_str or None,
            tipo_contrato_str or None,
            _cell_to_date_str(data_agendada) or None,
            _cell_to_str(nome_cliente) or None,
            _cell_to_str(cpf) or None,
            _cell_to_str(cidade) or None,
            _cell_to_str(motivo_abertura) or None,
            _cell_to_str(servico_abertura) or None,
        ))

    if not records:
        raise ValueError(
            "Nenhuma linha válida encontrada após aplicar os filtros "
            "STATUS=Agendada e TIPO_CONTRATO=REDE PRÓPRIA."
        )

    return records


def truncate_and_insert_crm_rows(records):
    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("TRUNCATE TABLE ofs_agendamentos_crm")

        cur.executemany("""
        INSERT INTO ofs_agendamentos_crm (
            numero_os,
            ocorrencia,
            appt_number,
            appt_number_norm,
            status,
            tipo_contrato,
            data_agendada,
            nome_cliente,
            cpf,
            cidade,
            motivo_abertura,
            servico_abertura
        )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, records)

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()
def format_data_agendada_display(value):
    s = str(value or "").strip()
    if not s:
        return ""

    # remove parte da hora se existir
    if " " in s:
        s = s.split(" ")[0].strip()

    # yyyy-mm-dd
    if len(s) == 10 and s[4] == "-" and s[7] == "-":
        yyyy, mm, dd = s.split("-")
        return f"{dd}/{mm}/{yyyy}"

    # dd/mm/yyyy
    if len(s) == 10 and s[2] == "/" and s[5] == "/":
        return s

    # dd/mm/yy
    if len(s) == 8 and s[2] == "/" and s[5] == "/":
        dd, mm, yy = s.split("/")
        return f"{dd}/{mm}/20{yy}"

    return s
def get_erros_agendamento_items(limit=50):
    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    try:
        sql = """
            SELECT
                crm.appt_number_norm,
                crm.status,
                crm.data_agendada,
                crm.nome_cliente,
                crm.cpf,
                crm.cidade,
                crm.motivo_abertura,
                crm.servico_abertura,
                CASE
                    WHEN base.activity_id IS NULL THEN 'Não'
                    ELSE 'Sim'
                END AS existe_no_ofs
            FROM ofs_agendamentos_crm crm
            LEFT JOIN ofs_atividades_base base
                ON base.appt_number_norm = crm.appt_number_norm
            ORDER BY
                CASE WHEN base.activity_id IS NULL THEN 0 ELSE 1 END,
                crm.data_agendada ASC,
                crm.appt_number_norm ASC
        """

        if limit is not None:
            sql += " LIMIT %s"
            cur.execute(sql, (limit,))
        else:
            cur.execute(sql)

        items = cur.fetchall()

        for item in items:
            item["data_agendada_fmt"] = format_data_agendada_display(item.get("data_agendada"))

        cur.execute("""
            SELECT COUNT(*) AS total_importados
            FROM ofs_agendamentos_crm
        """)
        total_importados = int((cur.fetchone() or {}).get("total_importados") or 0)

        cur.execute("""
            SELECT COUNT(*) AS total_erros
            FROM ofs_agendamentos_crm crm
            LEFT JOIN ofs_atividades_base base
                ON base.appt_number_norm = crm.appt_number_norm
            WHERE base.activity_id IS NULL
        """)
        total_erros = int((cur.fetchone() or {}).get("total_erros") or 0)

        cur.execute("""
            SELECT MIN(data_agendada) AS min_data_agendada
            FROM ofs_agendamentos_crm
        """)
        row_min = cur.fetchone() or {}
        min_data_agendada = format_data_agendada_display(row_min.get("min_data_agendada"))

        cur.execute("""
            SELECT
                crm.data_agendada
            FROM ofs_agendamentos_crm crm
            LEFT JOIN ofs_atividades_base base
                ON base.appt_number_norm = crm.appt_number_norm
            WHERE base.activity_id IS NULL
        """)
        rows_chart = cur.fetchall()

        chart_map = {}

        for row in rows_chart:
            label = format_data_agendada_display(row.get("data_agendada"))
            if not label:
                continue
            chart_map[label] = chart_map.get(label, 0) + 1

        def _sort_key_date_br(value):
            try:
                return datetime.strptime(value, "%d/%m/%Y")
            except Exception:
                return datetime.max

        sorted_labels = sorted(chart_map.keys(), key=_sort_key_date_br)

        dashboard_data = {
            "labels": sorted_labels,
            "values": [chart_map[label] for label in sorted_labels],
        }

        return items, total_erros, total_importados, min_data_agendada, dashboard_data

    finally:
        cur.close()
        conn.close()

def export_erros_agendamento_xlsx():
    items, _, _, _, _ = get_erros_agendamento_items(limit=None)

    wb = Workbook()
    ws = wb.active
    ws.title = "ErrosAgendamento"

    headers = [
        "appt_number_norm",
        "status",
        "data_agendada",
        "nome_cliente",
        "cpf",
        "cidade",
        "motivo_abertura",
        "servico_abertura",
        "existe_no_ofs",
    ]
    ws.append(headers)

    for row in items:
        ws.append([
            row.get("appt_number_norm"),
            row.get("status"),
            row.get("data_agendada"),
            row.get("nome_cliente"),
            row.get("cpf"),
            row.get("cidade"),
            row.get("motivo_abertura"),
            row.get("servico_abertura"),
            row.get("existe_no_ofs"),
        ])

    for col_idx, col_name in enumerate(headers, start=1):
        max_len = len(col_name)
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def _cell_to_date_str(value):
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")

    s = str(value).strip()
    if not s:
        return ""

    if " " in s:
        s = s.split(" ")[0].strip()

    # yyyy-mm-dd
    if len(s) == 10 and s[4] == "-" and s[7] == "-":
        yyyy, mm, dd = s.split("-")
        return f"{dd}/{mm}/{yyyy}"

    # dd/mm/yyyy
    if len(s) == 10 and s[2] == "/" and s[5] == "/":
        return s

    # dd/mm/yy
    if len(s) == 8 and s[2] == "/" and s[5] == "/":
        dd, mm, yy = s.split("/")
        return f"{dd}/{mm}/20{yy}"

    return s
def init_app(app):

    @app.route("/ofs/erros-agendamento", methods=["GET", "POST"])
    @login_required
    @perm_required("ofs.erros_agendamento")
    def ofs_erros_agendamento():
        items = []
        total = 0
        total_importados = 0
        min_data_agendada = None
        dashboard_data = {"labels": [], "values": []}

        if request.method == "POST":
            file = request.files.get("file")

            if not file or not file.filename:
                flash("Selecione um arquivo XLSX para importar.", "error")
                items, total, total_importados, min_data_agendada, dashboard_data = get_erros_agendamento_items(limit=50)
                return render_template(
                    "atividades_base/ofs_erros_agendamento.html",
                    items=items,
                    total=total,
                    total_importados=total_importados,
                    min_data_agendada=min_data_agendada,
                    dashboard_data=dashboard_data,
                )

            original_name = (file.filename or "").strip()
            if not original_name.lower().endswith(".xlsx"):
                flash("Arquivo inválido. Envie um arquivo .xlsx.", "error")
                items, total, total_importados, min_data_agendada, dashboard_data = get_erros_agendamento_items(limit=50)
                return render_template(
                    "atividades_base/ofs_erros_agendamento.html",
                    items=items,
                    total=total,
                    total_importados=total_importados,
                    min_data_agendada=min_data_agendada,
                    dashboard_data=dashboard_data,
                )

            try:
                records = load_crm_rows_from_xlsx(file)
                truncate_and_insert_crm_rows(records)

                flash(
                    f"Arquivo processado com sucesso. Registros importados: {len(records)}",
                    "success"
                )
                return redirect(url_for("ofs_erros_agendamento"))

            except Exception as e:
                flash(f"Erro ao processar arquivo: {str(e)}", "error")
                return redirect(url_for("ofs_erros_agendamento"))

        items, total, total_importados, min_data_agendada, dashboard_data = get_erros_agendamento_items(limit=50)

        return render_template(
            "atividades_base/ofs_erros_agendamento.html",
            items=items,
            total=total,
            total_importados=total_importados,
            min_data_agendada=min_data_agendada,
            dashboard_data=dashboard_data,
        )
    @app.route("/ofs/erros-agendamento/export/xlsx", methods=["GET"])
    @login_required
    @perm_required("ofs.erros_agendamento")
    def ofs_erros_agendamento_export_xlsx():
        try:
            output = export_erros_agendamento_xlsx()
            filename = f"erros_agendamento_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            flash(f"Erro ao exportar arquivo: {str(e)}", "error")
            return redirect(url_for("ofs_erros_agendamento"))