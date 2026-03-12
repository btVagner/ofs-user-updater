from flask import render_template, request, jsonify, send_file, flash, redirect, url_for, session
from datetime import datetime, date
from io import BytesIO
import threading
import xlsxwriter
import re

from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from database.connection import get_connection
from core.auth import login_required, perm_required, current_actor
from core.utils import xlsx_auto_width
from services.ofs_activities_errors_importer import (
    validate_max_range_7_days,
    run_import_job,
)
def _normalize_appt_number(value):
    s = str(value or "").strip()
    if not s:
        return ""
    return re.sub(r"-[^/]+(?=/)", "", s)
def _excel_date_to_str(value):
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    s = str(value).strip()
    if not s:
        return None

    formatos = [
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%d-%m-%Y",
    ]

    for fmt in formatos:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass

    return None

def _import_pending_close_xlsx_to_db(file_storage):
    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("O arquivo está vazio.")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    header_map = {h: idx for idx, h in enumerate(headers)}

    required_cols = ["NUMERO_OSE", "STATUS_OSE", "DATA_AGENDAMENTO"]
    missing = [col for col in required_cols if col not in header_map]
    if missing:
        raise ValueError(f"Colunas obrigatórias ausentes no XLSX: {', '.join(missing)}")

    idx_numero_ose = header_map["NUMERO_OSE"]
    idx_status_ose = header_map["STATUS_OSE"]
    idx_data_agendamento = header_map["DATA_AGENDAMENTO"]

    records = []

    for row in rows[1:]:
        numero_ose = row[idx_numero_ose] if idx_numero_ose < len(row) else None
        status_ose = row[idx_status_ose] if idx_status_ose < len(row) else None
        data_agendamento = row[idx_data_agendamento] if idx_data_agendamento < len(row) else None

        numero_ose_str = str(numero_ose or "").strip()
        numero_ose_norm = _normalize_appt_number(numero_ose)
        data_agendamento_str = _excel_date_to_str(data_agendamento)

        if not numero_ose_str or not numero_ose_norm or not data_agendamento_str:
            continue

        records.append((
            numero_ose_str,
            numero_ose_norm,
            str(status_ose or "").strip(),
            data_agendamento_str
        ))

    if not records:
        raise ValueError("Nenhuma linha válida foi encontrada no XLSX.")

    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("TRUNCATE TABLE ofs_pending_close_ng")

        cur.executemany("""
            INSERT INTO ofs_pending_close_ng (
                numero_ose,
                numero_ose_norm,
                status_ose,
                data_agendamento
            )
            VALUES (%s, %s, %s, %s)
        """, records)

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()
def _build_pending_close_context_from_db():
    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    try:
        cur.execute("""
            SELECT MIN(data_agendamento) AS min_date
            FROM ofs_pending_close_ng
        """)
        row = cur.fetchone() or {}
        min_date = row.get("min_date")
        min_date_str = min_date.strftime("%Y-%m-%d") if min_date else None

        dashboard_data = {
            "line_default": {
                "labels": [],
                "integration": [],
                "pending": [],
                "total_integration": 0,
                "total_pending": 0,
            },
            "activity_type_bar": {
                "labels": [],
                "integration": [],
                "pending": [],
            },
            "daily_by_activity_type": {},
        }

        if not min_date_str:
            return {
                "matched_items": [],
                "total_matches": 0,
                "min_date": None,
                "dashboard_data": dashboard_data,
            }

        # Base completa de integração:
        # tudo que tem erro NG válido no período, independente de existir no XLSX
        cur.execute("""
            SELECT DISTINCT
                e.activity_id,
                e.appt_number,
                e.appt_number_norm,
                e.status,
                e.activity_type,
                e.`date`
            FROM ofs_activities_errors e
            WHERE e.`date` >= %s
              AND (
                    NULLIF(TRIM(e.ng_dispatch_message), '') IS NOT NULL
                    OR NULLIF(TRIM(e.ng_response_message), '') IS NOT NULL
                  )
              AND COALESCE(e.activity_type, '') NOT IN (
                    'LUNCH',
                    'ALM',
                    'MAN_VEIC',
                    'EQP_DUP',
                    'REUNIAO',
                    'VIST_INV_TEC',
                    'MAN',
                    'EM ATIVIDADE B2B'
              )
        """, (min_date_str,))
        integration_items = cur.fetchall()

        # Base de pendentes:
        # subconjunto da integração que também existe no XLSX
        cur.execute("""
            SELECT DISTINCT
                e.activity_id,
                e.appt_number,
                e.appt_number_norm,
                e.status,
                e.activity_type,
                e.`date`,
                ng.numero_ose,
                ng.numero_ose_norm,
                ng.status_ose,
                ng.data_agendamento
            FROM ofs_activities_errors e
            INNER JOIN ofs_pending_close_ng ng
                ON e.appt_number_norm = ng.numero_ose_norm
            WHERE e.`date` >= %s
              AND (
                    NULLIF(TRIM(e.ng_dispatch_message), '') IS NOT NULL
                    OR NULLIF(TRIM(e.ng_response_message), '') IS NOT NULL
                  )
              AND COALESCE(e.activity_type, '') NOT IN (
                    'LUNCH',
                    'ALM',
                    'MAN_VEIC',
                    'EQP_DUP',
                    'REUNIAO',
                    'VIST_INV_TEC',
                    'MAN',
                    'EM ATIVIDADE B2B'
              )
        """, (min_date_str,))
        matched_rows = cur.fetchall()

        # Série diária total de integração
        cur.execute("""
            SELECT
                e.`date` AS ref_date,
                COUNT(DISTINCT e.activity_id) AS qtd
            FROM ofs_activities_errors e
            WHERE e.`date` >= %s
              AND (
                    NULLIF(TRIM(e.ng_dispatch_message), '') IS NOT NULL
                    OR NULLIF(TRIM(e.ng_response_message), '') IS NOT NULL
                  )
              AND COALESCE(e.activity_type, '') NOT IN (
                    'LUNCH',
                    'ALM',
                    'MAN_VEIC',
                    'EQP_DUP',
                    'REUNIAO',
                    'VIST_INV_TEC',
                    'MAN',
                    'EM ATIVIDADE B2B'
              )
            GROUP BY e.`date`
            ORDER BY e.`date` ASC
        """, (min_date_str,))
        integration_rows = cur.fetchall()

    finally:
        cur.close()
        conn.close()

    print("DEBUG integration_items:", len(integration_items), flush=True)
    print("DEBUG matched_rows:", len(matched_rows), flush=True)
    print("DEBUG integration_rows:", len(integration_rows), flush=True)
    print("DEBUG min_date helper:", min_date_str, flush=True)

    matched_items = []
    for item in matched_rows:
        matched_items.append({
            "numero_ose": item.get("numero_ose"),
            "numero_ose_norm": item.get("numero_ose_norm"),
            "status_ose": item.get("status_ose"),
            "data_agendamento": item.get("data_agendamento"),
            "appt_number": item.get("appt_number"),
            "appt_number_norm": item.get("appt_number_norm"),
            "status_painel": item.get("status"),
            "activity_id": item.get("activity_id"),
            "activity_type": item.get("activity_type") or "-",
            "date": item.get("date"),
        })

    integration_by_day = {
        str(item["ref_date"]): int(item["qtd"] or 0)
        for item in integration_rows
    }

    # Pendentes por dia usando a MESMA data da extração (e.date)
    pending_day_oses = {}
    for item in matched_items:
        day = str(item.get("date") or "").strip()
        ose = str(item.get("numero_ose_norm") or item.get("numero_ose") or "").strip()
        if not day or not ose:
            continue
        pending_day_oses.setdefault(day, set()).add(ose)

    pending_by_day = {
        day: len(oses)
        for day, oses in pending_day_oses.items()
    }

    all_dates = sorted(set(integration_by_day.keys()) | set(pending_by_day.keys()))

    # Total por activity_type - integração vem da base completa
    integration_by_activity_type_sets = {}
    for item in integration_items:
        activity_type = str(item.get("activity_type") or "-").strip() or "-"
        activity_id = str(item.get("activity_id") or "").strip()
        if not activity_id:
            continue
        integration_by_activity_type_sets.setdefault(activity_type, set()).add(activity_id)

    integration_by_activity_type = {
        k: len(v)
        for k, v in integration_by_activity_type_sets.items()
    }

    # Total por activity_type - pendente vem apenas da base com match no XLSX
    pending_by_activity_type_sets = {}
    for item in matched_items:
        activity_type = str(item.get("activity_type") or "-").strip() or "-"
        ose = str(item.get("numero_ose_norm") or item.get("numero_ose") or "").strip()
        if not ose:
            continue
        pending_by_activity_type_sets.setdefault(activity_type, set()).add(ose)

    pending_by_activity_type = {
        k: len(v)
        for k, v in pending_by_activity_type_sets.items()
    }

    activity_type_labels = sorted(
        set(integration_by_activity_type.keys()) | set(pending_by_activity_type.keys()),
        key=lambda x: (
            -(integration_by_activity_type.get(x, 0) + pending_by_activity_type.get(x, 0)),
            x
        )
    )

    # Série diária por tipo - integração vem da base completa
    integration_by_day_and_activity = {}
    for item in integration_items:
        activity_type = str(item.get("activity_type") or "-").strip() or "-"
        day = str(item.get("date") or "").strip()
        activity_id = str(item.get("activity_id") or "").strip()

        if not day or not activity_id:
            continue

        integration_by_day_and_activity \
            .setdefault(activity_type, {}) \
            .setdefault(day, set()) \
            .add(activity_id)

    # Série diária por tipo - pendente vem apenas da base com match no XLSX
    pending_by_day_and_activity = {}
    for item in matched_items:
        activity_type = str(item.get("activity_type") or "-").strip() or "-"
        day = str(item.get("date") or "").strip()
        ose = str(item.get("numero_ose_norm") or item.get("numero_ose") or "").strip()

        if not day or not ose:
            continue

        pending_by_day_and_activity \
            .setdefault(activity_type, {}) \
            .setdefault(day, set()) \
            .add(ose)

    daily_by_activity_type = {}
    for activity_type in activity_type_labels:
        integration_map = {
            day: len(ids)
            for day, ids in integration_by_day_and_activity.get(activity_type, {}).items()
        }
        pending_map = {
            day: len(oses)
            for day, oses in pending_by_day_and_activity.get(activity_type, {}).items()
        }

        dates_for_type = sorted(set(integration_map.keys()) | set(pending_map.keys()))

        daily_by_activity_type[activity_type] = {
            "labels": dates_for_type,
            "integration": [integration_map.get(day, 0) for day in dates_for_type],
            "pending": [pending_map.get(day, 0) for day in dates_for_type],
            "total_integration": sum(integration_map.values()),
            "total_pending": sum(pending_map.values()),
        }

    dashboard_data = {
        "line_default": {
            "labels": all_dates,
            "integration": [integration_by_day.get(day, 0) for day in all_dates],
            "pending": [pending_by_day.get(day, 0) for day in all_dates],
            "total_integration": sum(integration_by_day.values()),
            "total_pending": sum(pending_by_day.values()),
        },
        "activity_type_bar": {
            "labels": activity_type_labels,
            "integration": [integration_by_activity_type.get(label, 0) for label in activity_type_labels],
            "pending": [pending_by_activity_type.get(label, 0) for label in activity_type_labels],
        },
        "daily_by_activity_type": daily_by_activity_type,
    }

    return {
        "matched_items": matched_items,
        "total_matches": len(matched_items),
        "min_date": min_date_str,
        "dashboard_data": dashboard_data,
    }
def init_app(app):
    @app.route("/ofs/activities-errors", methods=["GET"])
    @login_required
    @perm_required("ofs.activities_errors")
    def ofs_activities_errors():
        today = datetime.now().strftime("%Y-%m-%d")
        date_from = (request.args.get("dateFrom") or today).strip()
        date_to = (request.args.get("dateTo") or today).strip()
        resources = (request.args.get("resources") or "02").strip()

        per_page = 50
        page = request.args.get("page", default=1, type=int)
        offset = (page - 1) * per_page

        conn = get_connection()
        cur = conn.cursor(dictionary=True)

        cur.execute("""
            SELECT COUNT(*) AS total
            FROM ofs_activities_errors
            WHERE `date` BETWEEN %s AND %s
        """, (date_from, date_to))
        total = (cur.fetchone() or {}).get("total", 0)
        total_pages = max(1, (total + per_page - 1) // per_page)

        cur.execute("""
        SELECT
            activity_id AS activityId,
            `date` AS date,
            city,
            activity_type AS activityType,
            appt_number AS apptNumber,
            status,
            xa_sap_crt AS XA_SAP_CRT,
            xa_sap_crt_ldg AS XA_SAP_CRT_LDG,
            xa_res_api_ng_response AS XA_RES_API_NG_RESPONSE,
            ng_dispatch_message AS ngDispatchMessage,
            ng_response_message AS ngResponseMessage,
            sap_error_raw_extracted AS sapErrorRawExtracted,
            sap_response_message AS sapResponseMessage,
            sap_error_category AS sapErrorCategory,
            CASE
                WHEN COALESCE(TRIM(xa_sap_crt), '') = '1'
                AND (
                        TRIM(COALESCE(xa_res_api_ng_response, '')) LIKE '<![CDATA[%'
                    OR TRIM(COALESCE(xa_api_ng_dispatch, ''))     LIKE '<![CDATA[%'
                )
                    THEN 'Erro SAP/NG'

                WHEN COALESCE(TRIM(xa_sap_crt), '') = '1'
                    THEN 'Erro SAP'

                WHEN COALESCE(TRIM(xa_sap_crt), '') <> '1'
                AND (
                        TRIM(COALESCE(xa_res_api_ng_response, '')) LIKE '<![CDATA[%'
                    OR TRIM(COALESCE(xa_api_ng_dispatch, ''))     LIKE '<![CDATA[%'
                )
                    THEN 'Erro NG'

                ELSE '-'
            END AS erro_tipo,
            last_seen_at
        FROM ofs_activities_errors
        WHERE `date` BETWEEN %s AND %s
        ORDER BY `date` DESC, last_seen_at DESC
        LIMIT %s OFFSET %s
        """, (date_from, date_to, per_page, offset))

        items = cur.fetchall()

        cur.close()
        conn.close()

        return render_template(
            "ofs_activities_errors/ofs_activities_errors.html",
            items=items,
            total=total,
            page=page,
            total_pages=total_pages,
            date_from=date_from,
            date_to=date_to,
            resources=resources
        )

    @app.route("/ofs/activities-errors/<activity_id>", methods=["GET"])
    @login_required
    @perm_required("ofs.activities_errors")
    def ofs_activities_errors_get(activity_id):
        activity_id = str(activity_id or "").strip()
        if not activity_id:
            return jsonify({"ok": False, "error": "activityId inválido"}), 400

        conn = get_connection()
        cur = conn.cursor(dictionary=True)

        cur.execute("""
            SELECT
                activity_id AS activityId,
                xa_api_ng_dispatch AS XA_API_NG_DISPATCH,
                xa_res_api_ng_response AS XA_RES_API_NG_RESPONSE,
                xa_sap_crt_ldg AS XA_SAP_CRT_LDG,
                xa_sap_crt AS XA_SAP_CRT,
                ng_dispatch_message AS ngDispatchMessage,
                ng_response_message AS ngResponseMessage,
                sap_error_raw_extracted AS sapErrorRawExtracted,
                sap_response_message AS sapResponseMessage,
                sap_error_category AS sapErrorCategory
            FROM ofs_activities_errors
            WHERE activity_id = %s
            LIMIT 1
        """, (activity_id,))

        row = cur.fetchone()
        cur.close()
        conn.close()

        if not row:
            return jsonify({"ok": False, "error": "Não encontrado"}), 404

        return jsonify({"ok": True, "item": row}), 200

    @app.route("/ofs/activities-errors/dashboard", methods=["GET"])
    @login_required
    @perm_required("ofs.activities_errors")
    def ofs_activities_errors_dashboard():
        today = datetime.now().strftime("%Y-%m-%d")
        date_from = (request.args.get("dateFrom") or today).strip()
        date_to = (request.args.get("dateTo") or today).strip()
        resources = (request.args.get("resources") or "02").strip()

        conn = get_connection()
        cur = conn.cursor(dictionary=True)

        # KPI: total linhas no período
        cur.execute("""
            SELECT COUNT(*) AS total
            FROM ofs_activities_errors
            WHERE `date` BETWEEN %s AND %s
        """, (date_from, date_to))
        total = (cur.fetchone() or {}).get("total", 0)

        # KPI: total com erro NG (dispatch OR response)
        cur.execute("""
            SELECT COUNT(*) AS total_ng
            FROM ofs_activities_errors
            WHERE `date` BETWEEN %s AND %s
            AND (ng_dispatch_message IS NOT NULL OR ng_response_message IS NOT NULL)
        """, (date_from, date_to))
        total_ng = (cur.fetchone() or {}).get("total_ng", 0)

        # Top mensagens (considerando dispatch e response juntos)
        cur.execute("""
            SELECT msg, COUNT(*) AS qtd
            FROM (
                SELECT ng_dispatch_message AS msg
                FROM ofs_activities_errors
                WHERE `date` BETWEEN %s AND %s AND ng_dispatch_message IS NOT NULL
                UNION ALL
                SELECT ng_response_message AS msg
                FROM ofs_activities_errors
                WHERE `date` BETWEEN %s AND %s AND ng_response_message IS NOT NULL
            ) x
            GROUP BY msg
            ORDER BY qtd DESC
            LIMIT 15
        """, (date_from, date_to, date_from, date_to))
        top_messages = cur.fetchall()

        # Erros por dia
        cur.execute("""
            SELECT `date`, COUNT(*) AS qtd
            FROM ofs_activities_errors
            WHERE `date` BETWEEN %s AND %s
            AND (ng_dispatch_message IS NOT NULL OR ng_response_message IS NOT NULL)
            GROUP BY `date`
            ORDER BY `date` DESC
            LIMIT 31
        """, (date_from, date_to))
        by_day = cur.fetchall()

        # Erros por activityType
        cur.execute("""
            SELECT COALESCE(activity_type,'-') AS activityType, COUNT(*) AS qtd
            FROM ofs_activities_errors
            WHERE `date` BETWEEN %s AND %s
            AND (ng_dispatch_message IS NOT NULL OR ng_response_message IS NOT NULL)
            GROUP BY COALESCE(activity_type,'-')
            ORDER BY qtd DESC
            LIMIT 20
        """, (date_from, date_to))
        by_type = cur.fetchall()
        cur.execute("""
        SELECT
            COALESCE(cfg.responsavel, 'Não mapeado') AS responsavel,
            COUNT(*) AS qtd
        FROM (

            SELECT
                'NG_DISPATCH' AS origem,
                TRIM(ng_dispatch_message) AS error_message
            FROM ofs_activities_errors
            WHERE `date` BETWEEN %s AND %s
            AND ng_dispatch_message IS NOT NULL
            AND NULLIF(TRIM(ng_dispatch_message),'') IS NOT NULL

            UNION ALL

            SELECT
                'NG_RESPONSE' AS origem,
                TRIM(ng_response_message) AS error_message
            FROM ofs_activities_errors
            WHERE `date` BETWEEN %s AND %s
            AND ng_response_message IS NOT NULL
            AND NULLIF(TRIM(ng_response_message),'') IS NOT NULL

        ) base

        LEFT JOIN ofs_error_owner_config cfg
            ON cfg.origem = base.origem
        AND TRIM(cfg.error_message) = base.error_message
        AND cfg.ativo = 1

        GROUP BY COALESCE(cfg.responsavel, 'Não mapeado')
        ORDER BY qtd DESC
        """, (date_from, date_to, date_from, date_to))

        by_owner = cur.fetchall()
        cur.close()
        conn.close()

        return render_template(
            "ofs_activities_errors/ofs_activities_errors_dashboard.html",
            total=total,
            total_ng=total_ng,
            top_messages=top_messages,
            by_day=by_day,
            by_type=by_type,
            date_from=date_from,
            date_to=date_to,
            by_owner=by_owner,
            resources=resources
        )

    @app.route("/ofs/activities-errors/dashboard/data", methods=["GET"])
    @login_required
    @perm_required("ofs.activities_errors")
    def ofs_activities_errors_dashboard_data():
        today = datetime.now().strftime("%Y-%m-%d")
        date_from = (request.args.get("dateFrom") or today).strip()
        date_to = (request.args.get("dateTo") or today).strip()
        resources = (request.args.get("resources") or "02").strip()

        conn = get_connection()
        cur = conn.cursor(dictionary=True)

        cur.execute("""
            SELECT COUNT(*) AS total
            FROM ofs_activities_errors
            WHERE `date` BETWEEN %s AND %s
        """, (date_from, date_to))
        total = (cur.fetchone() or {}).get("total", 0)

        cur.execute("""
            SELECT COUNT(*) AS total_ng
            FROM ofs_activities_errors
            WHERE `date` BETWEEN %s AND %s
            AND (ng_dispatch_message IS NOT NULL OR ng_response_message IS NOT NULL)
        """, (date_from, date_to))
        total_ng = (cur.fetchone() or {}).get("total_ng", 0)

        cur.execute("""
            SELECT msg, COUNT(*) AS qtd
            FROM (
                SELECT ng_dispatch_message AS msg
                FROM ofs_activities_errors
                WHERE `date` BETWEEN %s AND %s AND ng_dispatch_message IS NOT NULL
                UNION ALL
                SELECT ng_response_message AS msg
                FROM ofs_activities_errors
                WHERE `date` BETWEEN %s AND %s AND ng_response_message IS NOT NULL
            ) x
            GROUP BY msg
            ORDER BY qtd DESC
            LIMIT 15
        """, (date_from, date_to, date_from, date_to))
        top_messages = cur.fetchall()

        cur.execute("""
            SELECT `date`, COUNT(*) AS qtd
            FROM ofs_activities_errors
            WHERE `date` BETWEEN %s AND %s
            AND (ng_dispatch_message IS NOT NULL OR ng_response_message IS NOT NULL)
            GROUP BY `date`
            ORDER BY `date` ASC
            LIMIT 31
        """, (date_from, date_to))
        by_day = cur.fetchall()
        cur.execute("""
            SELECT
                `date`,
                COALESCE(e.activity_type, '-') AS activityType,
                COALESCE(c.descricao, COALESCE(e.activity_type, '-')) AS activityTypeLabel,
                COUNT(*) AS qtd
            FROM ofs_activities_errors e
            LEFT JOIN ofs_activity_type_config c
                ON c.activity_type = e.activity_type
            AND c.ativo = 1
            WHERE e.`date` BETWEEN %s AND %s
            AND (
                    e.ng_dispatch_message IS NOT NULL
                    OR e.ng_response_message IS NOT NULL
                )
            AND (
                    c.mostrar_dashboard = 1
                    OR c.id IS NULL
                )
            GROUP BY
                `date`,
                COALESCE(e.activity_type, '-'),
                COALESCE(c.descricao, COALESCE(e.activity_type, '-'))
            ORDER BY `date` ASC
        """, (date_from, date_to))
        by_day_type = cur.fetchall()
        cur.execute("""
            SELECT
                COALESCE(e.activity_type, '-') AS activityType,
                COALESCE(c.descricao, COALESCE(e.activity_type, '-')) AS activityTypeLabel,
                COUNT(*) AS qtd
            FROM ofs_activities_errors e
            LEFT JOIN ofs_activity_type_config c
                ON c.activity_type = e.activity_type
            AND c.ativo = 1
            WHERE e.`date` BETWEEN %s AND %s
            AND (
                    e.ng_dispatch_message IS NOT NULL
                    OR e.ng_response_message IS NOT NULL
                )
            AND (
                    c.mostrar_dashboard = 1
                    OR c.id IS NULL
                )
            GROUP BY
                COALESCE(e.activity_type, '-'),
                COALESCE(c.descricao, COALESCE(e.activity_type, '-'))
            ORDER BY qtd DESC
            LIMIT 20
        """, (date_from, date_to))
        by_type = cur.fetchall()
        cur.execute("""
            SELECT
                sap_response_message AS msg,
                COUNT(*) AS qtd
            FROM ofs_activities_errors
            WHERE `date` BETWEEN %s AND %s
            AND sap_response_message IS NOT NULL
            AND NULLIF(TRIM(sap_response_message), '') IS NOT NULL
            GROUP BY sap_response_message
            ORDER BY qtd DESC
            LIMIT 15
        """, (date_from, date_to))
        top_sap_messages = cur.fetchall()

        cur.execute("""
            SELECT
                COALESCE(e.activity_type, '-') AS activityType,
                COALESCE(c.descricao, COALESCE(e.activity_type, '-')) AS activityTypeLabel,
                COUNT(*) AS qtd
            FROM ofs_activities_errors e
            LEFT JOIN ofs_activity_type_config c
                ON c.activity_type = e.activity_type
                AND c.ativo = 1
            WHERE e.`date` BETWEEN %s AND %s
                AND COALESCE(TRIM(e.xa_sap_crt), '') = '1'
            GROUP BY
                COALESCE(e.activity_type, '-'),
                COALESCE(c.descricao, COALESCE(e.activity_type, '-'))
            ORDER BY qtd DESC
            LIMIT 20
        """, (date_from, date_to))
        sap_by_activity_type = cur.fetchall()
        cur.close()
        conn.close()

        return jsonify({
            "ok": True,
            "total": total,
            "total_ng": total_ng,
            "top_messages": top_messages,
            "by_day": by_day,
            "by_day_type": by_day_type,
            "by_type": by_type,
            "date_from": date_from,
            "date_to": date_to,
            "top_sap_messages": top_sap_messages,
            "sap_by_activity_type": sap_by_activity_type,
            "debug_marker": "ROTA_NOVA_SAP",
            "resources": resources
        }), 200
    @app.route("/ofs/activities-errors/export/xlsx", methods=["GET"])
    @login_required
    @perm_required("ofs.activities_errors")
    def ofs_activities_errors_export_xlsx():
        today = datetime.now().strftime("%Y-%m-%d")
        date_from = (request.args.get("dateFrom") or today).strip()
        date_to = (request.args.get("dateTo") or today).strip()
        resources = (request.args.get("resources") or "02").strip()

        conn = get_connection()
        cur = conn.cursor(dictionary=True)

        cur.execute("""
            SELECT
                activity_id,
                city,
                activity_type,
                appt_number,
                status,
                ng_dispatch_message,
                ng_response_message,
                sap_error_raw_extracted,
                sap_response_message,
                sap_error_category,
                xa_sap_crt_ldg,
                `date`
            FROM ofs_activities_errors
            WHERE `date` BETWEEN %s AND %s
            AND activity_type IN (
                    'INS',
                    'SUP_QUA',
                    'SUP_REP',
                    'SOL_SER',
                    'INS_DEV',
                    'SUP',
                    'MIG_PLA',
                    'QUA',
                    'MIG_TEC'
            )
            AND (
                    NULLIF(TRIM(ng_dispatch_message), '') IS NOT NULL
                    OR NULLIF(TRIM(ng_response_message), '') IS NOT NULL
                    OR NULLIF(TRIM(sap_response_message), '') IS NOT NULL
                    OR NULLIF(TRIM(xa_sap_crt_ldg), '') IS NOT NULL
            )
            ORDER BY `date` DESC, activity_type, city, appt_number
        """, (date_from, date_to))

        rows = cur.fetchall()
        cur.close()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Atividades com erro"

        headers = [
            "activity_id",
            "city",
            "activity_type",
            "appt_number",
            "status",
            "ng_dispatch_message",
            "ng_response_message",
            "sap_error_raw_extracted",
            "sap_response_message",
            "sap_error_category",
            "xa_sap_crt_ldg",
            "date",
        ]
        ws.append(headers)

        for row in rows:
            ws.append([
                row.get("activity_id"),
                row.get("city"),
                row.get("activity_type"),
                row.get("appt_number"),
                row.get("status"),
                row.get("ng_dispatch_message"),
                row.get("ng_response_message"),
                row.get("sap_error_raw_extracted"),
                row.get("sap_response_message"),
                row.get("sap_error_category"),
                row.get("xa_sap_crt_ldg"),
                row.get("date"),
            ])
        # Ajuste simples de largura
        for col_idx, col_name in enumerate(headers, start=1):
            max_len = len(col_name)
            for row_idx in range(2, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"ofs_activities_errors_{date_from}_a_{date_to}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    @app.route("/ofs/activities-errors/export/top-messages", methods=["GET"])
    @login_required
    @perm_required("ofs.activities_errors")
    def ofs_activities_errors_export_top_messages():

        date_from = (request.args.get("dateFrom") or "").strip()
        date_to = (request.args.get("dateTo") or "").strip()

        conn = get_connection()
        cur = conn.cursor(dictionary=True)

        cur.execute("""
            SELECT
                activity_id,
                `date`,
                city,
                activity_type,
                appt_number,
                status,
                ng_dispatch_message,
                ng_response_message
            FROM ofs_activities_errors
            WHERE `date` BETWEEN %s AND %s
            AND (
                    ng_dispatch_message IS NOT NULL
                    OR ng_response_message IS NOT NULL
                )
            ORDER BY `date` DESC
        """, (date_from, date_to))

        rows = cur.fetchall()

        cur.close()
        conn.close()

        output = BytesIO()

        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet("erros_ng")

        headers = [
            "activity_id",
            "date",
            "city",
            "activity_type",
            "appt_number",
            "status",
            "ng_dispatch_message",
            "ng_response_message"
        ]

        # escreve header
        for col, header in enumerate(headers):
            worksheet.write(0, col, header)

        # escreve dados
        for row_idx, row in enumerate(rows, start=1):
            for col_idx, header in enumerate(headers):
                worksheet.write(row_idx, col_idx, row.get(header))

        workbook.close()

        output.seek(0)

        filename = f"ofs_erros_ng_{date_from}_a_{date_to}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    @app.route("/ofs/activities-errors/export-top-sap-messages", methods=["GET"])
    @login_required
    @perm_required("ofs.activities_errors")
    def ofs_activities_errors_export_top_sap_messages():
        today = datetime.now().strftime("%Y-%m-%d")
        date_from = (request.args.get("dateFrom") or today).strip()
        date_to = (request.args.get("dateTo") or today).strip()

        conn = get_connection()
        cur = conn.cursor(dictionary=True)

        cur.execute("""
            SELECT
                activity_id,
                `date`,
                city,
                activity_type,
                appt_number,
                status,
                sap_response_message
            FROM ofs_activities_errors
            WHERE `date` BETWEEN %s AND %s
            AND sap_response_message IS NOT NULL
            AND NULLIF(TRIM(sap_response_message), '') IS NOT NULL
            ORDER BY `date` DESC, last_seen_at DESC
        """, (date_from, date_to))

        rows = cur.fetchall()
        cur.close()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Top erros SAP"

        ws.append([
            "activity_id",
            "date",
            "city",
            "activity_type",
            "appt_number",
            "status",
            "sap_response_message"
        ])

        for row in rows:
            ws.append([
                row.get("activity_id"),
                row.get("date"),
                row.get("city"),
                row.get("activity_type"),
                row.get("appt_number"),
                row.get("status"),
                row.get("sap_response_message"),
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"top_erros_sap_{date_from}_a_{date_to}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    @app.route("/ofs/activities-errors/importar/start", methods=["POST"])
    @login_required
    @perm_required("ofs.activities_errors")
    def ofs_activities_errors_importar_start():
        today = datetime.now().strftime("%Y-%m-%d")
        date_from = (request.form.get("dateFrom") or today).strip()
        date_to = (request.form.get("dateTo") or today).strip()
        resources = (request.form.get("resources") or "02").strip()

        try:
            validate_max_range_7_days(date_from, date_to)
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 400

        actor = current_actor()
        username = actor.get("username")

        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO ofs_import_jobs (module, status, progress, message, created_by)
            VALUES ('ofs.activities_errors', 'running', 0, 'Iniciando...', %s)
        """, (username,))
        job_id = cur.lastrowid
        conn.commit()
        cur.close()
        conn.close()

        t = threading.Thread(
            target=run_import_job,
            args=(job_id, date_from, date_to, resources, username),

        )
        t.start()

        return jsonify({"ok": True, "jobId": job_id}), 200

    @app.route("/ofs/activities-errors/importar/status/<int:job_id>", methods=["GET"])
    @login_required
    @perm_required("ofs.activities_errors")
    def ofs_activities_errors_importar_status(job_id):
        conn = get_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT id, status, progress, message, created_at, updated_at
            FROM ofs_import_jobs
            WHERE id=%s AND module='ofs.activities_errors'
            LIMIT 1
        """, (job_id,))
        row = cur.fetchone()
        cur.close()
        conn.close()

        if not row:
            return jsonify({"ok": False, "error": "Job não encontrado"}), 404

        return jsonify({"ok": True, "job": row}), 200
    @app.route("/ofs/activities-errors/importar/cancel/<int:job_id>", methods=["POST"])
    @login_required
    @perm_required("ofs.activities_errors")
    def ofs_activities_errors_importar_cancel(job_id):
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            UPDATE ofs_import_jobs
            SET cancel_requested=1, message='Cancelamento solicitado...'
            WHERE id=%s AND module='ofs.activities_errors' AND status='running'
        """, (job_id,))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"ok": True}), 200

    @app.route("/ofs/config/error-owners", methods=["GET"])
    @login_required
    @perm_required("ofs.activities_errors")
    def ofs_error_owners_config():
        q = str(request.args.get("q") or "").strip()
        status = str(request.args.get("status") or "all").strip().lower()
        page = request.args.get("page", default=1, type=int)
        per_page = 50
        offset = (page - 1) * per_page

        conn = get_connection()
        cur = conn.cursor(dictionary=True)

        summary_where_filters = []
        summary_params = []

        if q:
            like = f"%{q}%"
            summary_where_filters.append("(base.error_message LIKE %s OR base.origem LIKE %s)")
            summary_params.extend([like, like])

        summary_where_sql = ""
        if summary_where_filters:
            summary_where_sql = "WHERE " + " AND ".join(summary_where_filters)

        summary_base_sql = f"""
            FROM (
                SELECT
                    'NG_DISPATCH' AS origem,
                    TRIM(ng_dispatch_message) AS error_message,
                    COUNT(*) AS qtd
                FROM ofs_activities_errors
                WHERE ng_dispatch_message IS NOT NULL
                  AND NULLIF(TRIM(ng_dispatch_message), '') IS NOT NULL
                GROUP BY TRIM(ng_dispatch_message)

                UNION ALL

                SELECT
                    'NG_RESPONSE' AS origem,
                    TRIM(ng_response_message) AS error_message,
                    COUNT(*) AS qtd
                FROM ofs_activities_errors
                WHERE ng_response_message IS NOT NULL
                  AND NULLIF(TRIM(ng_response_message), '') IS NOT NULL
                GROUP BY TRIM(ng_response_message)
            ) base
            LEFT JOIN ofs_error_owner_config cfg
                ON cfg.origem = base.origem
               AND TRIM(cfg.error_message) = base.error_message
               AND cfg.ativo = 1
            {summary_where_sql}
        """

        cur.execute(f"""
            SELECT COUNT(*) AS total
            {summary_base_sql}
        """, tuple(summary_params))
        total = (cur.fetchone() or {}).get("total", 0)

        cur.execute(f"""
            SELECT COUNT(*) AS total_configured
            {summary_base_sql}
            {"AND" if summary_where_sql else "WHERE"} cfg.id IS NOT NULL
        """, tuple(summary_params))
        total_configured = (cur.fetchone() or {}).get("total_configured", 0)

        cur.execute(f"""
            SELECT COUNT(*) AS total_pending
            {summary_base_sql}
            {"AND" if summary_where_sql else "WHERE"} cfg.id IS NULL
        """, tuple(summary_params))
        total_pending = (cur.fetchone() or {}).get("total_pending", 0)

        where_filters = []
        params = []

        if q:
            like = f"%{q}%"
            where_filters.append("(base.error_message LIKE %s OR base.origem LIKE %s)")
            params.extend([like, like])

        if status == "configured":
            where_filters.append("cfg.id IS NOT NULL")
        elif status == "pending":
            where_filters.append("cfg.id IS NULL")

        where_sql = ""
        if where_filters:
            where_sql = "WHERE " + " AND ".join(where_filters)

        base_sql = f"""
            FROM (
                SELECT
                    'NG_DISPATCH' AS origem,
                    TRIM(ng_dispatch_message) AS error_message,
                    COUNT(*) AS qtd
                FROM ofs_activities_errors
                WHERE ng_dispatch_message IS NOT NULL
                  AND NULLIF(TRIM(ng_dispatch_message), '') IS NOT NULL
                GROUP BY TRIM(ng_dispatch_message)

                UNION ALL

                SELECT
                    'NG_RESPONSE' AS origem,
                    TRIM(ng_response_message) AS error_message,
                    COUNT(*) AS qtd
                FROM ofs_activities_errors
                WHERE ng_response_message IS NOT NULL
                  AND NULLIF(TRIM(ng_response_message), '') IS NOT NULL
                GROUP BY TRIM(ng_response_message)
            ) base
            LEFT JOIN ofs_error_owner_config cfg
                ON cfg.origem = base.origem
               AND TRIM(cfg.error_message) = base.error_message
               AND cfg.ativo = 1
            {where_sql}
        """

        cur.execute(f"""
            SELECT COUNT(*) AS total_filtered
            {base_sql}
        """, tuple(params))
        total_filtered = (cur.fetchone() or {}).get("total_filtered", 0)
        total_pages = max(1, (total_filtered + per_page - 1) // per_page)

        cur.execute(f"""
            SELECT
                base.origem,
                base.error_message,
                cfg.responsavel,
                CASE
                    WHEN cfg.id IS NULL THEN 0
                    ELSE 1
                END AS configurado,
                base.qtd
            {base_sql}
            ORDER BY
                base.qtd DESC,
                base.origem ASC
            LIMIT %s OFFSET %s
        """, tuple(params + [per_page, offset]))

        items = cur.fetchall()

        cur.close()
        conn.close()

        return render_template(
            "ofs_error_owner_config/ofs_error_owner_config.html",
            items=items,
            q=q,
            status=status,
            page=page,
            total=total,
            total_pages=total_pages,
            total_configured=total_configured,
            total_pending=total_pending,
            total_filtered=total_filtered
        )

    @app.route("/ofs/config/error-owners/save", methods=["POST"])
    @login_required
    @perm_required("ofs.activities_errors")
    def ofs_error_owners_config_save():
        origem = str(request.form.get("origem") or "").strip()
        error_message = request.form.get("error_message") or ""
        responsavel = str(request.form.get("responsavel") or "").strip()

        allowed = {"WFM", "NG", "Desconsiderar"}

        if not origem:
            flash("Origem obrigatória.", "error")
            return redirect(url_for("ofs_error_owners_config"))

        if not error_message:
            flash("Mensagem obrigatória.", "error")
            return redirect(url_for("ofs_error_owners_config"))

        if responsavel not in allowed:
            flash("Responsável inválido.", "error")
            return redirect(url_for("ofs_error_owners_config"))

        conn = get_connection()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO ofs_error_owner_config (origem, error_message, responsavel, ativo)
            VALUES (%s, %s, %s, 1)
            ON DUPLICATE KEY UPDATE
                responsavel = VALUES(responsavel),
                ativo = 1,
                atualizado_em = CURRENT_TIMESTAMP
        """, (origem, error_message, responsavel))

        conn.commit()
        cur.close()
        conn.close()

        flash("Responsável salvo com sucesso.", "success")
        return redirect(url_for("ofs_error_owners_config"))

    @app.route("/ofs/config/activity-types", methods=["GET"])
    @login_required
    @perm_required("ofs.activities_errors")
    def ofs_activity_types_config():
        conn = get_connection()
        cur = conn.cursor(dictionary=True)

        cur.execute("""
            SELECT
                base.activity_type,
                cfg.descricao,
                CASE
                    WHEN cfg.id IS NULL THEN 0
                    ELSE 1
                END AS configurado,
                base.qtd
            FROM (
                SELECT
                    COALESCE(activity_type, '-') AS activity_type,
                    COUNT(*) AS qtd
                FROM ofs_activities_errors
                GROUP BY COALESCE(activity_type, '-')
            ) base
            LEFT JOIN ofs_activity_type_config cfg
                ON cfg.activity_type = base.activity_type
               AND cfg.ativo = 1
            ORDER BY
                configurado ASC,
                base.qtd DESC,
                base.activity_type ASC
        """)
        items = cur.fetchall()

        cur.close()
        conn.close()

        return render_template(
            "ofs_activity_type_config/ofs_activity_type_config.html",
            items=items
        )

    @app.route("/ofs/config/activity-types/save", methods=["POST"])
    @login_required
    @perm_required("ofs.activities_errors")
    def ofs_activity_types_config_save():
        activity_type = str(request.form.get("activity_type") or "").strip()
        descricao = str(request.form.get("descricao") or "").strip()

        if not activity_type:
            flash("activity_type obrigatório.", "error")
            return redirect(url_for("ofs_activity_types_config"))

        if activity_type == "-":
            flash("Não é permitido configurar o tipo '-'.", "error")
            return redirect(url_for("ofs_activity_types_config"))

        if not descricao:
            flash("Descrição obrigatória.", "error")
            return redirect(url_for("ofs_activity_types_config"))

        conn = get_connection()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO ofs_activity_type_config (activity_type, descricao, ativo)
            VALUES (%s, %s, 1)
            ON DUPLICATE KEY UPDATE
                descricao = VALUES(descricao),
                ativo = 1,
                atualizado_em = CURRENT_TIMESTAMP
        """, (activity_type, descricao))

        conn.commit()
        cur.close()
        conn.close()

        flash(f"Tipo '{activity_type}' salvo com sucesso.", "success")
        return redirect(url_for("ofs_activity_types_config"))

    @app.route("/ofs/pending-close", methods=["GET", "POST"])
    @login_required
    @perm_required("ofs.activities_errors")
    def ofs_pending_close():
        dashboard_data = {
            "line_default": {
                "labels": [],
                "integration": [],
                "pending": [],
                "total_integration": 0,
                "total_pending": 0,
            },
            "activity_type_bar": {
                "labels": [],
                "integration": [],
                "pending": [],
            },
            "daily_by_activity_type": {},
        }

        matched_items = []
        total_matches = 0
        min_date = None
        uploaded_name = session.get("ofs_pending_close_original_name")

        if request.method == "POST":
            file = request.files.get("file")

            if not file or not file.filename:
                flash("Selecione um arquivo XLSX para importar.", "error")
                return render_template(
                    "ofs_activities_errors/ofs_pending_close.html",
                    matched_items=[],
                    total_matches=0,
                    min_date=None,
                    uploaded_name=uploaded_name,
                    dashboard_data=dashboard_data,
                )

            original_name = (file.filename or "").strip()

            if not original_name.lower().endswith(".xlsx"):
                flash("Arquivo inválido. Envie um arquivo .xlsx.", "error")
                return render_template(
                    "ofs_activities_errors/ofs_pending_close.html",
                    matched_items=[],
                    total_matches=0,
                    min_date=None,
                    uploaded_name=uploaded_name,
                    dashboard_data=dashboard_data,
                )

            try:
                _import_pending_close_xlsx_to_db(file)
                session["ofs_pending_close_original_name"] = original_name
                uploaded_name = original_name
                flash("Arquivo processado com sucesso.", "success")
            except Exception as e:
                flash(f"Erro ao processar arquivo: {str(e)}", "error")

        try:
            context = _build_pending_close_context_from_db()
            matched_items = (context["matched_items"] or [])[:50]
            total_matches = context["total_matches"] or 0
            min_date = context["min_date"]
            dashboard_data = context["dashboard_data"]
            print("DEBUG total_matches:", total_matches, flush=True)
            print("DEBUG matched_items_len:", len(matched_items), flush=True)
            print("DEBUG min_date:", min_date, flush=True)
        except Exception as e:
            import traceback
            print(traceback.format_exc(), flush=True)
            flash(f"Erro ao montar visão da tela: {str(e)}", "error")

        return render_template(
            "ofs_activities_errors/ofs_pending_close.html",
            matched_items=matched_items,
            total_matches=total_matches,
            min_date=min_date,
            uploaded_name=uploaded_name,
            dashboard_data=dashboard_data,
        )


    @app.route("/ofs/pending-close/export/xlsx", methods=["GET"])
    @login_required
    @perm_required("ofs.activities_errors")
    def ofs_pending_close_export_xlsx():
        try:
            context = _build_pending_close_context_from_db()
            rows = context["matched_items"] or []

            if not rows:
                flash("Não há dados processados para exportar.", "error")
                return redirect(url_for("ofs_pending_close"))

            wb = Workbook()
            ws = wb.active
            ws.title = "OSES pendentes fechamento"

            headers = [
                "NUMERO_OSE",
                "STATUS_OSE",
                "DATA_AGENDAMENTO",
                "activity_type",
                "appt_number_painel",
                "status_painel",
                "activity_id",
                "date_painel",
            ]
            ws.append(headers)

            for row in rows:
                ws.append([
                    row.get("numero_ose"),
                    row.get("status_ose"),
                    row.get("data_agendamento"),
                    row.get("activity_type"),
                    row.get("appt_number"),
                    row.get("status_painel"),
                    row.get("activity_id"),
                    row.get("date"),
                ])

            for col_idx, col_name in enumerate(headers, start=1):
                max_len = len(col_name)
                for row_idx in range(2, ws.max_row + 1):
                    value = ws.cell(row=row_idx, column=col_idx).value
                    if value is not None:
                        max_len = max(max_len, len(str(value)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            filename = "oses_pendentes_fechamento.xlsx"

            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            flash(f"Erro ao exportar arquivo: {str(e)}", "error")
            return redirect(url_for("ofs_pending_close"))