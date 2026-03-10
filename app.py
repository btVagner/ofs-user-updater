import traceback

from flask import Flask, render_template, request, send_file, redirect, url_for, flash, session, jsonify, send_file
import requests
from datetime import datetime, timedelta
import csv
from dotenv import load_dotenv
import xlsxwriter
from database.connection import get_connection
import bcrypt
import os
from functools import wraps
from io import BytesIO, StringIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import json
import time
from database.audit import audit_log
from werkzeug.middleware.proxy_fix import ProxyFix
from urllib.parse import urlencode
import re
import unicodedata
import threading

load_dotenv()

from ofs.cleanup import find_stale_users, execute_cleanup
from ofs.client import OFSClient

app = Flask(__name__)
APP_ROOT = os.getenv("APP_ROOT", "")  # em produção defina "/ofs", em dev deixe vazio
if APP_ROOT:
    app.config["APPLICATION_ROOT"] = APP_ROOT
app.secret_key = os.getenv("FLASK_SECRET_KEY", "minha_chave_secreta")
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)



@app.before_request
def atualizar_online():
    # Só registra se o usuário estiver logado
    if session.get("usuario_logado"):
        try:
            registrar_atividade_usuario()
        except Exception as e:
            # Em produção você pode logar isso
            print(f"[WARN] Falha ao registrar atividade do usuário: {e}")
@app.context_processor
def inject_online_count():
    """Disponibiliza 'usuarios_online_count' em todos os templates."""
    count = 0
    try:
        conn = get_connection()
        cur = conn.cursor()

        limite = datetime.now() - timedelta(minutes=5)  # janela de 5 minutos
        cur.execute(
            "SELECT COUNT(*) FROM usuarios_online WHERE last_seen >= %s",
            (limite,),
        )
        row = cur.fetchone()
        if row:
            count = row[0]

        cur.close()
        conn.close()
    except Exception as e:
        print(f"[WARN] Falha ao obter usuarios_online_count: {e}")

    return dict(usuarios_online_count=count)
def current_actor():
    return {
        "id": session.get("usuario_id"),
        "username": session.get("usuario_logado"),
        "nome": session.get("nome_usuario"),
        "tipo_id": session.get("tipo_id"),
    }

# Helpers de sessão/acesso
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "usuario_logado" not in session:
            flash("Faça login para acessar esta página.", "danger")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function


def _carregar_permissoes_por_perfil(perfil_id: int):
    """Carrega permissões do banco para um perfil (id do perfil)."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT p.recurso
        FROM perfil_permissao pp
        JOIN permissoes p ON p.id = pp.permissao_id
        WHERE pp.perfil_id = %s
    """, (perfil_id,))
    perms = [row[0] for row in cur.fetchall()]
    cur.close()
    conn.close()
    return perms


def has_perm(recurso: str) -> bool:
    # Fallback seguro: se não carregou permissões ainda, admin (tipo_id=1) enxerga tudo
    perms = session.get("permissoes")
    if perms is None:
        return session.get("tipo_id") == 1
    return recurso in perms


def perm_required(*recursos):
    """
    Uso: @perm_required('usuarios.criar') ou múltiplos @perm_required('a','b')
    Se o usuário tiver qualquer uma das permissões listadas, passa.
    """
    def deco(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if "usuario_logado" not in session:
                flash("Faça login para acessar esta página.", "danger")
                return redirect(url_for("login"))
            perms = session.get("permissoes", [])
            if not any(r in perms for r in recursos):
                flash("Acesso negado para este recurso.", "danger")
                return redirect(url_for("home"))
            return f(*args, **kwargs)
        return wrapper
    return deco


def any_perm(*recursos) -> bool:
    perms = session.get("permissoes")
    if perms is None:
        return session.get("tipo_id") == 1
    perms = set(perms)
    return any(r in perms for r in recursos)


def all_perms(*recursos) -> bool:
    perms = session.get("permissoes")
    if perms is None:
        return session.get("tipo_id") == 1
    perms = set(perms)
    return all(r in perms for r in recursos)
def _xlsx_auto_width(ws, max_width=60):
    """Ajuste simples de largura de colunas (sem exagerar)."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)



# Disponibiliza os helpers pra TODOS os templates
app.jinja_env.globals.update(
    has_perm=has_perm,
    any_perm=any_perm,
    all_perms=all_perms,
)


# ==========
# UTILIDADES
# ==========

def get_tipos_user():
    """Carrega lista de tipos do OFS (tabela local tipos_ofs)."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT codigo, descricao FROM tipos_ofs ORDER BY descricao")
    resultados = cursor.fetchall()
    cursor.close()
    conn.close()
    return [{"codigo": row[0], "descricao": row[1]} for row in resultados]


def get_perfis():
    """Carrega lista de perfis do painel (tabela perfis)."""
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id, nome FROM perfis ORDER BY nome")
    perfis = cur.fetchall()
    cur.close()
    conn.close()
    return perfis


def _only_alnum_upper(s: str) -> str:
    s = (s or "").strip().upper()
    # remove acentos
    s_norm = unicodedata.normalize("NFKD", s)
    s_noacc = "".join(ch for ch in s_norm if not unicodedata.combining(ch))
    # mantém só letras e números
    s_clean = re.sub(r"[^A-Z0-9]", "", s_noacc)
    return s_clean

def _bairro_variants(nome_bairro_original: str) -> list[str]:
    """
    Regra:
    - Remove acentos
    - Remove espaços
    - Remove caracteres especiais
    - Uppercase
    - Retorna sempre a versão limpa
    - Se houver diferença estrutural relevante, retorna segunda variante
    """

    original = (nome_bairro_original or "").strip()
    if not original:
        return []

    # Versão base (remove acentos e não alfanuméricos)
    base = _only_alnum_upper(original)

    variants = []

    if base:
        variants.append(base)

    # Segunda variante: remove qualquer caractere não ASCII da string original
    # e normaliza novamente
    ascii_strict = (
        unicodedata.normalize("NFKD", original)
        .encode("ASCII", "ignore")
        .decode("ASCII")
    )
    ascii_strict = re.sub(r"[^A-Za-z0-9]", "", ascii_strict).upper()

    if ascii_strict and ascii_strict not in variants:
        variants.append(ascii_strict)

    return variants

def _workzone(id_cidade_int: int, bairro_key: str) -> str:
    """
    workzone = 8 dígitos do idCidade com zero à esquerda + (bairro_key + zeros) completando 40 chars.
    Total 48.
    """
    id8 = str(int(id_cidade_int)).zfill(8)
    bairro_key = (bairro_key or "").strip().upper()
    bairro40 = (bairro_key + ("0" * 40))[:40]  # completa e corta em 40
    return f"{id8}{bairro40}"

def _validate_max_range_7_days(date_from: str, date_to: str):
    try:
        df = datetime.strptime(date_from, "%Y-%m-%d").date()
        dt = datetime.strptime(date_to, "%Y-%m-%d").date()
    except Exception:
        raise ValueError("Informe um período válido (De / Até).")

    if dt < df:
        raise ValueError("O campo 'Até' não pode ser menor que 'De'.")

    if (dt - df).days > 6:  # 7 dias = 0..6
        raise ValueError("Limite máximo: 7 dias por atualização.")

    return df, dt
def _job_should_cancel(job_id: int) -> bool:
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT cancel_requested, status FROM ofs_import_jobs WHERE id=%s", (job_id,))
    row = cur.fetchone() or {}
    cur.close(); conn.close()
    return int(row.get("cancel_requested") or 0) == 1 or (row.get("status") == "canceled")

def _job_update(job_id: int, **fields):
    if not fields:
        return
    sets = []
    params = []
    for k, v in fields.items():
        sets.append(f"{k}=%s")
        params.append(v)
    params.append(job_id)
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(f"UPDATE ofs_import_jobs SET {', '.join(sets)} WHERE id=%s", tuple(params))
    conn.commit()
    cur.close(); conn.close()

def _iter_days(date_from: str, date_to: str):
    start = datetime.strptime(date_from, "%Y-%m-%d").date()
    end = datetime.strptime(date_to, "%Y-%m-%d").date()

    current = start
    while current <= end:
        yield current.strftime("%Y-%m-%d")
        current += timedelta(days=1)

def _run_import_job(job_id: int, date_from: str, date_to: str, resources: str, actor_username: str):
    try:
        client = OFSClient()

        fields = [
            "activityId", "city", "activityType", "apptNumber", "date", "status",
            "XA_RES_API_NG_RESPONSE", "XA_API_NG_DISPATCH", "XA_SAP_CRT_LDG", "XA_SAP_CRT"
        ]

        days = list(_iter_days(date_from, date_to))
        total_days = len(days)

        total_inserted = 0
        total_updated = 0
        total_api_items = 0

        _job_update(job_id, message="Iniciando importação...", progress=1)

        sql = """
            INSERT INTO ofs_activities_errors
            (
                activity_id, `date`, city, activity_type, appt_number, status,
                xa_res_api_ng_response, xa_api_ng_dispatch, xa_sap_crt_ldg, xa_sap_crt,
                sap_error_raw_extracted, ng_dispatch_message, ng_response_message,
                sap_response_message, sap_error_category, last_seen_at
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
            ON DUPLICATE KEY UPDATE
                `date`=VALUES(`date`),
                city=VALUES(city),
                activity_type=VALUES(activity_type),
                appt_number=VALUES(appt_number),
                status=VALUES(status),
                xa_res_api_ng_response=VALUES(xa_res_api_ng_response),
                xa_api_ng_dispatch=VALUES(xa_api_ng_dispatch),
                xa_sap_crt_ldg=VALUES(xa_sap_crt_ldg),
                xa_sap_crt=VALUES(xa_sap_crt),
                sap_error_raw_extracted=VALUES(sap_error_raw_extracted),
                ng_dispatch_message=VALUES(ng_dispatch_message),
                ng_response_message=VALUES(ng_response_message),
                sap_response_message=VALUES(sap_response_message),
                sap_error_category=VALUES(sap_error_category),
                last_seen_at=NOW()
        """
        for day_index, day in enumerate(days, start=1):
            if _job_should_cancel(job_id):
                _job_update(
                    job_id,
                    status="canceled",
                    message="Operação cancelada pelo usuário.",
                    progress=0
                )
                return

            items = []
            has_more = True
            page = 0
            max_pages = 30
            offset = 0

            while has_more and page < max_pages:
                if _job_should_cancel(job_id):
                    _job_update(
                        job_id,
                        status="canceled",
                        message="Operação cancelada pelo usuário.",
                        progress=0
                    )
                    return

                base_params = {
                    "dateFrom": day,
                    "dateTo": day,
                    "resources": resources,
                    "q": "status == 'notdone' OR status == 'completed'",
                    "fields": ",".join(fields),
                    "limit": 2000,
                    "offset": offset,
                }

                qs = urlencode(base_params, safe="=,'")
                url = f"{client.base_url}/activities/?{qs}"
                data = client.authenticated_get(url)

                batch = data.get("items") or []
                items.extend(batch)

                has_more = bool(data.get("hasMore"))
                offset = len(items)
                page += 1

                pct_base = int(((day_index - 1) / max(1, total_days)) * 80)
                pct_page = int((page / max(1, max_pages)) * (80 / max(1, total_days)))
                progress = min(75, pct_base + pct_page + 5)

                _job_update(
                    job_id,
                    progress=progress,
                    message=f"API: dia {day} | página {page}/{max_pages}"
                )

            total_api_items += len(items)

            if _job_should_cancel(job_id):
                _job_update(
                    job_id,
                    status="canceled",
                    message="Operação cancelada pelo usuário.",
                    progress=0
                )
                return

            _job_update(
                job_id,
                message=f"Gravando dia {day} no banco ({len(items)} itens)...",
                progress=min(85, 5 + int((day_index / max(1, total_days)) * 80))
            )

            conn = get_connection()
            cur = conn.cursor()

            try:
                # limpa SOMENTE o dia que está sendo reprocessado
                cur.execute("""
                    DELETE FROM ofs_activities_errors
                    WHERE `date` = %s
                """, (day,))

                inserted_day = 0
                updated_day = 0

                for idx, a in enumerate(items, start=1):
                    if _job_should_cancel(job_id):
                        conn.rollback()
                        cur.close()
                        conn.close()

                        _job_update(
                            job_id,
                            status="canceled",
                            message="Operação cancelada pelo usuário.",
                            progress=0
                        )
                        return

                    activity_id = str(a.get("activityId") or "").strip()
                    if not activity_id:
                        continue

                    ng_dispatch_raw = a.get("XA_API_NG_DISPATCH")
                    ng_response_raw = a.get("XA_RES_API_NG_RESPONSE")

                    ng_dispatch_msg = _extract_message(ng_dispatch_raw)
                    ng_response_msg = _extract_message(ng_response_raw)
                    sap_raw = a.get("XA_SAP_CRT_LDG")
                    xa_sap_crt = str(a.get("XA_SAP_CRT") or "").strip() or None

                    sap_info = parse_sap_error(
                        raw_value=sap_raw,
                        xa_sap_crt=xa_sap_crt
                    )
                    cur.execute(sql, (
                        activity_id,
                        str(a.get("date") or "") or None,
                        str(a.get("city") or "") or None,
                        str(a.get("activityType") or "") or None,
                        str(a.get("apptNumber") or "") or None,
                        str(a.get("status") or "") or None,
                        ng_response_raw,
                        ng_dispatch_raw,
                        sap_raw,
                        xa_sap_crt,
                        sap_info["sap_error_raw_extracted"],
                        ng_dispatch_msg,
                        ng_response_msg,
                        sap_info["sap_response_message"],
                        sap_info["sap_error_category"],
                    ))

                    if cur.rowcount == 1:
                        inserted_day += 1
                    elif cur.rowcount == 2:
                        updated_day += 1

                    if idx % 250 == 0:
                        progress = 85 + int((day_index / max(1, total_days)) * 10)
                        _job_update(
                            job_id,
                            progress=min(95, progress),
                            message=f"Gravando dia {day}... {idx}/{len(items)}"
                        )

                conn.commit()

                total_inserted += inserted_day
                total_updated += updated_day

            except Exception:
                conn.rollback()
                raise
            finally:
                cur.close()
                conn.close()

        _job_update(
            job_id,
            status="done",
            progress=100,
            message=(
                f"Concluído. Novos: {total_inserted} | "
                f"Atualizados: {total_updated} | "
                f"Total API: {total_api_items}"
            )
        )

    except Exception as e:
        print(traceback.format_exc(), flush=True)
        _job_update(
            job_id,
            status="error",
            message=f"{type(e).__name__}: {str(e)[:220]}",
            progress=0
        )

_MSG_RE = re.compile(r'"message"\s*:\s*"([^"]+)"')


def _extract_message(val):
    """
    Extrai SOMENTE o campo JSON "message":"..."
    Funciona mesmo quando vem JSON + texto depois.
    """
    if val is None:
        return None

    s = str(val)
    m = _MSG_RE.search(s)
    if not m:
        return None

    msg = (m.group(1) or "").strip()
    return msg or None

_SAP_HTTP_SUFFIX_RE = re.compile(
    r'\.The 500 Internal Server Error.*$',
    re.IGNORECASE | re.DOTALL
)

_SAP_CDATA_START_RE = re.compile(r'^\s*<!\[CDATA\[',
    re.IGNORECASE)
_SAP_CDATA_END_RE = re.compile(r'\]\]>\s*$',
    re.IGNORECASE)


def _normalize_space(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _strip_sap_wrapper(value):
    """
    Remove CDATA e o sufixo padrão de HTTP 500 do texto bruto do SAP.
    """
    if value is None:
        return None

    s = str(value).strip()
    if not s:
        return None

    s = _SAP_CDATA_START_RE.sub("", s)
    s = _SAP_CDATA_END_RE.sub("", s)
    s = _SAP_HTTP_SUFFIX_RE.sub("", s)
    s = s.strip()

    return s or None


def _extract_first_json_object(text):
    """
    Extrai o primeiro JSON {...} válido do texto.
    Não depende do conteúdo terminar exatamente no JSON.
    """
    if not text:
        return None

    start = text.find("{")
    if start == -1:
        return None

    depth = 0
    in_string = False
    escape = False

    for i in range(start, len(text)):
        ch = text[i]

        if in_string:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == '"':
                in_string = False
            continue

        if ch == '"':
            in_string = True
        elif ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                candidate = text[start:i + 1]
                try:
                    return json.loads(candidate)
                except Exception:
                    return None

    return None


def _try_parse_json_string(value):
    """
    Tenta parsear string JSON pura.
    """
    if value is None:
        return None

    s = str(value).strip()
    if not s:
        return None

    try:
        return json.loads(s)
    except Exception:
        return None


def _extract_best_text_from_obj(obj):
    """
    Prioriza os campos mais úteis para causa raiz.
    """
    if not isinstance(obj, dict):
        return None

    for key in ("data", "Documento", "Resposta", "message"):
        raw = obj.get(key)
        if raw is None:
            continue

        raw_str = str(raw).strip()
        if not raw_str:
            continue

        nested_obj = _try_parse_json_string(raw_str)
        if isinstance(nested_obj, dict):
            nested_best = _extract_best_text_from_obj(nested_obj)
            if nested_best:
                return nested_best

        return raw_str

    return None
SAP_ERROR_RULES = [
    {
        "category": "mac_duplicado",
        "patterns": [
            r"não é possivel inserir mac duplicado",
            r"nao e possivel inserir mac duplicado",
            r"mac duplicado",
        ],
        "message": "MAC duplicado"
    },
    {
        "category": "tecnico_filial_inexistente",
        "patterns": [
            r"o tecnico \d+ nao existe na filial \d+",
        ],
        "message": "Técnico não existe na filial"
    },
    {
        "category": "localizacao_tecnico_inexistente",
        "patterns": [
            r"localizacao do tecnico \d+ nao existe no deposito",
            r"localizacao \d+ no deposito .* nao existe",
        ],
        "message": "Localização do técnico não existe no depósito"
    },
    {
        "category": "warehouse_filial_invalida",
        "patterns": [
            r"warehouse is not assigned to the same branch as the document",
        ],
        "message": "Warehouse não pertence à mesma filial do documento"
    },
    {
        "category": "item_nao_encontrado_warehouse",
        "patterns": [
            r"item .* not found in warehouse",
        ],
        "message": "Item não encontrado no warehouse"
    },
    {
        "category": "bin_sem_saldo",
        "patterns": [
            r"allocated quantity exceeds available quantity",
        ],
        "message": "Quantidade alocada excede o saldo disponível"
    },
    {
        "category": "bin_inativo",
        "patterns": [
            r"inactive bin location",
        ],
        "message": "Bin location inativo"
    },
    {
        "category": "row_without_tax",
        "patterns": [
            r"row without tax was found",
        ],
        "message": "Linha sem imposto"
    },
    {
        "category": "protocolo_duplicado",
        "patterns": [
            r"protocolo adapter já existe no documento",
            r"protocolo já existe no documento",
            r"protocolo .* ja existe no documento",
        ],
        "message": "Protocolo já existe no documento"
    },
    {
        "category": "timeout_integracao",
        "patterns": [
            r"timeoutexception",
            r"read timeout",
            r"stream closed: read timeout",
        ],
        "message": "Timeout na integração SAP"
    },
    {
        "category": "connection_reset",
        "patterns": [
            r"connection reset",
            r"socketexception",
        ],
        "message": "Falha de conexão com o SAP"
    },
    {
        "category": "erro_interno_index_out_of_bounds",
        "patterns": [
            r"index was outside the bounds of the array",
        ],
        "message": "Erro interno SAP - índice fora do limite"
    },
    {
        "category": "erro_parse_json",
        "patterns": [
            r"after parsing a value an unexpected character was encountered",
        ],
        "message": "Erro ao interpretar retorno do SAP"
    },
    {
        "category": "string_too_long",
        "patterns": [
            r"string is too long",
            r"input string is longer than the maximum length",
        ],
        "message": "Valor maior que o tamanho permitido no SAP"
    },
    {
        "category": "serie_sem_custodia",
        "patterns": [
            r"numero de serie .* nao existe na custodia",
        ],
        "message": "Número de série não está na custódia do técnico"
    },
    {
        "category": "quantidade_maior_documento_base",
        "patterns": [
            r"quantity cannot exceed the quantity in the base document",
        ],
        "message": "Quantidade maior que a permitida no documento base"
    },
    {
        "category": "sequencia_nf_nao_definida",
        "patterns": [
            r"default sequence not defined or locked",
        ],
        "message": "Sequência padrão de nota fiscal não definida ou bloqueada"
    },
    {
        "category": "erro_localizacao_cliente",
        "patterns": [
            r"erro ao criar localizacao do cliente",
            r"g2_get_whs_id",
            r"no data found",
        ],
        "message": "Erro ao criar localização do cliente"
    },
    {
        "category": "erro_geral_hana",
        "patterns": [
            r"hdbodbc",
            r"transaction rolled back by an internal error",
            r"trexcolumnupdate failed",
        ],
        "message": "Erro interno de banco no SAP"
    },
    {
        "category": "erro_generico_sap",
        "patterns": [],
        "message": "Erro genérico no SAP"
    },
]
def parse_sap_error(raw_value, xa_sap_crt):
    """
    Retorna:
        {
            "sap_response_message": ...,
            "sap_error_category": ...,
            "sap_error_raw_extracted": ...
        }
    """
    if str(xa_sap_crt or "").strip() != "1":
        return {
            "sap_response_message": None,
            "sap_error_category": None,
            "sap_error_raw_extracted": None,
        }

    cleaned = _strip_sap_wrapper(raw_value)
    if not cleaned:
        return {
            "sap_response_message": "Erro SAP sem detalhe",
            "sap_error_category": "erro_sem_detalhe",
            "sap_error_raw_extracted": None,
        }

    obj = _extract_first_json_object(cleaned)

    if isinstance(obj, dict):
        best_text = _extract_best_text_from_obj(obj) or cleaned
    else:
        best_text = cleaned

    best_text = _normalize_space(best_text)
    for rule in SAP_ERROR_RULES:
        for pattern in rule["patterns"]:
            if re.search(pattern, best_text, re.IGNORECASE):
                return {
                    "sap_response_message": rule["message"],
                    "sap_error_category": rule["category"],
                    "sap_error_raw_extracted": best_text,
                }

    if best_text == '"null"' or best_text.lower() == "null":
        return {
            "sap_response_message": "Erro SAP sem detalhe",
            "sap_error_category": "erro_sem_detalhe",
            "sap_error_raw_extracted": best_text,
        }

    return {
        "sap_response_message": "Erro genérico no SAP",
        "sap_error_category": "erro_generico_sap",
        "sap_error_raw_extracted": best_text,
    }

@app.route("/atividades-notdone/exportar", methods=["POST"])
@login_required
@perm_required("ofs.atividades_notdone")
def atividades_notdone_exportar():
    """
    Exporta XLSX:
      - tipo=clientes -> tabela ofs_atividades_notdone filtrada por date (agendamento)
      - tipo=tratativas -> history + join com notdone, filtrada por h.created_at
    """
    tipo = (request.form.get("tipo") or "").strip().lower()
    date_from = (request.form.get("dateFrom") or "").strip()
    date_to = (request.form.get("dateTo") or "").strip()

    if tipo not in {"clientes", "tratativas"}:
        flash("Tipo de exportação inválido.", "danger")
        return redirect(url_for("atividades_notdone"))

    # valida datas
    try:
        dt_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        dt_to = datetime.strptime(date_to, "%Y-%m-%d").date()
    except Exception:
        flash("Informe um período válido (De / Até).", "danger")
        return redirect(url_for("atividades_notdone"))

    if dt_to < dt_from:
        flash("O campo 'Até' não pode ser menor que 'De'.", "danger")
        return redirect(url_for("atividades_notdone"))

    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    try:
        if tipo == "clientes":
            # (IMPORTANTE) date é VARCHAR, então compare como string.
            # E use crase porque "date" é um nome sensível.
            cur.execute("""
                SELECT
                    activity_id,
                    `date`,
                    city,
                    customer_number,
                    customer_phone,
                    customer_name,
                    appt_number,
                    origin_bucket,
                    ser_clo_imp_ada,
                    resource_id,
                    tratativa_status,
                    tratado_por_username,
                    tratado_em,
                    created_at
                FROM ofs_atividades_notdone
                WHERE `date` BETWEEN %s AND %s
                ORDER BY `date` ASC, created_at DESC
            """, (date_from, date_to))  # <-- strings mesmo

            rows = cur.fetchall()

            sheet_name = "Clientes"
            headers = [
                "activity_id", "date", "city",
                "customer_number", "customer_phone", "customer_name",
                "appt_number", "origin_bucket",
                "ser_clo_imp_ada", "resource_id",
                "tratativa_status", "tratado_por_username", "tratado_em", "created_at"
            ]

        else:
            # Filtra por created_at da history (data/hora da ação)
            # Join para trazer customer_name/number/appt_number
            cur.execute("""
                SELECT
                    h.id AS history_id,
                    h.activity_id,
                    n.customer_name,
                    n.customer_number,
                    n.appt_number,

                    h.action,
                    h.status,
                    h.obs,
                    h.actor_username,
                    h.created_at
                FROM ofs_atividades_notdone_history h
                LEFT JOIN ofs_atividades_notdone n
                  ON n.activity_id = h.activity_id
                WHERE h.created_at BETWEEN %s AND %s
                ORDER BY h.created_at DESC
            """, (f"{dt_from} 00:00:00", f"{dt_to} 23:59:59"))
            rows = cur.fetchall()

            sheet_name = "Tratativas"
            headers = [
                "history_id", "activity_id",
                "customer_name", "customer_number", "appt_number",
                "action", "status", "obs", "actor_username", "created_at"
            ]

    finally:
        cur.close()
        conn.close()

    # Monta XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header
    ws.append(headers)

    # Dados
    for r in rows:
        ws.append([r.get(h) for h in headers])

    _xlsx_auto_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"ofs_notdone_{tipo}_{dt_from}_{dt_to}_{stamp}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/toquio/td-bucket/inserir-mapeamento-bairro", methods=["GET", "POST"])
@login_required
@perm_required("toquio.td_bucket_insert")
def toquio_td_bucket_inserir_mapeamento_bairro():
    # =========================
    # CONSULTA (GET)
    # =========================
    id_cidade_q_raw = (request.args.get("idCidade") or "").strip()
    nome_cidade_q = (request.args.get("nomeCidade") or "").strip()
    chave_like_q = (request.args.get("chave") or "").strip()

    # Para manter o valor no input do template
    idCidade_q = id_cidade_q_raw

    resultados = []

    # Monta WHERE dinamicamente
    where = []
    params = []

    # idCidade (igualdade)
    if id_cidade_q_raw:
        if not id_cidade_q_raw.isdigit():
            # se preferir, pode retornar mensagem de erro ao invés de ignorar
            # flash("idCidade deve ser numérico", "warning")
            pass
        else:
            where.append("idCidade = %s")
            params.append(int(id_cidade_q_raw))

    # nomeCidade (LIKE)
    if nome_cidade_q:
        where.append("nomeCidade LIKE %s")
        params.append(f"%{nome_cidade_q}%")

    # chave (LIKE)
    if chave_like_q:
        where.append("chave LIKE %s")
        params.append(f"%{chave_like_q}%")

    # Só consulta se tiver pelo menos 1 filtro
    if where:
        conn = get_connection()
        cur = conn.cursor(dictionary=True)

        sql = """
            SELECT *
            FROM projToquio.td_bucket
        """

        sql += " WHERE " + " AND ".join(where)
        sql += " LIMIT 200"

        cur.execute(sql, params)
        resultados = cur.fetchall() or []

        cur.close()
        conn.close()
    # =========================
    # INSERT (POST)
    # =========================
    if request.method == "POST":

        bucket = (request.form.get("bucket") or "").strip()
        sistema = (request.form.get("sistema") or "").strip().upper()
        nome_cidade = (request.form.get("nomeCidade") or "").strip()
        nome_bairro = (request.form.get("nomeBairro") or "").strip()
        uf = (request.form.get("uf") or "").strip().upper()
        id_cidade_raw = (request.form.get("idCidade") or "").strip()
        area_bucket = (request.form.get("areaBucket") or "").strip()
        filial_cidade_raw = (request.form.get("filialCidade") or "").strip()
        regional_pbi = (request.form.get("Regional_PBI") or "").strip()

        # Normaliza UF
        uf = re.sub(r"[^A-Z]", "", uf)[:3]

        # =========================
        # Validação enxuta
        # =========================
        required_fields = {
            "bucket": "Bucket",
            "sistema": "Sistema de origem",
            "nomeCidade": "Nome da cidade",
            "uf": "UF",
            "idCidade": "idCidade",
            "areaBucket": "Região",
            "filialCidade": "filialCidade",
            "Regional_PBI": "Regional_PBI",
        }

        values = {
            "bucket": bucket,
            "sistema": sistema,
            "nomeCidade": nome_cidade,
            "uf": uf,
            "idCidade": id_cidade_raw,
            "areaBucket": area_bucket,
            "filialCidade": filial_cidade_raw,
            "Regional_PBI": regional_pbi,
        }

        missing = [label for key, label in required_fields.items() if not values.get(key)]

        invalid_numeric = []
        if id_cidade_raw and not id_cidade_raw.isdigit():
            invalid_numeric.append("idCidade")
        if filial_cidade_raw and not filial_cidade_raw.isdigit():
            invalid_numeric.append("filialCidade")

        if missing or invalid_numeric:
            parts = []
            if missing:
                parts.append("Campos obrigatórios vazios: " + ", ".join(missing))
            if invalid_numeric:
                parts.append("Campos numéricos inválidos: " + ", ".join(invalid_numeric))
            flash(" | ".join(parts), "error")
            return redirect(url_for("toquio_td_bucket_inserir_mapeamento_bairro"))

        id_cidade = int(id_cidade_raw)
        filial_cidade = int(filial_cidade_raw)

        supervisor = "DEFAULT"
        is_ativo = 1
        id_bairro = None

        variants = _bairro_variants(nome_bairro)

        rows_to_insert = []

        if not variants:
            bairro_key = "0000000"
            chave = f"{sistema}{id_cidade}DEFAULT"
            workzone = _workzone(id_cidade, bairro_key)

            rows_to_insert.append({
                "bucket": bucket,
                "chave": chave,
                "idCidade": id_cidade,
                "idBairro": id_bairro,
                "nomeCidade": nome_cidade,
                "nomeBairro": nome_bairro,
                "uf": uf,
                "workzone": workzone,
                "supervisor": supervisor,
                "sistema": sistema,
                "areaBucket": area_bucket,
                "filialCidade": filial_cidade,
                "Regional_PBI": regional_pbi,
                "isAtivo": is_ativo
            })
        else:
            for bairro_key in variants:
                chave = f"{id_cidade}{bairro_key}"
                workzone = _workzone(id_cidade, bairro_key)

                rows_to_insert.append({
                    "bucket": bucket,
                    "chave": chave,
                    "idCidade": id_cidade,
                    "idBairro": id_bairro,
                    "nomeCidade": nome_cidade,
                    "nomeBairro": nome_bairro,
                    "uf": uf,
                    "workzone": workzone,
                    "supervisor": supervisor,
                    "sistema": sistema,
                    "areaBucket": area_bucket,
                    "filialCidade": filial_cidade,
                    "Regional_PBI": regional_pbi,
                    "isAtivo": is_ativo
                })

        try:
            conn = get_connection()
            cur = conn.cursor()

            sql_insert = """
                INSERT INTO projToquio.td_bucket
                (bucket, chave, idCidade, idBairro, nomeCidade, nomeBairro, uf, workzone,
                 supervisor, sistema, areaBucket, filialCidade, Regional_PBI, isAtivo)
                VALUES
                (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """

            for r in rows_to_insert:
                cur.execute(sql_insert, (
                    r["bucket"], r["chave"], r["idCidade"], r["idBairro"],
                    r["nomeCidade"], r["nomeBairro"], r["uf"], r["workzone"],
                    r["supervisor"], r["sistema"], r["areaBucket"],
                    r["filialCidade"], r["Regional_PBI"], r["isAtivo"]
                ))

            conn.commit()
            cur.close()
            conn.close()

            flash(
                "Inserido com sucesso. "
                + " | ".join([f"Key para o OFS={x['workzone']}" for x in rows_to_insert]),
                "success"
            )

        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            flash(f"Erro ao inserir: {str(e)}", "error")

        return redirect(url_for("toquio_td_bucket_inserir_mapeamento_bairro"))

    return render_template(
        "toquio_inserir_mapeamento_bairro.html",
        resultados=resultados,
        idCidade_q=idCidade_q,
        nomeCidade_q=nome_cidade_q,
        chave_q=chave_like_q
    )

# =============================
# OFS - Activities Errors
# =============================
@app.route("/ofs/activities-errors/export/xlsx", methods=["GET"])
@login_required
@perm_required("ofs.activities_errors")
def ofs_activities_errors_export_xlsx():
    today = datetime.now().strftime("%Y-%m-%d")
    date_from = (request.args.get("dateFrom") or today).strip()
    date_to = (request.args.get("dateTo") or today).strip()
    resources = (request.args.get("resources") or "02").strip()

    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute("""
        SELECT
            activity_id,
            city,
            activity_type,
            appt_number,
            status,
            ng_dispatch_message,
            ng_response_message,
            sap_error_raw_extracted,
            sap_response_message,
            sap_error_category,
            xa_sap_crt_ldg,
            `date`
        FROM ofs_activities_errors
        WHERE `date` BETWEEN %s AND %s
        AND activity_type IN (
                'INS',
                'SUP_QUA',
                'SUP_REP',
                'SOL_SER',
                'INS_DEV',
                'SUP',
                'MIG_PLA',
                'QUA',
                'MIG_TEC'
        )
        AND (
                NULLIF(TRIM(ng_dispatch_message), '') IS NOT NULL
                OR NULLIF(TRIM(ng_response_message), '') IS NOT NULL
                OR NULLIF(TRIM(sap_response_message), '') IS NOT NULL
                OR NULLIF(TRIM(xa_sap_crt_ldg), '') IS NOT NULL
        )
        ORDER BY `date` DESC, activity_type, city, appt_number
    """, (date_from, date_to))

    rows = cur.fetchall()
    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Atividades com erro"

    headers = [
        "activity_id",
        "city",
        "activity_type",
        "appt_number",
        "status",
        "ng_dispatch_message",
        "ng_response_message",
        "sap_error_raw_extracted",
        "sap_response_message",
        "sap_error_category",
        "xa_sap_crt_ldg",
        "date",
    ]
    ws.append(headers)

    for row in rows:
        ws.append([
            row.get("activity_id"),
            row.get("city"),
            row.get("activity_type"),
            row.get("appt_number"),
            row.get("status"),
            row.get("ng_dispatch_message"),
            row.get("ng_response_message"),
            row.get("sap_error_raw_extracted"),
            row.get("sap_response_message"),
            row.get("sap_error_category"),
            row.get("xa_sap_crt_ldg"),
            row.get("date"),
        ])
    # Ajuste simples de largura
    for col_idx, col_name in enumerate(headers, start=1):
        max_len = len(col_name)
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"ofs_activities_errors_{date_from}_a_{date_to}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
@app.route("/ofs/activities-errors/dashboard/data", methods=["GET"])
@login_required
@perm_required("ofs.activities_errors")
def ofs_activities_errors_dashboard_data():
    today = datetime.now().strftime("%Y-%m-%d")
    date_from = (request.args.get("dateFrom") or today).strip()
    date_to = (request.args.get("dateTo") or today).strip()
    resources = (request.args.get("resources") or "02").strip()

    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute("""
        SELECT COUNT(*) AS total
        FROM ofs_activities_errors
        WHERE `date` BETWEEN %s AND %s
    """, (date_from, date_to))
    total = (cur.fetchone() or {}).get("total", 0)

    cur.execute("""
        SELECT COUNT(*) AS total_ng
        FROM ofs_activities_errors
        WHERE `date` BETWEEN %s AND %s
          AND (ng_dispatch_message IS NOT NULL OR ng_response_message IS NOT NULL)
    """, (date_from, date_to))
    total_ng = (cur.fetchone() or {}).get("total_ng", 0)

    cur.execute("""
        SELECT msg, COUNT(*) AS qtd
        FROM (
            SELECT ng_dispatch_message AS msg
            FROM ofs_activities_errors
            WHERE `date` BETWEEN %s AND %s AND ng_dispatch_message IS NOT NULL
            UNION ALL
            SELECT ng_response_message AS msg
            FROM ofs_activities_errors
            WHERE `date` BETWEEN %s AND %s AND ng_response_message IS NOT NULL
        ) x
        GROUP BY msg
        ORDER BY qtd DESC
        LIMIT 15
    """, (date_from, date_to, date_from, date_to))
    top_messages = cur.fetchall()

    cur.execute("""
        SELECT `date`, COUNT(*) AS qtd
        FROM ofs_activities_errors
        WHERE `date` BETWEEN %s AND %s
          AND (ng_dispatch_message IS NOT NULL OR ng_response_message IS NOT NULL)
        GROUP BY `date`
        ORDER BY `date` ASC
        LIMIT 31
    """, (date_from, date_to))
    by_day = cur.fetchall()
    cur.execute("""
        SELECT
            `date`,
            COALESCE(e.activity_type, '-') AS activityType,
            COALESCE(c.descricao, COALESCE(e.activity_type, '-')) AS activityTypeLabel,
            COUNT(*) AS qtd
        FROM ofs_activities_errors e
        LEFT JOIN ofs_activity_type_config c
            ON c.activity_type = e.activity_type
           AND c.ativo = 1
        WHERE e.`date` BETWEEN %s AND %s
          AND (
                e.ng_dispatch_message IS NOT NULL
                OR e.ng_response_message IS NOT NULL
              )
          AND (
                c.mostrar_dashboard = 1
                OR c.id IS NULL
              )
        GROUP BY
            `date`,
            COALESCE(e.activity_type, '-'),
            COALESCE(c.descricao, COALESCE(e.activity_type, '-'))
        ORDER BY `date` ASC
    """, (date_from, date_to))
    by_day_type = cur.fetchall()
    cur.execute("""
        SELECT
            COALESCE(e.activity_type, '-') AS activityType,
            COALESCE(c.descricao, COALESCE(e.activity_type, '-')) AS activityTypeLabel,
            COUNT(*) AS qtd
        FROM ofs_activities_errors e
        LEFT JOIN ofs_activity_type_config c
            ON c.activity_type = e.activity_type
        AND c.ativo = 1
        WHERE e.`date` BETWEEN %s AND %s
        AND (
                e.ng_dispatch_message IS NOT NULL
                OR e.ng_response_message IS NOT NULL
            )
        AND (
                c.mostrar_dashboard = 1
                OR c.id IS NULL
            )
        GROUP BY
            COALESCE(e.activity_type, '-'),
            COALESCE(c.descricao, COALESCE(e.activity_type, '-'))
        ORDER BY qtd DESC
        LIMIT 20
    """, (date_from, date_to))
    by_type = cur.fetchall()
    cur.execute("""
        SELECT
            sap_response_message AS msg,
            COUNT(*) AS qtd
        FROM ofs_activities_errors
        WHERE `date` BETWEEN %s AND %s
          AND sap_response_message IS NOT NULL
          AND NULLIF(TRIM(sap_response_message), '') IS NOT NULL
        GROUP BY sap_response_message
        ORDER BY qtd DESC
        LIMIT 15
    """, (date_from, date_to))
    top_sap_messages = cur.fetchall()

    cur.execute("""
        SELECT
            COALESCE(e.activity_type, '-') AS activityType,
            COALESCE(c.descricao, COALESCE(e.activity_type, '-')) AS activityTypeLabel,
            COUNT(*) AS qtd
        FROM ofs_activities_errors e
        LEFT JOIN ofs_activity_type_config c
            ON c.activity_type = e.activity_type
            AND c.ativo = 1
        WHERE e.`date` BETWEEN %s AND %s
            AND COALESCE(TRIM(e.xa_sap_crt), '') = '1'
        GROUP BY
            COALESCE(e.activity_type, '-'),
            COALESCE(c.descricao, COALESCE(e.activity_type, '-'))
        ORDER BY qtd DESC
        LIMIT 20
    """, (date_from, date_to))
    sap_by_activity_type = cur.fetchall()
    cur.close()
    conn.close()

    return jsonify({
        "ok": True,
        "total": total,
        "total_ng": total_ng,
        "top_messages": top_messages,
        "by_day": by_day,
        "by_day_type": by_day_type,
        "by_type": by_type,
        "date_from": date_from,
        "date_to": date_to,
        "top_sap_messages": top_sap_messages,
        "sap_by_activity_type": sap_by_activity_type,
        "debug_marker": "ROTA_NOVA_SAP",
        "resources": resources
    }), 200

@app.route("/ofs/config/error-owners", methods=["GET"])
@login_required
@perm_required("ofs.activities_errors")
def ofs_error_owners_config():
    q = str(request.args.get("q") or "").strip()
    status = str(request.args.get("status") or "all").strip().lower()
    page = request.args.get("page", default=1, type=int)
    per_page = 50
    offset = (page - 1) * per_page

    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    # Base sem filtro de status, usada para os cards de resumo
    summary_where_filters = []
    summary_params = []

    if q:
        like = f"%{q}%"
        summary_where_filters.append("(base.error_message LIKE %s OR base.origem LIKE %s)")
        summary_params.extend([like, like])

    summary_where_sql = ""
    if summary_where_filters:
        summary_where_sql = "WHERE " + " AND ".join(summary_where_filters)

    summary_base_sql = f"""
        FROM (
            SELECT
                'NG_DISPATCH' AS origem,
                TRIM(ng_dispatch_message) AS error_message,
                COUNT(*) AS qtd
            FROM ofs_activities_errors
            WHERE ng_dispatch_message IS NOT NULL
              AND NULLIF(TRIM(ng_dispatch_message), '') IS NOT NULL
            GROUP BY TRIM(ng_dispatch_message)

            UNION ALL

            SELECT
                'NG_RESPONSE' AS origem,
                TRIM(ng_response_message) AS error_message,
                COUNT(*) AS qtd
            FROM ofs_activities_errors
            WHERE ng_response_message IS NOT NULL
              AND NULLIF(TRIM(ng_response_message), '') IS NOT NULL
            GROUP BY TRIM(ng_response_message)
        ) base
        LEFT JOIN ofs_error_owner_config cfg
            ON cfg.origem = base.origem
           AND TRIM(cfg.error_message) = base.error_message
           AND cfg.ativo = 1
        {summary_where_sql}
    """

    cur.execute(f"""
        SELECT COUNT(*) AS total
        {summary_base_sql}
    """, tuple(summary_params))
    total = (cur.fetchone() or {}).get("total", 0)

    cur.execute(f"""
        SELECT COUNT(*) AS total_configured
        {summary_base_sql}
        {"AND" if summary_where_sql else "WHERE"} cfg.id IS NOT NULL
    """, tuple(summary_params))
    total_configured = (cur.fetchone() or {}).get("total_configured", 0)

    cur.execute(f"""
        SELECT COUNT(*) AS total_pending
        {summary_base_sql}
        {"AND" if summary_where_sql else "WHERE"} cfg.id IS NULL
    """, tuple(summary_params))
    total_pending = (cur.fetchone() or {}).get("total_pending", 0)

    # Base paginada, com filtro de status
    where_filters = []
    params = []

    if q:
        like = f"%{q}%"
        where_filters.append("(base.error_message LIKE %s OR base.origem LIKE %s)")
        params.extend([like, like])

    if status == "configured":
        where_filters.append("cfg.id IS NOT NULL")
    elif status == "pending":
        where_filters.append("cfg.id IS NULL")

    where_sql = ""
    if where_filters:
        where_sql = "WHERE " + " AND ".join(where_filters)

    base_sql = f"""
        FROM (
            SELECT
                'NG_DISPATCH' AS origem,
                TRIM(ng_dispatch_message) AS error_message,
                COUNT(*) AS qtd
            FROM ofs_activities_errors
            WHERE ng_dispatch_message IS NOT NULL
              AND NULLIF(TRIM(ng_dispatch_message), '') IS NOT NULL
            GROUP BY TRIM(ng_dispatch_message)

            UNION ALL

            SELECT
                'NG_RESPONSE' AS origem,
                TRIM(ng_response_message) AS error_message,
                COUNT(*) AS qtd
            FROM ofs_activities_errors
            WHERE ng_response_message IS NOT NULL
              AND NULLIF(TRIM(ng_response_message), '') IS NOT NULL
            GROUP BY TRIM(ng_response_message)
        ) base
        LEFT JOIN ofs_error_owner_config cfg
            ON cfg.origem = base.origem
           AND TRIM(cfg.error_message) = base.error_message
           AND cfg.ativo = 1
        {where_sql}
    """

    cur.execute(f"""
        SELECT COUNT(*) AS total_filtered
        {base_sql}
    """, tuple(params))
    total_filtered = (cur.fetchone() or {}).get("total_filtered", 0)
    total_pages = max(1, (total_filtered + per_page - 1) // per_page)

    cur.execute(f"""
        SELECT
            base.origem,
            base.error_message,
            cfg.responsavel,
            CASE
                WHEN cfg.id IS NULL THEN 0
                ELSE 1
            END AS configurado,
            base.qtd
        {base_sql}
        ORDER BY
            base.qtd DESC,
            base.origem ASC
        LIMIT %s OFFSET %s
    """, tuple(params + [per_page, offset]))

    items = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "ofs_error_owner_config/ofs_error_owner_config.html",
        items=items,
        q=q,
        status=status,
        page=page,
        total=total,
        total_pages=total_pages,
        total_configured=total_configured,
        total_pending=total_pending,
        total_filtered=total_filtered
    )

@app.route("/ofs/config/error-owners/save", methods=["POST"])
@login_required
@perm_required("ofs.activities_errors")
def ofs_error_owners_config_save():
    origem = str(request.form.get("origem") or "").strip()
    error_message = request.form.get("error_message") or ""
    responsavel = str(request.form.get("responsavel") or "").strip()

    allowed = {"WFM", "NG", "Desconsiderar"}

    if not origem:
        flash("Origem obrigatória.", "error")
        return redirect(url_for("ofs_error_owners_config"))

    if not error_message:
        flash("Mensagem obrigatória.", "error")
        return redirect(url_for("ofs_error_owners_config"))

    if responsavel not in allowed:
        flash("Responsável inválido.", "error")
        return redirect(url_for("ofs_error_owners_config"))

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO ofs_error_owner_config (origem, error_message, responsavel, ativo)
        VALUES (%s, %s, %s, 1)
        ON DUPLICATE KEY UPDATE
            responsavel = VALUES(responsavel),
            ativo = 1,
            atualizado_em = CURRENT_TIMESTAMP
    """, (origem, error_message, responsavel))

    conn.commit()
    cur.close()
    conn.close()

    flash("Responsável salvo com sucesso.", "success")
    return redirect(url_for("ofs_error_owners_config"))

@app.route("/ofs/config/activity-types", methods=["GET"])
@login_required
@perm_required("ofs.activities_errors")
def ofs_activity_types_config():
    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute("""
        SELECT
            base.activity_type,
            cfg.descricao,
            CASE
                WHEN cfg.id IS NULL THEN 0
                ELSE 1
            END AS configurado,
            base.qtd
        FROM (
            SELECT
                COALESCE(activity_type, '-') AS activity_type,
                COUNT(*) AS qtd
            FROM ofs_activities_errors
            GROUP BY COALESCE(activity_type, '-')
        ) base
        LEFT JOIN ofs_activity_type_config cfg
            ON cfg.activity_type = base.activity_type
           AND cfg.ativo = 1
        ORDER BY
            configurado ASC,
            base.qtd DESC,
            base.activity_type ASC
    """)
    items = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "ofs_activity_type_config/ofs_activity_type_config.html",
        items=items
    )


@app.route("/ofs/config/activity-types/save", methods=["POST"])
@login_required
@perm_required("ofs.activities_errors")
def ofs_activity_types_config_save():
    activity_type = str(request.form.get("activity_type") or "").strip()
    descricao = str(request.form.get("descricao") or "").strip()

    if not activity_type:
        flash("activity_type obrigatório.", "error")
        return redirect(url_for("ofs_activity_types_config"))

    if activity_type == "-":
        flash("Não é permitido configurar o tipo '-'.", "error")
        return redirect(url_for("ofs_activity_types_config"))

    if not descricao:
        flash("Descrição obrigatória.", "error")
        return redirect(url_for("ofs_activity_types_config"))

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO ofs_activity_type_config (activity_type, descricao, ativo)
        VALUES (%s, %s, 1)
        ON DUPLICATE KEY UPDATE
            descricao = VALUES(descricao),
            ativo = 1,
            atualizado_em = CURRENT_TIMESTAMP
    """, (activity_type, descricao))

    conn.commit()
    cur.close()
    conn.close()

    flash(f"Tipo '{activity_type}' salvo com sucesso.", "success")
    return redirect(url_for("ofs_activity_types_config"))

@app.route("/ofs/activities-errors", methods=["GET"])
@login_required
@perm_required("ofs.activities_errors")
def ofs_activities_errors():
    today = datetime.now().strftime("%Y-%m-%d")
    date_from = (request.args.get("dateFrom") or today).strip()
    date_to = (request.args.get("dateTo") or today).strip()
    resources = (request.args.get("resources") or "02").strip()

    per_page = 50
    page = request.args.get("page", default=1, type=int)
    offset = (page - 1) * per_page

    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute("""
        SELECT COUNT(*) AS total
        FROM ofs_activities_errors
        WHERE `date` BETWEEN %s AND %s
    """, (date_from, date_to))
    total = (cur.fetchone() or {}).get("total", 0)
    total_pages = max(1, (total + per_page - 1) // per_page)

    cur.execute("""
    SELECT
        activity_id AS activityId,
        `date` AS date,
        city,
        activity_type AS activityType,
        appt_number AS apptNumber,
        status,
        xa_sap_crt AS XA_SAP_CRT,
        xa_sap_crt_ldg AS XA_SAP_CRT_LDG,
        xa_res_api_ng_response AS XA_RES_API_NG_RESPONSE,
        ng_dispatch_message AS ngDispatchMessage,
        ng_response_message AS ngResponseMessage,
        sap_error_raw_extracted AS sapErrorRawExtracted,
        sap_response_message AS sapResponseMessage,
        sap_error_category AS sapErrorCategory,
        CASE
            WHEN COALESCE(TRIM(xa_sap_crt), '') = '1'
            AND (
                    TRIM(COALESCE(xa_res_api_ng_response, '')) LIKE '<![CDATA[%'
                OR TRIM(COALESCE(xa_api_ng_dispatch, ''))     LIKE '<![CDATA[%'
            )
                THEN 'Erro SAP/NG'

            WHEN COALESCE(TRIM(xa_sap_crt), '') = '1'
                THEN 'Erro SAP'

            WHEN COALESCE(TRIM(xa_sap_crt), '') <> '1'
            AND (
                    TRIM(COALESCE(xa_res_api_ng_response, '')) LIKE '<![CDATA[%'
                OR TRIM(COALESCE(xa_api_ng_dispatch, ''))     LIKE '<![CDATA[%'
            )
                THEN 'Erro NG'

            ELSE '-'
        END AS erro_tipo,
        last_seen_at
    FROM ofs_activities_errors
    WHERE `date` BETWEEN %s AND %s
    ORDER BY `date` DESC, last_seen_at DESC
    LIMIT %s OFFSET %s
    """, (date_from, date_to, per_page, offset))

    items = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "ofs_activities_errors/ofs_activities_errors.html",
        items=items,
        total=total,
        page=page,
        total_pages=total_pages,
        date_from=date_from,
        date_to=date_to,
        resources=resources
    )


# -----------------------------
# IMPORT (start/status/cancel)
# -----------------------------
@app.route("/ofs/activities-errors/importar/start", methods=["POST"])
@login_required
@perm_required("ofs.activities_errors")
def ofs_activities_errors_importar_start():
    today = datetime.now().strftime("%Y-%m-%d")
    date_from = (request.form.get("dateFrom") or today).strip()
    date_to = (request.form.get("dateTo") or today).strip()
    resources = (request.form.get("resources") or "02").strip()

    try:
        _validate_max_range_7_days(date_from, date_to)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

    actor = current_actor()
    username = actor.get("username")

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO ofs_import_jobs (module, status, progress, message, created_by)
        VALUES ('ofs.activities_errors', 'running', 0, 'Iniciando...', %s)
    """, (username,))
    job_id = cur.lastrowid
    conn.commit()
    cur.close()
    conn.close()

    t = threading.Thread(
        target=_run_import_job,
        args=(job_id, date_from, date_to, resources, username),

    )
    t.start()

    return jsonify({"ok": True, "jobId": job_id}), 200


@app.route("/ofs/activities-errors/importar/status/<int:job_id>", methods=["GET"])
@login_required
@perm_required("ofs.activities_errors")
def ofs_activities_errors_importar_status(job_id):
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id, status, progress, message, created_at, updated_at
        FROM ofs_import_jobs
        WHERE id=%s AND module='ofs.activities_errors'
        LIMIT 1
    """, (job_id,))
    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row:
        return jsonify({"ok": False, "error": "Job não encontrado"}), 404

    return jsonify({"ok": True, "job": row}), 200


@app.route("/ofs/activities-errors/importar/cancel/<int:job_id>", methods=["POST"])
@login_required
@perm_required("ofs.activities_errors")
def ofs_activities_errors_importar_cancel(job_id):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        UPDATE ofs_import_jobs
        SET cancel_requested=1, message='Cancelamento solicitado...'
        WHERE id=%s AND module='ofs.activities_errors' AND status='running'
    """, (job_id,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"ok": True}), 200


# -----------------------------
# DETALHE (modal)
# -----------------------------
@app.route("/ofs/activities-errors/<activity_id>", methods=["GET"])
@login_required
@perm_required("ofs.activities_errors")
def ofs_activities_errors_get(activity_id):
    activity_id = str(activity_id or "").strip()
    if not activity_id:
        return jsonify({"ok": False, "error": "activityId inválido"}), 400

    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute("""
        SELECT
            activity_id AS activityId,
            xa_api_ng_dispatch AS XA_API_NG_DISPATCH,
            xa_res_api_ng_response AS XA_RES_API_NG_RESPONSE,
            xa_sap_crt_ldg AS XA_SAP_CRT_LDG,
            xa_sap_crt AS XA_SAP_CRT,
            ng_dispatch_message AS ngDispatchMessage,
            ng_response_message AS ngResponseMessage,
            sap_error_raw_extracted AS sapErrorRawExtracted,
            sap_response_message AS sapResponseMessage,
            sap_error_category AS sapErrorCategory
        FROM ofs_activities_errors
        WHERE activity_id = %s
        LIMIT 1
    """, (activity_id,))

    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row:
        return jsonify({"ok": False, "error": "Não encontrado"}), 404

    return jsonify({"ok": True, "item": row}), 200
@app.route("/ofs/activities-errors/export-top-sap-messages", methods=["GET"])
@login_required
@perm_required("ofs.activities_errors")
def ofs_activities_errors_export_top_sap_messages():
    today = datetime.now().strftime("%Y-%m-%d")
    date_from = (request.args.get("dateFrom") or today).strip()
    date_to = (request.args.get("dateTo") or today).strip()

    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute("""
        SELECT
            activity_id,
            `date`,
            city,
            activity_type,
            appt_number,
            status,
            sap_response_message
        FROM ofs_activities_errors
        WHERE `date` BETWEEN %s AND %s
        AND sap_response_message IS NOT NULL
        AND NULLIF(TRIM(sap_response_message), '') IS NOT NULL
        ORDER BY `date` DESC, last_seen_at DESC
    """, (date_from, date_to))

    rows = cur.fetchall()
    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Top erros SAP"

    ws.append([
        "activity_id",
        "date",
        "city",
        "activity_type",
        "appt_number",
        "status",
        "sap_response_message"
    ])

    for row in rows:
        ws.append([
            row.get("activity_id"),
            row.get("date"),
            row.get("city"),
            row.get("activity_type"),
            row.get("appt_number"),
            row.get("status"),
            row.get("sap_response_message"),
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"top_erros_sap_{date_from}_a_{date_to}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
@app.route("/ofs/activities-errors/export/top-messages", methods=["GET"])
@login_required
@perm_required("ofs.activities_errors")
def ofs_activities_errors_export_top_messages():

    date_from = (request.args.get("dateFrom") or "").strip()
    date_to = (request.args.get("dateTo") or "").strip()

    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute("""
        SELECT
            activity_id,
            `date`,
            city,
            activity_type,
            appt_number,
            status,
            ng_dispatch_message,
            ng_response_message
        FROM ofs_activities_errors
        WHERE `date` BETWEEN %s AND %s
          AND (
                ng_dispatch_message IS NOT NULL
                OR ng_response_message IS NOT NULL
              )
        ORDER BY `date` DESC
    """, (date_from, date_to))

    rows = cur.fetchall()

    cur.close()
    conn.close()

    output = BytesIO()

    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet("erros_ng")

    headers = [
        "activity_id",
        "date",
        "city",
        "activity_type",
        "appt_number",
        "status",
        "ng_dispatch_message",
        "ng_response_message"
    ]

    # escreve header
    for col, header in enumerate(headers):
        worksheet.write(0, col, header)

    # escreve dados
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, header in enumerate(headers):
            worksheet.write(row_idx, col_idx, row.get(header))

    workbook.close()

    output.seek(0)

    filename = f"ofs_erros_ng_{date_from}_a_{date_to}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
@app.route("/ofs/activities-errors/dashboard", methods=["GET"])
@login_required
@perm_required("ofs.activities_errors")
def ofs_activities_errors_dashboard():
    today = datetime.now().strftime("%Y-%m-%d")
    date_from = (request.args.get("dateFrom") or today).strip()
    date_to = (request.args.get("dateTo") or today).strip()
    resources = (request.args.get("resources") or "02").strip()

    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    # KPI: total linhas no período
    cur.execute("""
        SELECT COUNT(*) AS total
        FROM ofs_activities_errors
        WHERE `date` BETWEEN %s AND %s
    """, (date_from, date_to))
    total = (cur.fetchone() or {}).get("total", 0)

    # KPI: total com erro NG (dispatch OR response)
    cur.execute("""
        SELECT COUNT(*) AS total_ng
        FROM ofs_activities_errors
        WHERE `date` BETWEEN %s AND %s
          AND (ng_dispatch_message IS NOT NULL OR ng_response_message IS NOT NULL)
    """, (date_from, date_to))
    total_ng = (cur.fetchone() or {}).get("total_ng", 0)

    # Top mensagens (considerando dispatch e response juntos)
    cur.execute("""
        SELECT msg, COUNT(*) AS qtd
        FROM (
            SELECT ng_dispatch_message AS msg
            FROM ofs_activities_errors
            WHERE `date` BETWEEN %s AND %s AND ng_dispatch_message IS NOT NULL
            UNION ALL
            SELECT ng_response_message AS msg
            FROM ofs_activities_errors
            WHERE `date` BETWEEN %s AND %s AND ng_response_message IS NOT NULL
        ) x
        GROUP BY msg
        ORDER BY qtd DESC
        LIMIT 15
    """, (date_from, date_to, date_from, date_to))
    top_messages = cur.fetchall()

    # Erros por dia
    cur.execute("""
        SELECT `date`, COUNT(*) AS qtd
        FROM ofs_activities_errors
        WHERE `date` BETWEEN %s AND %s
          AND (ng_dispatch_message IS NOT NULL OR ng_response_message IS NOT NULL)
        GROUP BY `date`
        ORDER BY `date` DESC
        LIMIT 31
    """, (date_from, date_to))
    by_day = cur.fetchall()

    # Erros por activityType
    cur.execute("""
        SELECT COALESCE(activity_type,'-') AS activityType, COUNT(*) AS qtd
        FROM ofs_activities_errors
        WHERE `date` BETWEEN %s AND %s
          AND (ng_dispatch_message IS NOT NULL OR ng_response_message IS NOT NULL)
        GROUP BY COALESCE(activity_type,'-')
        ORDER BY qtd DESC
        LIMIT 20
    """, (date_from, date_to))
    by_type = cur.fetchall()
    cur.execute("""
    SELECT
        COALESCE(cfg.responsavel, 'Não mapeado') AS responsavel,
        COUNT(*) AS qtd
    FROM (

        SELECT
            'NG_DISPATCH' AS origem,
            TRIM(ng_dispatch_message) AS error_message
        FROM ofs_activities_errors
        WHERE `date` BETWEEN %s AND %s
        AND ng_dispatch_message IS NOT NULL
        AND NULLIF(TRIM(ng_dispatch_message),'') IS NOT NULL

        UNION ALL

        SELECT
            'NG_RESPONSE' AS origem,
            TRIM(ng_response_message) AS error_message
        FROM ofs_activities_errors
        WHERE `date` BETWEEN %s AND %s
        AND ng_response_message IS NOT NULL
        AND NULLIF(TRIM(ng_response_message),'') IS NOT NULL

    ) base

    LEFT JOIN ofs_error_owner_config cfg
        ON cfg.origem = base.origem
    AND TRIM(cfg.error_message) = base.error_message
    AND cfg.ativo = 1

    GROUP BY COALESCE(cfg.responsavel, 'Não mapeado')
    ORDER BY qtd DESC
    """, (date_from, date_to, date_from, date_to))

    by_owner = cur.fetchall()
    cur.close()
    conn.close()

    return render_template(
        "ofs_activities_errors/ofs_activities_errors_dashboard.html",
        total=total,
        total_ng=total_ng,
        top_messages=top_messages,
        by_day=by_day,
        by_type=by_type,
        date_from=date_from,
        date_to=date_to,
        by_owner=by_owner,
        resources=resources
    )
# =====
# Login
# =====
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        # sempre tratar login como e-mail em minúsculas
        username = (request.form.get("username") or "").strip().lower()
        password = (request.form.get("password") or "").strip()

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM usuarios WHERE username = %s", (username,))
        user = cursor.fetchone()
        cursor.close()
        conn.close()

        if user and bcrypt.checkpw(password.encode(), user["password_hash"].encode()):
            # atualiza last_login
            conn = get_connection()
            cur_upd = conn.cursor()
            cur_upd.execute(
                "UPDATE usuarios SET last_login = %s WHERE id = %s",
                (datetime.now(), user["id"]),
            )
            conn.commit()
            cur_upd.close()
            conn.close()

            # guarda dados básicos na sessão
            session["usuario_id"] = user["id"]
            session["usuario_logado"] = user["username"]
            session["nome_usuario"] = user["nome"]
            session["tipo_id"] = int(user["tipo_id"]) if user.get("tipo_id") is not None else 3

            # carrega permissões do perfil
            session["permissoes"] = _carregar_permissoes_por_perfil(session["tipo_id"])

            # AUDIT: login ok
            audit_log(
                actor_user_id=user["id"],
                actor_username=user["username"],
                module="auth",
                action="login",
                entity_type="usuario",
                entity_id=user["id"],
                entity_ref=user["username"],
                summary=f"Login realizado com sucesso: {user['username']}",
                meta={"ip": request.remote_addr, "ua": request.user_agent.string},
            )

            return redirect(url_for("home"))

        # AUDIT: login falho (não registra senha)
        audit_log(
            actor_user_id=None,
            actor_username=username,
            module="auth",
            action="login_failed",
            entity_type="usuario",
            entity_ref=username,
            summary=f"Tentativa de login falhou: {username}",
            meta={"ip": request.remote_addr, "ua": request.user_agent.string},
        )

        flash("Usuário ou senha inválidos.", "danger")

    return render_template("login.html")


# Página Home
@app.route("/")
@login_required
def home():
    # carrega lista de tipos de usuário do OFS (para telas de atualização) 1x por sessão
    if "tipos_user" not in session:
        session["tipos_user"] = get_tipos_user()
    return render_template("home.html")


# ===========================
# Atualizar userType no OFS
# ===========================

@app.route("/atualizar", methods=["GET", "POST"])
@login_required
@perm_required("ofs.atualizar_tipo")
def atualizar_user_type():
    if request.method == "POST":
        resource_id = request.form.get("resource_id")
        new_user_type = request.form.get("user_type")

        client = OFSClient()
        try:
            login = client.get_login_by_resource_id(resource_id)
            status, _ = client.update_user_type(login, new_user_type)
            flash(f"✅ Login {login} atualizado com sucesso! (Status: {status})", "success")
        except Exception as e:
            flash(f"❌ Falha: {e}", "danger")

        return redirect(url_for("atualizar_user_type"))

    return render_template("atualizar_user_type.html")


@app.route("/atualizar-um", methods=["GET", "POST"])
@login_required
@perm_required("ofs.atualizar_tipo")
def atualizar_um():
    tipos_user = session.get("tipos_user", [])

    if request.method == "POST":
        resource_id = request.form.get("resource_id")
        user_type_codigo = request.form.get("user_type")

        session["ultimo_user_type"] = user_type_codigo

        username = os.getenv("OFS_USERNAME")
        password = os.getenv("OFS_PASSWORD")
        client = OFSClient(username, password)

        try:
            login = client.get_login_by_resource_id(resource_id)
            status, _ = client.update_user_type(login, user_type_codigo)
            flash(f"✅ Login {login} atualizado com sucesso! (Status: {status})", "success")

            # AUDIT
            actor = current_actor()
            audit_log(
                actor_user_id=actor.get("id"),
                actor_username=actor.get("username"),
                module="ofs",
                action="update_user_type",
                entity_type="ofs_user",
                entity_ref=str(resource_id),
                summary=f"Atualizou userType no OFS (um): resourceId={resource_id} login={login} userType={user_type_codigo}",
                meta={"resource_id": resource_id, "login": login, "userType": user_type_codigo, "status": status},
            )

        except Exception as e:
            flash(f"❌ Erro ao atualizar o userType: {e}", "danger")

        return redirect(url_for("atualizar_um"))

    selected = session.pop("ultimo_user_type", "")
    return render_template("atualizar_um.html", tipos=tipos_user, selected=selected)


@app.route("/atualizar-varios", methods=["GET", "POST"])
@login_required
@perm_required("ofs.atualizar_tipo")
def atualizar_varios():
    tipos_user = session.get("tipos_user", [])

    if request.method == "POST":
        modo = request.form.get("modo")  # "resourceId" ou "email"
        valores_raw = request.form.get("identificadores", "")
        user_type = request.form.get("user_type")

        valores = [v.strip() for v in valores_raw.split(",") if v.strip()]
        logs = []

        username = os.getenv("OFS_USERNAME")
        password = os.getenv("OFS_PASSWORD")
        client = OFSClient(username, password)

        ok = 0
        fail = 0

        for item in valores:
            try:
                if modo == "email":
                    login = item
                else:
                    login = client.get_login_by_resource_id(item)

                status, _ = client.update_user_type(login, user_type)
                logs.append(f"✅ {item} → {login} atualizado com sucesso (Status: {status})")
                ok += 1
            except Exception as e:
                logs.append(f"❌ {item} → Erro: {e}")
                fail += 1

        # AUDIT: bulk
        actor = current_actor()
        audit_log(
            actor_user_id=actor.get("id"),
            actor_username=actor.get("username"),
            module="ofs",
            action="bulk_update_user_type",
            entity_type="ofs_user",
            summary=f"Atualizou userType em lote: modo={modo}, userType={user_type}, total={len(valores)}, ok={ok}, fail={fail}",
            meta={"modo": modo, "userType": user_type, "total": len(valores), "ok": ok, "fail": fail, "itens": valores[:200]},
        )

        session["log_varios"] = logs
        return redirect(url_for("log_varios"))

    return render_template("atualizar_varios.html", tipos=tipos_user)


@app.route("/log-varios")
@login_required
@perm_required("ofs.atualizar_tipo")
def log_varios():
    logs = session.pop("log_varios", [])
    return render_template("log_varios.html", logs=logs)


# ===========================
# Criar técnicos via CSV
# ===========================

@app.route("/criar-tecnicos", methods=["GET", "POST"])
@login_required
@perm_required("ofs.criar_tecnicos")
def criar_tecnicos():
    logs = []

    if request.method == "POST":
        if "csv_file" not in request.files or request.files["csv_file"].filename == "":
            flash("Envie um arquivo CSV válido.", "danger")
            return render_template("criar_tecnicos.html", logs=logs)

        file = request.files["csv_file"]
        try:
            data = file.read().decode("utf-8-sig")  # trata BOM
        except Exception:
            flash("Falha ao ler o CSV. Verifique se está em UTF-8.", "danger")
            return render_template("criar_tecnicos.html", logs=logs)

        reader = csv.DictReader(StringIO(data))
        expected = [
            "idSAP",
            "depositoTecnico",
            "tipoDeRecurso",
            "nomeCompleto",
            "areaDoTecnico",
            "tipoDeUsuario",
            "email",
            "Senha",
        ]
        missing = [h for h in expected if h not in reader.fieldnames]
        if missing:
            flash(f"Cabeçalhos ausentes no CSV: {', '.join(missing)}", "danger")
            return render_template("criar_tecnicos.html", logs=logs)

        client = OFSClient()

        linha = 1  # +1 do header para exibir ao usuário
        for row in reader:
            linha += 1
            id_sap = (row.get("idSAP") or "").strip()
            deposito_tecnico = (row.get("depositoTecnico") or "").strip()
            # tipoDeRecurso está no CSV, mas hoje usamos fixo "TCV" na API de criação
            nome_completo = (row.get("nomeCompleto") or "").strip()
            area_tecnico = (row.get("areaDoTecnico") or "").strip()
            tipo_usuario = (row.get("tipoDeUsuario") or "").strip()
            email = (row.get("email") or "").strip()
            senha = (row.get("Senha") or "").strip()

            msg_parts = []
            rec_status = "-"
            usr_status = "-"
            dep_status = "-"

            # validações mínimas
            if not id_sap or not nome_completo or not area_tecnico or not tipo_usuario or not email or not senha:
                logs.append({
                    "linha": linha, "idSAP": id_sap, "email": email,
                    "recurso_status": rec_status, "usuario_status": usr_status,
                    "deposito_status": dep_status,
                    "msg": "Dados obrigatórios ausentes na linha."
                })
                continue

            # 1) cria recurso (PUT)
            try:
                r1 = client.create_resource(
                    id_sap=id_sap,
                    parent_resource_id=area_tecnico,
                    name=nome_completo,
                    email=email
                )
                rec_status = f"{r1.status_code}"
                r1_text = (r1.text or "") if hasattr(r1, "text") else ""
                if r1.status_code in (200, 201):
                    msg_parts.append("Recurso criado/atualizado com sucesso.")
                elif r1.status_code == 409:
                    msg_parts.append("Recurso já existia (409).")
                else:
                    msg_parts.append(f"Falha ao criar recurso: {r1.status_code} {r1_text}")

            except Exception as e:
                msg_parts.append(f"Exceção na criação do recurso: {e}")

            # 2) cria usuário (PUT) — se recurso ok ou já existia
            if rec_status in ("200", "201", "409"):
                try:
                    r2 = client.create_user(
                        email=email,
                        name=nome_completo,
                        id_sap=id_sap,
                        user_type=tipo_usuario,
                        password=senha
                    )
                    usr_status = f"{r2.status_code}"
                    r2_text = (r2.text or "") if hasattr(r2, "text") else ""
                    if r2.status_code in (200, 201):
                        msg_parts.append("Usuário criado/atualizado com sucesso.")
                    elif r2.status_code == 409:
                        msg_parts.append("Usuário já existia (409).")
                    else:
                        msg_parts.append(f"Falha ao criar usuário: {r2.status_code} {r2_text}")
                except Exception as e:
                    msg_parts.append(f"Exceção na criação do usuário: {e}")
            else:
                msg_parts.append("Usuário não criado pois o recurso não foi criado.")

            # 3) atualiza depósito (PATCH no recurso)
            if rec_status in ("200", "201", "409"):
                if deposito_tecnico:
                    try:
                        r3 = client.update_resource_deposito(id_sap=id_sap, deposito_tecnico=deposito_tecnico)
                        dep_status = f"{r3.status_code}"
                        r3_text = (r3.text or "") if hasattr(r3, "text") else ""
                        if r3.status_code in (200, 204):
                            msg_parts.append("Depósito atualizado com sucesso.")
                        else:
                            msg_parts.append(f"Falha ao atualizar depósito: {r3.status_code} {r3_text}")
                    except Exception as e:
                        msg_parts.append(f"Exceção no update do depósito: {e}")
                else:
                    dep_status = "-"
                    msg_parts.append("Depósito não enviado (campo vazio).")

            logs.append({
                "linha": linha,
                "idSAP": id_sap,
                "email": email,
                "recurso_status": rec_status,
                "usuario_status": usr_status,
                "deposito_status": dep_status,
                "msg": "\n".join(msg_parts)
            })

        flash(f"Processamento concluído. Linhas processadas: {len(logs)}", "success")
        actor = current_actor()
        audit_log(
            actor_user_id=actor.get("id"),
            actor_username=actor.get("username"),
            module="ofs",
            action="create_tecnicos_csv",
            entity_type="ofs_user",
            summary=f"Criou técnicos via CSV: linhas={len(logs)}",
            meta={"linhas": len(logs)},
        )

    return render_template("criar_tecnicos.html", logs=logs)


# ===========================
# Trocar a própria senha
# ===========================

@app.route("/trocar-senha", methods=["GET", "POST"])
@login_required
@perm_required("usuarios.trocar_senha")
def trocar_senha():
    if request.method == "POST":
        senha_atual = (request.form.get("senha_atual") or "").strip()
        nova_senha = (request.form.get("nova_senha") or "").strip()
        confirmar = (request.form.get("confirmar_senha") or "").strip()

        if not senha_atual or not nova_senha or not confirmar:
            flash("Preencha todos os campos.", "danger")
            return redirect(url_for("trocar_senha"))

        if len(nova_senha) < 8:
            flash("A nova senha deve ter pelo menos 8 caracteres.", "danger")
            return redirect(url_for("trocar_senha"))

        if nova_senha != confirmar:
            flash("A confirmação não confere com a nova senha.", "danger")
            return redirect(url_for("trocar_senha"))

        username = session.get("usuario_logado")

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, password_hash FROM usuarios WHERE username = %s", (username,))
        user = cursor.fetchone()

        if not user:
            cursor.close(); conn.close()
            flash("Usuário não encontrado.", "danger")
            return redirect(url_for("trocar_senha"))

        if not bcrypt.checkpw(senha_atual.encode(), user["password_hash"].encode()):
            cursor.close(); conn.close()
            flash("Senha atual incorreta.", "danger")
            return redirect(url_for("trocar_senha"))

        novo_hash = bcrypt.hashpw(nova_senha.encode(), bcrypt.gensalt()).decode()
        cursor.execute("UPDATE usuarios SET password_hash = %s WHERE id = %s", (novo_hash, user["id"]))
        conn.commit()
        cursor.close(); conn.close()

        # AUDIT: troca de senha (sem logar a senha)
        actor = current_actor()
        audit_log(
            actor_user_id=actor.get("id"),
            actor_username=actor.get("username"),
            module="usuarios",
            action="change_password",
            entity_type="usuario",
            entity_id=user["id"],
            entity_ref=actor.get("username"),
            summary=f"Trocou a própria senha: {actor.get('username')}",
            meta={"ip": request.remote_addr},
        )

        flash("Senha alterada com sucesso!", "success")
        return redirect(url_for("home"))

    return render_template("trocar_senha.html")


# ===========================
# Criar usuário do painel
# ===========================

@app.route("/criar-usuario", methods=["GET", "POST"])
@login_required
@perm_required("usuarios.criar")
def criar_usuario():
    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip()

        local = (request.form.get("username_local") or "").strip().lower()
        dominio = "@verointernet.com.br"
        username = local + dominio  # monta o e-mail final

        senha = (request.form.get("senha") or "").strip()
        confirmar = (request.form.get("confirmar") or "").strip()
        tipo_id_raw = (request.form.get("tipo_id") or "").strip()

        if not nome or not local or not senha or not confirmar or not tipo_id_raw:
            flash("Preencha todos os campos.", "danger")
            return redirect(url_for("criar_usuario"))

        # valida apenas a parte local
        if not local.replace(".", "").replace("_", "").replace("-", "").isalnum():
            flash("A parte inicial do e-mail contém caracteres inválidos.", "danger")
            return redirect(url_for("criar_usuario"))

        if len(senha) < 8:
            flash("A senha deve ter pelo menos 8 caracteres.", "danger")
            return redirect(url_for("criar_usuario"))

        if senha != confirmar:
            flash("A confirmação não confere com a senha.", "danger")
            return redirect(url_for("criar_usuario"))

        try:
            tipo_id = int(tipo_id_raw)
        except ValueError:
            flash("Perfil inválido.", "danger")
            return redirect(url_for("criar_usuario"))

        conn = get_connection()
        cur = conn.cursor(dictionary=True)

        # confirma se o perfil existe
        cur.execute("SELECT id, nome FROM perfis WHERE id = %s", (tipo_id,))
        perfil_row = cur.fetchone()
        if not perfil_row:
            cur.close(); conn.close()
            flash("Perfil informado não existe.", "danger")
            return redirect(url_for("criar_usuario"))

        cur.execute("SELECT id FROM usuarios WHERE username = %s", (username,))
        if cur.fetchone():
            cur.close(); conn.close()
            flash("Já existe um usuário com esse e-mail.", "danger")
            return redirect(url_for("criar_usuario"))

        password_hash = bcrypt.hashpw(senha.encode(), bcrypt.gensalt()).decode()
        cur.execute(
            "INSERT INTO usuarios (nome, username, password_hash, tipo_id) VALUES (%s, %s, %s, %s)",
            (nome, username, password_hash, tipo_id)
        )
        conn.commit()
        cur.close(); conn.close()

        # AUDIT: criação de usuário do painel (não loga senha)
        actor = current_actor()
        audit_log(
            actor_user_id=actor.get("id"),
            actor_username=actor.get("username"),
            module="usuarios",
            action="create",
            entity_type="usuario",
            entity_ref=username,
            summary=f"Criou usuário do painel: {username}",
            after={"nome": nome, "username": username, "tipo_id": tipo_id, "perfil": perfil_row.get("nome")},
        )

        flash("Usuário criado com sucesso!", "success")
        return redirect(url_for("criar_usuario"))

    perfis = get_perfis()
    return render_template("criar_usuario.html", perfis=perfis)


# ===========================
# Consultar usuários OFS
# ===========================

@app.route("/consultar-usuarios")
@login_required
@perm_required("ofs.consultar")
def consultar_usuarios():
    client = OFSClient()
    usuarios_raw = client.get_usuarios()

    bucket_cache = {}
    usuarios = []
    for u in usuarios_raw:
        main_res = u.get("mainResourceId") or u.get("main_resource_id")
        if main_res:
            if main_res in bucket_cache:
                bucket = bucket_cache[main_res]
            else:
                try:
                    bucket = client.get_bucket_by_resource_id(main_res)
                except Exception:
                    bucket = "-"
                bucket_cache[main_res] = bucket
        else:
            bucket = "-"

        usuarios.append({
            "name": u.get("name", "-"),
            "userType": u.get("userType", "-"),
            "bucket": bucket,
            "code_sap": u.get("XU_CODE_SAP", "-"),
            "status": u.get("status", "-"),
            "login": u.get("login", "-"),
            "lastLoginTime": u.get("lastLoginTime", "-"),
        })

    ativos = sum(1 for u in usuarios if u["status"] == "active")

    return render_template(
        "consultar_usuarios.html",
        usuarios=usuarios,
        total_ativos=ativos
    )


# ===========================
# Desativar inativos / sem login
# ===========================

@app.route("/desativar_inativos", methods=["GET", "POST"])
@login_required
@perm_required("ofs.desativar")
def desativar_inativos():
    raw_days = (request.values.get("cutoff_days") or "80").strip()
    cutoff_days = int(raw_days) if raw_days.isdigit() else 80
    only_active = request.values.get("only_active") is not None

    vencidos, meta = find_stale_users(cutoff_days=cutoff_days, only_active=only_active)

    results = []
    mode = "SIMULACAO"
    if request.method == "POST":
        apply = request.form.get("apply_changes") == "1"
        results = execute_cleanup(vencidos, apply_changes=apply)
        mode = "APLICACAO" if apply else "SIMULACAO"
        flash(f"{'Aplicado' if apply else 'Simulado'} para {len(vencidos)} usuários.", "success")

        # AUDIT: cleanup
        actor = current_actor()
        audit_log(
            actor_user_id=actor.get("id"),
            actor_username=actor.get("username"),
            module="ofs",
            action="cleanup" if apply else "cleanup_simulation",
            entity_type="ofs_user",
            summary=f"Cleanup OFS: mode={mode}, cutoff_days={cutoff_days}, only_active={only_active}, total={len(vencidos)}",
            meta={"mode": mode, "cutoff_days": cutoff_days, "only_active": only_active, "total": len(vencidos)},
        )

    if request.values.get("export") == "1":
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = f"/tmp/users_vencidos_{stamp}.csv"
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["login", "status", "lastLoginTime", "userType", "mainResourceId"])
            for u in vencidos:
                w.writerow([u.get("login"), u.get("status"), u.get("lastLoginTime"), u.get("userType"), u.get("mainResourceId")])
        return send_file(path, as_attachment=True, download_name=os.path.basename(path), mimetype="text/csv")

    return render_template(
        "desativar_inativos.html",
        cutoff_days=cutoff_days,
        only_active=only_active,
        vencidos=vencidos,
        results=results,
        mode=mode,
        meta=meta,
    )


# ===========================
# Gestão de perfis e permissões
# ===========================

# Gerenciar perfis e permissões
@app.route("/perfis", methods=["GET", "POST"])
@login_required
@perm_required("perfis.gerenciar")
def perfis_view():
    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    # ---------- POST: criar, salvar, deletar ----------
    if request.method == "POST":
        acao = request.form.get("acao")
        perfil_id_raw = (request.form.get("perfil_id") or "").strip()

        # --- CRIAR PERFIL ---
        if acao == "criar":
            novo_nome = (request.form.get("novo_perfil") or "").strip()
            if not novo_nome:
                flash("Informe um nome para o novo perfil.", "danger")
                cur.close(); conn.close()
                return redirect(url_for("perfis_view"))

            slug = novo_nome.lower().strip().replace(" ", "_")

            # Gera próximo ID manualmente (MAX(id) + 1)
            cur.execute("SELECT COALESCE(MAX(id), 0) + 1 AS prox_id FROM perfis")
            row = cur.fetchone()
            prox_id = row["prox_id"] if row and "prox_id" in row else 1

            cur.execute(
                "INSERT INTO perfis (id, nome, slug) VALUES (%s, %s, %s)",
                (prox_id, novo_nome, slug),
            )
            conn.commit()

            # AUDIT: criar perfil
            actor = current_actor()
            audit_log(
                actor_user_id=actor.get("id"),
                actor_username=actor.get("username"),
                module="perfis",
                action="create",
                entity_type="perfil",
                entity_id=prox_id,
                entity_ref=slug,
                summary=f"Criou perfil: {novo_nome}",
                after={"id": prox_id, "nome": novo_nome, "slug": slug},
            )

            flash("Perfil criado com sucesso.", "success")
            cur.close(); conn.close()
            return redirect(url_for("perfis_view"))

        # --- SALVAR ALTERAÇÕES DE PERFIL ---
        if acao == "salvar" and perfil_id_raw:
            try:
                perfil_id = int(perfil_id_raw)
            except ValueError:
                flash("Perfil inválido.", "danger")
                cur.close(); conn.close()
                return redirect(url_for("perfis_view"))

            nome_editado = (request.form.get("nome_perfil") or "").strip()
            ids_permissoes = request.form.getlist("permissoes[]")

            if not nome_editado:
                flash("O nome do perfil não pode ser vazio.", "danger")
                cur.close(); conn.close()
                return redirect(url_for("perfis_view", perfil_id=perfil_id))

            # BEFORE: perfil + permissões
            cur.execute("SELECT id, nome, slug FROM perfis WHERE id = %s", (perfil_id,))
            before_perfil = cur.fetchone()

            cur.execute("""
                SELECT p.recurso
                FROM perfil_permissao pp
                JOIN permissoes p ON p.id = pp.permissao_id
                WHERE pp.perfil_id = %s
                ORDER BY p.recurso
            """, (perfil_id,))
            before_perms = [r["recurso"] for r in cur.fetchall()]

            # update nome
            cur.execute(
                "UPDATE perfis SET nome = %s WHERE id = %s",
                (nome_editado, perfil_id),
            )

            # Atualiza permissões
            cur.execute("DELETE FROM perfil_permissao WHERE perfil_id = %s", (perfil_id,))
            for pid in ids_permissoes:
                try:
                    pid_int = int(pid)
                    cur.execute(
                        "INSERT INTO perfil_permissao (perfil_id, permissao_id) VALUES (%s, %s)",
                        (perfil_id, pid_int),
                    )
                except ValueError:
                    continue

            conn.commit()

            # AFTER: permissões atualizadas
            cur.execute("""
                SELECT p.recurso
                FROM perfil_permissao pp
                JOIN permissoes p ON p.id = pp.permissao_id
                WHERE pp.perfil_id = %s
                ORDER BY p.recurso
            """, (perfil_id,))
            after_perms = [r["recurso"] for r in cur.fetchall()]

            # AUDIT: update perfil
            actor = current_actor()
            audit_log(
                actor_user_id=actor.get("id"),
                actor_username=actor.get("username"),
                module="perfis",
                action="update",
                entity_type="perfil",
                entity_id=perfil_id,
                entity_ref=(before_perfil or {}).get("slug"),
                summary=f"Atualizou perfil: {nome_editado}",
                before={"perfil": before_perfil, "permissoes": before_perms},
                after={"perfil": {"id": perfil_id, "nome": nome_editado, "slug": (before_perfil or {}).get("slug")},
                       "permissoes": after_perms},
            )

            flash("Perfil atualizado com sucesso.", "success")
            cur.close(); conn.close()
            return redirect(url_for("perfis_view", perfil_id=perfil_id))

        # --- DELETAR PERFIL ---
        if acao == "deletar" and perfil_id_raw:
            try:
                perfil_id = int(perfil_id_raw)
            except ValueError:
                flash("Perfil inválido.", "danger")
                cur.close(); conn.close()
                return redirect(url_for("perfis_view"))

            # BEFORE: perfil (para audit)
            cur.execute("SELECT id, nome, slug FROM perfis WHERE id = %s", (perfil_id,))
            perfil_row = cur.fetchone()

            # Verifica se há usuários usando esse perfil
            cur.execute(
                "SELECT COUNT(*) AS total FROM usuarios WHERE tipo_id = %s",
                (perfil_id,),
            )
            qtd_usuarios = cur.fetchone()["total"]

            if qtd_usuarios > 0:
                flash(
                    f"Não é possível apagar: existem {qtd_usuarios} usuário(s) usando este perfil.",
                    "danger",
                )
                cur.close(); conn.close()
                return redirect(url_for("perfis_view", perfil_id=perfil_id))

            # Remove vínculos antes de apagar
            cur.execute("DELETE FROM perfil_permissao WHERE perfil_id = %s", (perfil_id,))
            cur.execute("DELETE FROM perfis WHERE id = %s", (perfil_id,))
            conn.commit()

            # AUDIT: delete perfil
            actor = current_actor()
            audit_log(
                actor_user_id=actor.get("id"),
                actor_username=actor.get("username"),
                module="perfis",
                action="delete",
                entity_type="perfil",
                entity_id=perfil_id,
                entity_ref=(perfil_row or {}).get("slug"),
                summary=f"Removeu perfil: {(perfil_row or {}).get('nome', perfil_id)}",
                before={"perfil": perfil_row},
            )

            flash("Perfil removido com sucesso.", "success")
            cur.close(); conn.close()
            return redirect(url_for("perfis_view"))

    # ---------- GET: Carregamento da tela ----------
    cur.execute("SELECT id, nome, slug FROM perfis ORDER BY nome")
    perfis = cur.fetchall()

    # Seleção do perfil atual via GET
    perfil_id = request.args.get("perfil_id", type=int)
    if not perfil_id and perfis:
        perfil_id = perfis[0]["id"]

    # Perfil selecionado
    perfil_atual = None
    if perfil_id:
        for p in perfis:
            if p["id"] == perfil_id:
                perfil_atual = p
                break

    # Carrega todas as permissões existentes
    cur.execute("SELECT id, recurso, descricao FROM permissoes ORDER BY recurso")
    permissoes = cur.fetchall()

    # Carrega permissões do perfil selecionado
    perfil_permissoes = set()
    user_count = 0

    if perfil_atual:
        cur.execute(
            "SELECT permissao_id FROM perfil_permissao WHERE perfil_id = %s",
            (perfil_atual["id"],),
        )
        perfil_permissoes = {row["permissao_id"] for row in cur.fetchall()}

        cur.execute(
            "SELECT COUNT(*) AS total FROM usuarios WHERE tipo_id = %s",
            (perfil_atual["id"],),
        )
        user_count = cur.fetchone()["total"]

    cur.close()
    conn.close()

    return render_template(
        "perfis.html",
        perfis=perfis,
        perfil_atual=perfil_atual,
        permissoes=permissoes,
        perfil_permissoes=perfil_permissoes,
        user_count=user_count,
    )


@app.route("/fechar-os-adapter", methods=["GET", "POST"])
@login_required
@perm_required("adapter.fechar_os")
def fechar_os_adapter():
    """
    Fluxo:
      - GET: exibe formulário
      - POST acao=preview: faz GET Activity + GET Resource e mostra preview (sem enviar)
      - POST acao=confirmar: envia POST para o Adapter usando dados do preview e grava log
    """
    acao = (request.form.get("acao") or "").strip().lower()
    activity_id = (request.form.get("activity_id") or "").strip()

    # Preview fica na sessão para evitar adulteração via form
    preview = session.get("adapter_preview")

    def _get_usuario_id_logado():
        # Preferência: se você já tem session["usuario_id"] no projeto
        uid = session.get("usuario_id")
        if uid:
            return int(uid)

        # Fallback: buscar por username
        conn = get_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT id FROM usuarios WHERE username = %s", (session.get("usuario_logado"),))
        row = cur.fetchone()
        cur.close()
        conn.close()
        return int(row["id"]) if row else 0

    def _log_fechamento(**kwargs):
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO adapter_fechamento_os_log
            (usuario_id, activity_id, resource_id, cod_atendimento, id_fechamento,
             payload_json, response_status, response_body, error_message)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            kwargs.get("usuario_id"),
            kwargs.get("activity_id"),
            kwargs.get("resource_id"),
            kwargs.get("cod_atendimento"),
            kwargs.get("id_fechamento"),
            kwargs.get("payload_json"),
            kwargs.get("response_status"),
            kwargs.get("response_body"),
            kwargs.get("error_message"),
        ))
        conn.commit()
        cur.close()
        conn.close()

    def _to_int_or_keep(v):
        """Converte para int se for numérico; caso contrário mantém."""
        if v is None:
            return None
        s = str(v).strip()
        return int(s) if s.isdigit() else v

    # -------- GET: tela vazia --------
    if request.method == "GET":
        # não mantém preview antigo ao entrar na página
        session.pop("adapter_preview", None)
        return render_template("fechar_os_adapter.html", stage="form", activity_id="")

    # -------- POST: preview --------
    if acao == "preview":
        if not activity_id:
            flash("Informe o ID da atividade OFS.", "danger")
            return render_template("fechar_os_adapter.html", stage="form", activity_id="")

        try:
            client = OFSClient()

            # 1) Get Activity
            atividade = client.authenticated_get(f"{client.base_url}/activities/{activity_id}")

            resource_id = atividade.get("resourceId")
            if not resource_id:
                raise ValueError("A atividade não possui resourceId.")

            cod_atendimento = atividade.get("XA_SOL_ID")
            start_time = atividade.get("startTime")  # conforme você disse: já vem "AAAA-MM-DD HH:MM:SS"
            obs = atividade.get("XA_TSK_NOT")
            id_fechamento = atividade.get("XA_SER_CLO_PRO_ADA") or atividade.get("XA_SER_CLO_IMP_ADA")

            if not cod_atendimento:
                raise ValueError("Atividade sem XA_SOL_ID (CodAtendimento).")
            if not id_fechamento:
                raise ValueError("Atividade sem XA_SER_CLO_PRO_ADA e sem XA_SER_CLO_IMP_ADA.")
            if not start_time:
                raise ValueError("Atividade sem startTime (DataInicioAtendimento).")

            # 2) Get Resource
            recurso = client.authenticated_get(f"{client.base_url}/resources/{resource_id}")
            resource_name = recurso.get("name")
            xr_user = recurso.get("XR_USER_ADAPTER")
            xr_pass = recurso.get("XR_PASSWORD_ADAPTER")

            if not resource_name:
                resource_name = "Recurso sem nome"

            if not xr_user or not xr_pass:
                raise ValueError("Recurso sem XR_USER_ADAPTER ou XR_PASSWORD_ADAPTER.")

            # 3) Monta payload (não envia ainda) — modelo atualizado
            payload = {
                "usuario": xr_user,
                "senha": xr_pass,
                "DadosFechamento": {
                    "CodAtendimento": str(cod_atendimento),

                    # novos campos do modelo do Adapter
                    "WifiUsuario": "NULL",
                    "WifiSenha": "NULL",

                    # mantém como vem do OFS
                    "DataInicioAtendimento": str(start_time),
                    "IDFechamento": str(id_fechamento),

                    # novos campos do modelo do Adapter (null real)
                    "MACONU": None,
                    "IDSaidaCaixaEscolhida": None,

                    # mantém
                    "IDInterface": None,

                    # correção: null real (não string "null")
                    "JustificativaReagendamento": None,
                    "IDMotivoReagendamento": None,

                    "ObsFechamento": obs,
                    "obsFechamentoLog": "NULL",

                    # no exemplo é numérico; converte se der
                    "CodTecnico": _to_int_or_keep(resource_id),

                    "MovimentouEquipamento": True,
                    "MovimentouMaterial": True,
                    "MovimentouEquipamentoCliente": True
                }
            }

            preview = {
                "activity_id": activity_id,
                "resource_id": str(resource_id),
                "resource_name": resource_name,
                "xr_user": xr_user,
                "xr_pass": xr_pass,
                "cod_atendimento": str(cod_atendimento),
                "start_time": str(start_time),
                "id_fechamento": str(id_fechamento),
                "obs": obs,
                "payload": payload,  # fica na sessão
            }
            session["adapter_preview"] = preview

            return render_template("fechar_os_adapter.html", stage="preview", preview=preview, activity_id=activity_id)

        except Exception as e:
            session.pop("adapter_preview", None)
            flash(f"Erro ao montar preview: {e}", "danger")
            return render_template("fechar_os_adapter.html", stage="form", activity_id=activity_id)

    # -------- POST: confirmar envio --------
    if acao == "confirmar":
        if not preview:
            flash("Preview expirado. Gere o preview novamente.", "danger")
            return render_template("fechar_os_adapter.html", stage="form", activity_id="")

        try:
            close_url = os.getenv("URL_CLOSE_ADAPTER")
            auth_ada = os.getenv("AUTH_ADA")     # valor depois de "Basic "
            cookie_ada = os.getenv("COOKIE_ADA")

            if not close_url:
                raise RuntimeError("URL_CLOSE_ADAPTER não configurado no .env.")
            if not auth_ada:
                raise RuntimeError("AUTH_ADA não configurado no .env.")
            if not cookie_ada:
                raise RuntimeError("COOKIE_ADA não configurado no .env.")

            headers = {
                "Content-Type": "application/json",
                "Accept": "application/json",
                "Authorization": f"Basic {auth_ada}",
                "Cookie": cookie_ada,
            }

            payload = preview["payload"]
            resp = requests.post(close_url, json=payload, headers=headers, timeout=30)
            try:
                api_response = resp.json()
            except Exception:
                api_response = {"raw": (resp.text or "")}

            usuario_id = _get_usuario_id_logado()

            _log_fechamento(
                usuario_id=usuario_id,
                activity_id=preview.get("activity_id"),
                resource_id=preview.get("resource_id"),
                cod_atendimento=preview.get("cod_atendimento"),
                id_fechamento=preview.get("id_fechamento"),
                payload_json=json.dumps(payload, ensure_ascii=False),
                response_status=resp.status_code,
                response_body=(resp.text or "")[:65000],
                error_message=None,
            )

            actor = current_actor()
            audit_log(
                actor_user_id=actor.get("id"),
                actor_username=actor.get("username"),
                module="adapter",
                action="close_os",
                entity_type="activity",
                entity_ref=preview.get("activity_id"),
                summary=f"Fechou OS via Adapter: activityId={preview.get('activity_id')} HTTP={resp.status_code}",
                meta={
                    "activity_id": preview.get("activity_id"),
                    "resource_id": preview.get("resource_id"),
                    "resource_name": preview.get("resource_name"),
                    "cod_atendimento": preview.get("cod_atendimento"),
                    "id_fechamento": preview.get("id_fechamento"),
                    "status_code": resp.status_code,
                },
                api_response=api_response,
            )

            # limpa preview após tentativa
            session.pop("adapter_preview", None)

            if 200 <= resp.status_code < 300:
                flash("Fechamento enviado com sucesso para o Adapter.", "success")
            else:
                flash(f"Adapter retornou erro HTTP {resp.status_code}.", "danger")

            return render_template(
                "fechar_os_adapter.html",
                stage="result",
                result={"status_code": resp.status_code, "body": (resp.text or "")[:5000]},
                activity_id=preview.get("activity_id"),
            )

        except Exception as e:
            # tenta logar erro também
            try:
                usuario_id = _get_usuario_id_logado()
                _log_fechamento(
                    usuario_id=usuario_id,
                    activity_id=(preview or {}).get("activity_id") or activity_id,
                    resource_id=(preview or {}).get("resource_id"),
                    cod_atendimento=(preview or {}).get("cod_atendimento"),
                    id_fechamento=(preview or {}).get("id_fechamento"),
                    payload_json=json.dumps((preview or {}).get("payload") or {}, ensure_ascii=False),
                    response_status=None,
                    response_body=None,
                    error_message=str(e),
                )
            except Exception:
                pass

            flash(f"Erro ao enviar fechamento: {e}", "danger")
            return render_template(
                "fechar_os_adapter.html",
                stage="preview",
                preview=preview,
                activity_id=(preview or {}).get("activity_id") or activity_id,
            )

    # fallback
    flash("Ação inválida.", "danger")
    return redirect(url_for("fechar_os_adapter"))

# ===========================
# Atividades notdone - DB + Import + Tratativa
# ===========================

@app.route("/atividades-notdone", methods=["GET"])
@login_required
@perm_required("ofs.atividades_notdone")
def atividades_notdone():
    today = datetime.now().strftime("%Y-%m-%d")
    date_from = (request.args.get("dateFrom") or today).strip()
    date_to = (request.args.get("dateTo") or today).strip()
    resources = (request.args.get("resources") or "MG").strip()

    # Lista SEMPRE do banco
    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute("""
        SELECT
            activity_id AS activityId,
            city,
            customer_number AS customerNumber,
            customer_name AS customerName,
            appt_number AS apptNumber,
            origin_bucket AS XA_ORIGIN_BUCKET,
            tsk_not AS XA_TSK_NOT,
            ser_clo_imp_ada AS XA_SER_CLO_IMP_ADA,
            resource_id AS resourceId,
            date AS date,
            tratativa_status,
            tratativa_obs,
            tratado_por_username,
            tratado_em
        FROM ofs_atividades_notdone
        ORDER BY
            (tratado_em IS NULL) DESC,   -- não tratados primeiro
            created_at DESC
        LIMIT 5000
    """)
    items = cur.fetchall()
    cur.close()
    conn.close()

    total = len(items)
    tratados = sum(1 for i in items if i.get("tratado_em"))
    pendentes = total - tratados

    return render_template(
        "atividades_notdone.html",
        items=items,
        date_from=date_from,
        date_to=date_to,
        resources=resources,
        total=total,
        tratados=tratados,
        pendentes=pendentes,
    )


@app.route("/atividades-notdone/importar", methods=["POST"])
@login_required
@perm_required("ofs.atividades_notdone")
def atividades_notdone_importar():
    """
    Chama a API SOMENTE via botão e insere apenas os novos (activityId UNIQUE + INSERT IGNORE).
    """
    today = datetime.now().strftime("%Y-%m-%d")
    date_from = (request.form.get("dateFrom") or today).strip()
    date_to = (request.form.get("dateTo") or today).strip()
    resources = (request.form.get("resources") or "MG").strip()

    client = OFSClient()

    fields = [
        "activityId",
        "city",
        "customerNumber",
        "customerName",
        "customerPhone",
        "apptNumber",
        "XA_ORIGIN_BUCKET",
        "XA_TSK_NOT",
        "XA_SER_CLO_IMP_ADA",
        "resourceId",
        "date",
    ]

    base_params = {
        "dateFrom": date_from,
        "dateTo": date_to,
        "resources": resources,
        "q": "status=='notdone'",
        "fields": ",".join(fields),
        "limit": 2000,
        "offset": 0,
    }

    items = []
    has_more = True
    max_pages = 20
    page = 0

    try:
        while has_more and page < max_pages:
            qs = urlencode(base_params, safe="=,'")
            url = f"{client.base_url}/activities/?{qs}"
            data = client.authenticated_get(url)

            batch = data.get("items") or []
            items.extend(batch)

            has_more = bool(data.get("hasMore"))
            if has_more:
                base_params["offset"] = len(items)
            page += 1
    except Exception as e:
        flash(f"❌ Falha ao importar da API: {e}", "danger")
        return redirect(url_for("atividades_notdone", dateFrom=date_from, dateTo=date_to, resources=resources))

    # Insere no DB apenas os novos
    conn = get_connection()
    cur = conn.cursor()

    inserted = 0
    skipped = 0

    sql = """
        INSERT IGNORE INTO ofs_atividades_notdone
        (activity_id, city, customer_number, customer_phone, customer_name, appt_number, origin_bucket, tsk_not, ser_clo_imp_ada, resource_id,date)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """

    for a in items:
        activity_id = str(a.get("activityId") or "").strip()
        if not activity_id:
            continue

        cur.execute(sql, (
            activity_id,
            str(a.get("city") or "") or None,
            str(a.get("customerNumber") or "") or None,
            str(a.get("customerPhone") or "") or None,
            str(a.get("customerName") or "") or None,
            str(a.get("apptNumber") or "") or None,
            str(a.get("XA_ORIGIN_BUCKET") or "") or None,
            a.get("XA_TSK_NOT"),  # pode ser texto grande (já é str normalmente)
            str(a.get("XA_SER_CLO_IMP_ADA") or "") or None,
            str(a.get("resourceId") or "") or None,
            str(a.get("date") or "") or None,
        ))
        if cur.rowcount == 1:
            inserted += 1
        else:
            skipped += 1

    conn.commit()
    cur.close()
    conn.close()

    flash(f"✅ Importação concluída. Novos: {inserted} | Já existiam: {skipped}", "success")

    return redirect(url_for("atividades_notdone", dateFrom=date_from, dateTo=date_to, resources=resources))


@app.route("/atividades-notdone/tratar", methods=["POST"])
@login_required
@perm_required("ofs.atividades_notdone")
def atividades_notdone_tratar():
    """
    Salva tratativa de forma ATÔMICA:
    - Só trata se tratado_em IS NULL
    - Se já tratado, retorna 409
    - Registra histórico
    """
    data = request.get_json(silent=True) or {}

    activity_id = str(data.get("activityId") or "").strip()
    status = (data.get("status") or "").strip()
    obs = (data.get("observacoes") or "").strip()

    allowed = {
        "Reagendado", 
        "Sem contato", 
        "Reagendado sem contato",
        "Visita cancelada",
        "Aberto Lecom (Crescimento Organico)"
        }
    if not activity_id:
        return jsonify({"ok": False, "error": "activityId obrigatório"}), 400
    if status not in allowed:
        return jsonify({"ok": False, "error": "status inválido"}), 400

    actor = current_actor()
    user_id = actor.get("id")
    username = actor.get("username")

    conn = get_connection()
    cur = conn.cursor()

    try:
        # UPDATE ATÔMICO: só trata se ainda não estiver tratado
        update_sql = """
            UPDATE ofs_atividades_notdone
            SET
                tratativa_status = %s,
                tratativa_obs = %s,
                tratado_por_user_id = %s,
                tratado_por_username = %s,
                tratado_em = NOW()
            WHERE activity_id = %s
              AND tratado_em IS NULL
        """
        cur.execute(update_sql, (
            status,
            obs if obs else None,
            int(user_id) if user_id else None,
            username,
            activity_id
        ))

        if cur.rowcount == 0:
            conn.rollback()
            return jsonify({
                "ok": False,
                "error": "Esta atividade já foi tratada por outro usuário (ou não existe no banco)."
            }), 409

        # HISTÓRICO
        cur.execute("""
            INSERT INTO ofs_atividades_notdone_history
            (activity_id, action, status, obs, actor_user_id, actor_username)
            VALUES (%s, 'TRATAR', %s, %s, %s, %s)
        """, (
            activity_id,
            status,
            obs if obs else None,
            int(user_id) if user_id else None,
            username
        ))

        conn.commit()

    except Exception as e:
        conn.rollback()
        return jsonify({"ok": False, "error": f"Erro ao salvar tratativa: {e}"}), 500

    finally:
        cur.close()
        conn.close()

    # AUDIT (opcional)
    try:
        audit_log(
            actor_user_id=user_id,
            actor_username=username,
            module="ofs",
            action="tratativa_notdone",
            entity_type="activity",
            entity_ref=activity_id,
            summary=f"Tratou atividade notdone: activityId={activity_id} status={status}",
            meta={"status": status},
        )
    except Exception:
        pass

    return jsonify({
        "ok": True,
        "activityId": activity_id,
        "tratadoPor": username,
        "status": status
    }), 200

@app.route("/atividades-notdone/<activity_id>", methods=["GET"])
@login_required
@perm_required("ofs.atividades_notdone")
def atividades_notdone_get(activity_id):
    activity_id = str(activity_id or "").strip()
    if not activity_id:
        return jsonify({"ok": False, "error": "activityId inválido"}), 400

    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute("""
        SELECT
            activity_id AS activityId,
            city,
            customer_number AS customerNumber,
            customer_phone AS customerPhone,
            customer_name AS customerName,
            appt_number AS apptNumber,
            origin_bucket AS XA_ORIGIN_BUCKET,
            tsk_not AS XA_TSK_NOT,
            ser_clo_imp_ada AS XA_SER_CLO_IMP_ADA,
            resource_id AS resourceId,
            date,
            tratativa_status,
            tratativa_obs,
            tratado_por_username,
            tratado_em
        FROM ofs_atividades_notdone
        WHERE activity_id = %s
        LIMIT 1
    """, (activity_id,))
    row = cur.fetchone()

    cur.close()
    conn.close()

    if not row:
        return jsonify({"ok": False, "error": "Não encontrado"}), 404

    return jsonify({"ok": True, "item": row}), 200

@app.route("/atividades-notdone/revogar", methods=["POST"])
@login_required
@perm_required("ofs.atividades_notdone")
def atividades_notdone_revogar():
    """
    Revoga tratativa de forma ATÔMICA:
    - Só revoga se tratado_em IS NOT NULL
    - Exige observação
    - Registra histórico com status/obs anterior + motivo da revogação
    """
    data = request.get_json(silent=True) or {}

    activity_id = str(data.get("activityId") or "").strip()
    obs = (data.get("observacoes") or "").strip()

    if not activity_id:
        return jsonify({"ok": False, "error": "activityId obrigatório"}), 400
    if len(obs) < 3:
        return jsonify({"ok": False, "error": "Observação obrigatória para revogar"}), 400

    actor = current_actor()
    user_id = actor.get("id")
    username = actor.get("username")

    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    try:
        # Pega estado atual (para histórico)
        cur.execute("""
            SELECT tratativa_status, tratativa_obs, tratado_em, tratado_por_username
            FROM ofs_atividades_notdone
            WHERE activity_id = %s
            LIMIT 1
        """, (activity_id,))
        row = cur.fetchone()

        if not row:
            conn.rollback()
            return jsonify({"ok": False, "error": "activityId não encontrado no banco"}), 404

        if row.get("tratado_em") is None:
            conn.rollback()
            return jsonify({"ok": False, "error": "Este caso não está tratado"}), 409

        before_status = row.get("tratativa_status")
        before_obs = row.get("tratativa_obs")
        before_user = row.get("tratado_por_username")
        before_dt = row.get("tratado_em")

        # HISTÓRICO: revogação
        cur.execute("""
            INSERT INTO ofs_atividades_notdone_history
            (activity_id, action, status, obs, actor_user_id, actor_username)
            VALUES (%s, 'REVOGAR', %s, %s, %s, %s)
        """, (
            activity_id,
            before_status,
            (
                f"[REVOGAÇÃO] {obs}\n\n"
                f"[ANTES] status={before_status or '-'} | por={before_user or '-'} | em={before_dt or '-'}\n"
                f"[OBS ANTERIOR] {before_obs or '-'}"
            ),
            int(user_id) if user_id else None,
            username
        ))

        # UPDATE ATÔMICO: só revoga se ainda estiver tratado
        cur.execute("""
            UPDATE ofs_atividades_notdone
            SET
                tratativa_status = NULL,
                tratativa_obs = NULL,
                tratado_por_user_id = NULL,
                tratado_por_username = NULL,
                tratado_em = NULL
            WHERE activity_id = %s
              AND tratado_em IS NOT NULL
        """, (activity_id,))

        if cur.rowcount == 0:
            conn.rollback()
            return jsonify({"ok": False, "error": "Este caso já foi revogado por outro usuário."}), 409

        conn.commit()

    except Exception as e:
        conn.rollback()
        return jsonify({"ok": False, "error": f"Erro ao revogar: {e}"}), 500

    finally:
        cur.close()
        conn.close()

    # AUDIT (opcional)
    try:
        audit_log(
            actor_user_id=user_id,
            actor_username=username,
            module="ofs",
            action="revogar_tratativa_notdone",
            entity_type="activity",
            entity_ref=activity_id,
            summary=f"Revogou tratativa notdone: activityId={activity_id}",
            meta={"obs": obs[:500]},
        )
    except Exception:
        pass

    return jsonify({"ok": True, "activityId": activity_id}), 200

@app.route("/logs", methods=["GET"])
@login_required
@perm_required("logs.visualizar")
def logs_view():
    import json

    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    # filtros
    user = request.args.get("user", "").strip()
    module = request.args.get("module", "").strip()
    action = request.args.get("action", "").strip()
    q = request.args.get("q", "").strip()
    date_ini = request.args.get("date_ini", "").strip()
    date_fim = request.args.get("date_fim", "").strip()
    export = request.args.get("export") == "1"

    base_query = """
        FROM audit_log
        WHERE 1=1
    """
    params = []

    if user:
        base_query += " AND actor_username LIKE %s"
        params.append(f"%{user}%")

    if module:
        base_query += " AND module = %s"
        params.append(module)

    if action:
        base_query += " AND action = %s"
        params.append(action)

    if q:
        base_query += " AND (summary LIKE %s OR entity_ref LIKE %s)"
        params.extend([f"%{q}%", f"%{q}%"])

    if date_ini:
        base_query += " AND created_at >= %s"
        params.append(f"{date_ini} 00:00:00")

    if date_fim:
        base_query += " AND created_at <= %s"
        params.append(f"{date_fim} 23:59:59")

    # -------- EXPORT CSV --------
    if export:
        cur.execute(
            f"""
            SELECT
                created_at,
                actor_username,
                module,
                action,
                summary,
                entity_type,
                entity_ref
            {base_query}
            ORDER BY created_at DESC
            """,
            tuple(params),
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()

        text_buffer = StringIO()
        writer = csv.writer(text_buffer)
        writer.writerow([
            "data_hora",
            "usuario",
            "modulo",
            "acao",
            "resumo",
            "tipo_entidade",
            "referencia",
        ])

        for r in rows:
            writer.writerow([
                r["created_at"].strftime("%Y-%m-%d %H:%M:%S") if r.get("created_at") else "",
                r.get("actor_username") or "",
                r.get("module") or "",
                r.get("action") or "",
                r.get("summary") or "",
                r.get("entity_type") or "",
                r.get("entity_ref") or "",
            ])

        csv_bytes = text_buffer.getvalue().encode("utf-8-sig")
        output = BytesIO(csv_bytes)

        return send_file(
            output,
            mimetype="text/csv; charset=utf-8",
            as_attachment=True,
            download_name=f"audit_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        )

    # -------- HTML --------
    cur.execute(
        f"""
        SELECT
            id,
            created_at,
            actor_username,
            module,
            action,
            summary,
            entity_type,
            entity_ref,
            api_response
        {base_query}
        ORDER BY created_at DESC
        LIMIT 500
        """,
        tuple(params),
    )
    logs = cur.fetchall()

    # prepara texto do api_response para UI
    for r in logs:
        raw = r.get("api_response")
        if raw is None:
            r["api_response_text"] = ""
        elif isinstance(raw, (dict, list)):
            r["api_response_text"] = json.dumps(raw, ensure_ascii=False)
        else:
            r["api_response_text"] = str(raw)

        # limite de segurança para não estourar tela/HTML
        r["api_response_text"] = (r["api_response_text"] or "")[:10000]

    # filtros auxiliares
    cur.execute("SELECT DISTINCT module FROM audit_log ORDER BY module")
    modules = [r["module"] for r in cur.fetchall()]

    cur.execute("SELECT DISTINCT action FROM audit_log ORDER BY action")
    actions = [r["action"] for r in cur.fetchall()]

    cur.close()
    conn.close()

    return render_template(
        "logs.html",
        logs=logs,
        modules=modules,
        actions=actions,
        filtros={
            "user": user,
            "module": module,
            "action": action,
            "q": q,
            "date_ini": date_ini,
            "date_fim": date_fim,
        },
    )


# Listar usuários de um perfil específico
@app.route("/usuarios-por-perfil/<int:perfil_id>")
@login_required
@perm_required("perfis.gerenciar")
def usuarios_por_perfil(perfil_id):
    # Redireciona para a nova tela única de usuários, já com filtro aplicado
    return redirect(url_for("usuarios_painel", perfil_id=perfil_id))

@app.route("/usuarios", methods=["GET"])
@login_required
@perm_required("usuarios.criar")  # ou "perfis.gerenciar" se quiser restringir ainda mais
def usuarios_painel():
    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    # carregar perfis para o filtro
    cur.execute("SELECT id, nome FROM perfis ORDER BY nome")
    perfis = cur.fetchall()

    perfil_id = request.args.get("perfil_id", type=int)

    query = """
        SELECT u.id,
               u.nome,
               u.username,
               p.nome AS perfil_nome,
               u.last_login
        FROM usuarios u
        LEFT JOIN perfis p ON p.id = u.tipo_id
    """
    params = []
    if perfil_id:
        query += " WHERE u.tipo_id = %s"
        params.append(perfil_id)

    query += " ORDER BY u.nome"

    cur.execute(query, tuple(params))
    usuarios = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "usuarios_painel.html",
        usuarios=usuarios,
        perfis=perfis,
        perfil_id_selecionado=perfil_id,
    )
def registrar_atividade_usuario():
    """Atualiza a tabela usuarios_online com o último acesso do usuário logado."""
    usuario_id = session.get("usuario_id")
    if not usuario_id:
        return

    conn = get_connection()
    cur = conn.cursor()

    agora = datetime.now()
    # se quiser UTC, use datetime.utcnow()

    cur.execute(
        """
        INSERT INTO usuarios_online (usuario_id, last_seen)
        VALUES (%s, %s)
        ON DUPLICATE KEY UPDATE last_seen = VALUES(last_seen)
        """,
        (usuario_id, agora),
    )
    conn.commit()
    cur.close()
    conn.close()

@app.route("/status-online")
@login_required
def status_online():
    conn = get_connection()
    cur = conn.cursor()

    limite = datetime.now() - timedelta(minutes=5)

    cur.execute(
        "SELECT COUNT(*) FROM usuarios_online WHERE last_seen >= %s",
        (limite,),
    )
    count = cur.fetchone()[0]

    cur.close()
    conn.close()

    return jsonify({
        "usuarios_online": count
    })


@app.route("/sap/acompanhamento-critica", methods=["GET"])
@login_required
@perm_required("sap.acompanhamento_critica")
def sap_acompanhamento_critica():
    today = datetime.now().strftime("%Y-%m-%d")
    date_from = (request.args.get("dateFrom") or today).strip()
    date_to = (request.args.get("dateTo") or today).strip()
    resources = (request.args.get("resources") or "MG").strip()

    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT
            activity_id AS activityId,
            city,
            activity_type AS activityType,
            appt_number AS apptNumber,
            origin_bucket AS XA_ORIGIN_BUCKET,
            resource_id AS resourceId,
            xa_sap_crt AS XA_SAP_CRT,
            `date` AS date,
            created_at
        FROM sap_criticas_atividades
        WHERE `date` BETWEEN %s AND %s
        ORDER BY `date` ASC, created_at DESC
        LIMIT 5000
    """, (date_from, date_to))
    items = cur.fetchall()
    total= len(items)
    cur.close()
    conn.close()

    return render_template(
        "criticas_sap/sap_acompanhamento_critica.html",
        items=items,
        date_from=date_from,
        date_to=date_to,
        resources=resources,
        total = total,
    )

#CRITICAS SAP

@app.route("/sap/acompanhamento-critica/importar", methods=["POST"])
@login_required
@perm_required("sap.acompanhamento_critica")
def sap_acompanhamento_critica_importar():
    today = datetime.now().strftime("%Y-%m-%d")
    date_from = (request.form.get("dateFrom") or today).strip()
    date_to = (request.form.get("dateTo") or today).strip()
    resources = (request.form.get("resources") or "MG").strip()

    client = OFSClient()

    fields = [
        "activityId",
        "city",
        "activityType",
        "apptNumber",
        "XA_ORIGIN_BUCKET",
        "resourceId",
        "XA_SAP_CRT",
        "XA_SAP_CRT_LDG",
        "date",
    ]

    base_params = {
        "dateFrom": date_from,
        "dateTo": date_to,
        "resources": resources,
        "q": "XA_SAP_CRT==1",
        "fields": ",".join(fields),
        "limit": 2000,
        "offset": 0,
    }

    items = []
    has_more = True
    max_pages = 20
    page = 0

    try:
        while has_more and page < max_pages:
            qs = urlencode(base_params, safe="=,'")
            url = f"{client.base_url}/activities/?{qs}"
            data = client.authenticated_get(url)

            batch = data.get("items") or []
            items.extend(batch)

            has_more = bool(data.get("hasMore"))
            if has_more:
                base_params["offset"] = len(items)
            page += 1

    except Exception as e:
        flash(f"❌ Falha ao importar da API: {e}", "danger")
        return redirect(url_for("sap_acompanhamento_critica", dateFrom=date_from, dateTo=date_to, resources=resources))

    conn = get_connection()
    cur = conn.cursor()

    inserted = 0
    skipped = 0

    sql = """
        INSERT IGNORE INTO sap_criticas_atividades
        (activity_id, city, activity_type, appt_number, origin_bucket, resource_id, xa_sap_crt, xa_sap_crt_ldg, `date`)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """

    for a in items:
        activity_id = str(a.get("activityId") or "").strip()
        if not activity_id:
            continue

        cur.execute(sql, (
            activity_id,
            str(a.get("city") or "") or None,
            str(a.get("activityType") or "") or None,
            str(a.get("apptNumber") or "") or None,
            str(a.get("XA_ORIGIN_BUCKET") or "") or None,
            str(a.get("resourceId") or "") or None,
            1 if str(a.get("XA_SAP_CRT") or "").strip() == "1" else 0,
            a.get("XA_SAP_CRT_LDG"),  # LONGTEXT
            str(a.get("date") or "") or None,
        ))

        if cur.rowcount == 1:
            inserted += 1
        else:
            skipped += 1

    conn.commit()
    cur.close()
    conn.close()

    flash(f"✅ Importação concluída. Novos: {inserted} | Já existiam: {skipped}", "success")
    return redirect(url_for("sap_acompanhamento_critica", dateFrom=date_from, dateTo=date_to, resources=resources))

@app.route("/sap/acompanhamento-critica/<activity_id>", methods=["GET"])
@login_required
@perm_required("sap.acompanhamento_critica")
def sap_acompanhamento_critica_get(activity_id):
    activity_id = str(activity_id or "").strip()
    if not activity_id:
        return jsonify({"ok": False, "error": "activityId inválido"}), 400

    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute("""
        SELECT
            activity_id AS activityId,
            xa_sap_crt_ldg AS XA_SAP_CRT_LDG
        FROM sap_criticas_atividades
        WHERE activity_id = %s
        LIMIT 1
    """, (activity_id,)) 
    row = cur.fetchone()

    cur.close()
    conn.close()

    if not row:
        return jsonify({"ok": False, "error": "Não encontrado"}), 404

    return jsonify({"ok": True, "item": row}), 200

@app.route("/sap/acompanhamento-critica/dashboard", methods=["GET"])
@login_required
@perm_required("sap.acompanhamento_critica")
def sap_dashboard_critica():
    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute("SELECT MIN(`date`) AS min_date, MAX(`date`) AS max_date FROM sap_criticas_atividades")
    mm = cur.fetchone() or {}
    min_date = (mm.get("min_date") or "")
    max_date = (mm.get("max_date") or "")

    cur.execute("""
    SELECT activity_type
    FROM sap_criticas_atividades
    WHERE activity_type IS NOT NULL AND activity_type <> ''
    GROUP BY activity_type
    ORDER BY activity_type
    """)
    types = [r["activity_type"] for r in cur.fetchall()]

    cur.execute("""
    SELECT DISTINCT tb.bucket
    FROM projToquio.td_bucket tb
    JOIN (
        SELECT DISTINCT city
        FROM sap_criticas_atividades
        WHERE city IS NOT NULL AND city <> ''
    ) c ON c.city = tb.nomeCidade
    ORDER BY tb.bucket
    """)
    buckets = [r["bucket"] for r in cur.fetchall()]

    cur.close()
    conn.close()

    return render_template(
    "criticas_sap/dashboard_critica.html",
    min_date=min_date,
    max_date=max_date,
    types=types,
    buckets=buckets
    )

@app.route("/sap/acompanhamento-critica/dashboard/data", methods=["GET"])
@login_required
@perm_required("sap.acompanhamento_critica")
def sap_dashboard_critica_data():
    date_from = (request.args.get("dateFrom") or "").strip()
    date_to = (request.args.get("dateTo") or "").strip()
    activity_type = (request.args.get("activityType") or "").strip()

    # múltiplos buckets: buckets=a&buckets=b
    buckets = [b.strip() for b in request.args.getlist("buckets") if (b or "").strip()]

    if not date_from or not date_to:
        return jsonify({"ok": False, "error": "Informe dateFrom e dateTo."}), 400
    if date_to < date_from:
        return jsonify({"ok": False, "error": "dateTo não pode ser menor que dateFrom."}), 400

    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    try:
        sql = """
            SELECT `date` AS d, COUNT(*) AS total
            FROM sap_criticas_atividades
            WHERE `date` BETWEEN %s AND %s
        """
        params = [date_from, date_to]

        if activity_type:
            sql += " AND activity_type = %s"
            params.append(activity_type)

        if buckets:
            placeholders = ",".join(["%s"] * len(buckets))
            sql += f" AND origin_bucket IN ({placeholders})"
            params.extend(buckets)

        sql += " GROUP BY `date` ORDER BY `date` ASC"

        cur.execute(sql, tuple(params))
        rows = cur.fetchall()

        labels = [r["d"] for r in rows]
        values = [int(r["total"] or 0) for r in rows]

        return jsonify({"ok": True, "labels": labels, "values": values}), 200

    except Exception as e:
        return jsonify({"ok": False, "error": f"Erro no dashboard/data: {e}"}), 500

    finally:
        cur.close()
        conn.close()

@app.route("/sap/acompanhamento-critica/dashboard/data2", methods=["GET"])
@login_required
@perm_required("sap.acompanhamento_critica")
def sap_dashboard_critica_data2():
    date_from = (request.args.get("dateFrom") or "").strip()
    date_to = (request.args.get("dateTo") or "").strip()
    activity_type = (request.args.get("activityType") or "").strip()

    buckets = [b.strip() for b in request.args.getlist("buckets") if (b or "").strip()]

    if not date_from or not date_to:
        return jsonify({"ok": False, "error": "Informe dateFrom e dateTo."}), 400
    if date_to < date_from:
        return jsonify({"ok": False, "error": "dateTo não pode ser menor que dateFrom."}), 400

    produced_types = [
        "INS_DEV", "SOL_SER", "MIG_PLA", "MUD_END", "RET",
        "INS", "SUP", "SUP_REP", "SUP_QUA", "MIG_TEC", "QUA",
    ]

    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    try:
        # 1) Criticadas -> labels base (como você pediu)
        sql_c = """
            SELECT `date` AS d, COUNT(*) AS total
            FROM sap_criticas_atividades
            WHERE `date` BETWEEN %s AND %s
        """
        params_c = [date_from, date_to]

        if activity_type:
            sql_c += " AND activity_type = %s"
            params_c.append(activity_type)

        if buckets:
            ph = ",".join(["%s"] * len(buckets))
            sql_c += f" AND origin_bucket IN ({ph})"
            params_c.extend(buckets)

        sql_c += " GROUP BY `date` ORDER BY `date` ASC"

        cur.execute(sql_c, tuple(params_c))
        rows_c = cur.fetchall()

        labels = [r["d"] for r in rows_c]
        criticadas = [int(r["total"] or 0) for r in rows_c]

        if not labels:
            return jsonify({"ok": True, "labels": [], "criticadas": [], "produzidas": []}), 200

        # 2) Produzidas (completed + whitelist + mesmos filtros)
        sql_p = """
            SELECT b.`date` AS d, COUNT(*) AS total
            FROM ofs_atividades_base b
            WHERE b.`date` BETWEEN %s AND %s
              AND b.status = 'completed'
        """
        params_p = [date_from, date_to]

        ph = ",".join(["%s"] * len(produced_types))
        sql_p += f" AND b.activity_type IN ({ph})"
        params_p.extend(produced_types)

        if activity_type:
            sql_p += " AND b.activity_type = %s"
            params_p.append(activity_type)

        if buckets:
            ph = ",".join(["%s"] * len(buckets))
            sql_p += f" AND b.origin_bucket IN ({ph})"
            params_p.extend(buckets)

        sql_p += " GROUP BY b.`date` ORDER BY b.`date` ASC"

        cur.execute(sql_p, tuple(params_p))
        rows_p = cur.fetchall()

        prod_map = {r["d"]: int(r["total"] or 0) for r in rows_p}
        # null quando não há produção no dia (pra “linha só aparecer quando houver”)
        produzidas = [prod_map.get(d, None) for d in labels]

        return jsonify({
            "ok": True,
            "labels": labels,
            "criticadas": criticadas,
            "produzidas": produzidas,
        }), 200

    except Exception as e:
        return jsonify({"ok": False, "error": f"Erro no dashboard/data2: {e}"}), 500
    finally:
        cur.close()
        conn.close()

@app.route("/ofs/atividades-base", methods=["GET"])
@login_required
@perm_required("ofs.atividades_base")
def ofs_atividades_base():
    # filtros (agora server-side)
    date_from = (request.args.get("dateFrom") or "").strip()
    date_to = (request.args.get("dateTo") or "").strip()
    resources = (request.args.get("resources") or "02").strip()  # mantido no form, sem filtrar DB por padrão

    # multi (querystring repetida)
    selected_types = [t.strip() for t in request.args.getlist("activityType") if (t or "").strip()]
    selected_statuses = [s.strip() for s in request.args.getlist("status") if (s or "").strip()]

    # paginação
    per_page = 100
    page = request.args.get("page", default=1, type=int)
    if page < 1:
        page = 1
    offset = (page - 1) * per_page

    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    try:
        # DISTINCTs para UI (modal export / filtros server-side)
        cur.execute("""
            SELECT DISTINCT activity_type
            FROM ofs_atividades_base
            WHERE activity_type IS NOT NULL AND activity_type <> ''
            ORDER BY activity_type
        """)
        activity_types = [r["activity_type"] for r in cur.fetchall()]

        cur.execute("""
            SELECT DISTINCT origin_bucket
            FROM ofs_atividades_base
            WHERE origin_bucket IS NOT NULL AND origin_bucket <> ''
            ORDER BY origin_bucket
        """)
        buckets = [r["origin_bucket"] for r in cur.fetchall()]

        cur.execute("""
            SELECT DISTINCT status
            FROM ofs_atividades_base
            WHERE status IS NOT NULL AND status <> ''
            ORDER BY status
        """)
        statuses = [r["status"] for r in cur.fetchall()]

        # WHERE dinâmico (server-side)
        where = ["1=1"]
        params = []

        # date range (se nenhum informado -> traz tudo)
        if date_from and date_to:
            where.append("b.`date` BETWEEN %s AND %s")
            params.extend([date_from, date_to])
        elif date_from:
            where.append("b.`date` >= %s")
            params.append(date_from)
        elif date_to:
            where.append("b.`date` <= %s")
            params.append(date_to)

        # activityType IN (...)
        if selected_types:
            placeholders = ",".join(["%s"] * len(selected_types))
            where.append(f"b.activity_type IN ({placeholders})")
            params.extend(selected_types)

        # status IN (...)
        if selected_statuses:
            placeholders = ",".join(["%s"] * len(selected_statuses))
            where.append(f"b.status IN ({placeholders})")
            params.extend(selected_statuses)

        # Se você quiser filtrar por resource_id na base (opcional):
        # if resources:
        #     where.append("b.resource_id = %s")
        #     params.append(resources)

        where_sql = " AND ".join(where)

        # total filtrado
        cur.execute(f"""
            SELECT COUNT(*) AS total
            FROM ofs_atividades_base b
            WHERE {where_sql}
        """, tuple(params))
        total = int((cur.fetchone() or {}).get("total") or 0)

        total_pages = max(1, (total + per_page - 1) // per_page)
        if page > total_pages:
            page = total_pages
            offset = (page - 1) * per_page

        # itens paginados (já filtrados no DB)
        cur.execute(f"""
            SELECT
                b.activity_id AS activityId,
                b.city,
                b.activity_type AS activityType,
                COALESCE(atm.label_pt, b.activity_type) AS activityType_pt,

                b.appt_number AS apptNumber,
                b.origin_bucket AS XA_ORIGIN_BUCKET,
                b.resource_id AS resourceId,

                b.status AS status,
                COALESCE(sm.label_pt, b.status) AS status_pt,

                b.xa_org_sys AS XA_ORG_SYS,
                b.`date` AS date
            FROM ofs_atividades_base b
            LEFT JOIN ofs_activity_type_map atm
              ON atm.code = b.activity_type AND atm.is_active = 1
            LEFT JOIN ofs_status_map sm
              ON sm.code = b.status AND sm.is_active = 1
            WHERE {where_sql}
            ORDER BY b.`date` ASC, b.activity_id ASC
            LIMIT %s OFFSET %s
        """, tuple(params + [per_page, offset]))
        items = cur.fetchall()

    finally:
        cur.close()
        conn.close()

    return render_template(
        "atividades_base/ofs_atividades_base.html",
        items=items,
        date_from=date_from,
        date_to=date_to,
        resources=resources,
        total=total,
        page=page,
        total_pages=total_pages,
        per_page=per_page,

        activity_types=activity_types,
        buckets=buckets,
        statuses=statuses,

        # novos para preservar seleção
        selected_types=selected_types,
        selected_statuses=selected_statuses,
    )


@app.route("/ofs/atividades-base/exportar", methods=["POST"])
@login_required
@perm_required("ofs.atividades_base")
def ofs_atividades_base_exportar():
    activity_types = [t.strip() for t in request.form.getlist("activityType") if (t or "").strip()]
    statuses = [s.strip() for s in request.form.getlist("status") if (s or "").strip()]
    date_from = (request.form.get("dateFrom") or "").strip()
    date_to = (request.form.get("dateTo") or "").strip()

    # buckets pode vir como lista
    buckets = [b.strip() for b in request.form.getlist("buckets") if (b or "").strip()]

    # valida datas
    try:
        dt_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        dt_to = datetime.strptime(date_to, "%Y-%m-%d").date()
    except Exception:
        flash("Informe um período válido (De / Até).", "danger")
        return redirect(url_for("ofs_atividades_base"))

    if dt_to < dt_from:
        flash("O campo 'Até' não pode ser menor que 'De'.", "danger")
        return redirect(url_for("ofs_atividades_base", dateFrom=date_from, dateTo=date_to))

    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    try:
        sql = """
            SELECT
                b.activity_id AS activityId,
                b.city,
                b.activity_type AS activityType,
                COALESCE(atm.label_pt, b.activity_type) AS activityType_pt,

                b.appt_number AS apptNumber,
                b.origin_bucket AS XA_ORIGIN_BUCKET,
                b.resource_id AS resourceId,

                b.status AS status,
                COALESCE(sm.label_pt, b.status) AS status_pt,

                b.xa_org_sys AS XA_ORG_SYS,
                b.`date` AS date
                FROM ofs_atividades_base b
                LEFT JOIN ofs_activity_type_map atm
                ON atm.code = b.activity_type AND atm.is_active = 1
                LEFT JOIN ofs_status_map sm
                ON sm.code = b.status AND sm.is_active = 1
                WHERE b.`date` BETWEEN %s AND %s
        """
        params = [date_from, date_to]

        if activity_types:
            placeholders = ",".join(["%s"] * len(activity_types))
            sql += f" AND activity_type IN ({placeholders})"
            params.extend(activity_types)

        if buckets:
            placeholders = ",".join(["%s"] * len(buckets))
            sql += f" AND origin_bucket IN ({placeholders})"
            params.extend(buckets)

        if statuses:
            placeholders = ",".join(["%s"] * len(statuses))
            sql += f" AND status IN ({placeholders})"
            params.extend(statuses)

        sql += " ORDER BY `date` ASC, last_seen_at DESC"

        cur.execute(sql, tuple(params))
        rows = cur.fetchall()

    finally:
        cur.close()
        conn.close()

    # Monta XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "AtividadesBase"

    headers = [
        "activityId",
        "city",
        "activityType",
        "activityType_pt",
        "apptNumber",
        "XA_ORIGIN_BUCKET",
        "resourceId",
        "status",
        "status_pt",
        "XA_ORG_SYS",
        "date",
        "last_seen_at",
    ]
    ws.append(headers)

    for r in rows:
        ws.append([r.get(h) for h in headers])

    _xlsx_auto_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"ofs_atividades_base_{dt_from}_{dt_to}_{stamp}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
@app.route("/ofs/atividades-base/importar", methods=["POST"])
@login_required
@perm_required("ofs.atividades_base")
def ofs_atividades_base_importar():
    today = datetime.now().strftime("%Y-%m-%d")
    date_from = (request.form.get("dateFrom") or today).strip()
    date_to = (request.form.get("dateTo") or today).strip()
    resources = (request.form.get("resources") or "02").strip()

    # valida datas mínimas
    try:
        dt_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        dt_to = datetime.strptime(date_to, "%Y-%m-%d").date()
        if dt_to < dt_from:
            flash("O campo 'Até' não pode ser menor que 'De'.", "danger")
            return redirect(url_for("ofs_atividades_base", dateFrom=date_from, dateTo=date_to, resources=resources))
    except Exception:
        flash("Informe um período válido (De / Até).", "danger")
        return redirect(url_for("ofs_atividades_base", dateFrom=date_from, dateTo=date_to, resources=resources))

    client = OFSClient()

    fields = [
        "activityId",
        "city",
        "activityType",
        "apptNumber",
        "XA_ORIGIN_BUCKET",
        "resourceId",
        "status",
        "XA_ORG_SYS",
        "date",
    ]

    base_params = {
        "dateFrom": date_from,
        "dateTo": date_to,
        "resources": resources,
        "fields": ",".join(fields),
        "limit": 2000,
        "offset": 0,
    }

    items = []
    has_more = True
    max_pages = 30
    page = 0

    try:
        while has_more and page < max_pages:
            qs = urlencode(base_params, safe="=,'")
            url = f"{client.base_url}/activities/?{qs}"
            data = client.authenticated_get(url)

            batch = data.get("items") or []
            items.extend(batch)

            has_more = bool(data.get("hasMore"))
            if has_more:
                base_params["offset"] = len(items)
            page += 1

    except Exception as e:
        flash(f"❌ Falha ao importar da API: {e}", "danger")
        return redirect(url_for("ofs_atividades_base", dateFrom=date_from, dateTo=date_to, resources=resources))

    conn = get_connection()
    cur = conn.cursor()

    inserted = 0
    updated = 0

    # UPSERT (insere novo; se já existe, atualiza campos + last_seen_at)
    sql = """
        INSERT INTO ofs_atividades_base
        (activity_id, city, activity_type, appt_number, origin_bucket, resource_id, status, xa_org_sys, `date`)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON DUPLICATE KEY UPDATE
            city = VALUES(city),
            activity_type = VALUES(activity_type),
            appt_number = VALUES(appt_number),
            origin_bucket = VALUES(origin_bucket),
            resource_id = VALUES(resource_id),
            status = VALUES(status),
            xa_org_sys = VALUES(xa_org_sys),
            `date` = VALUES(`date`),
            last_seen_at = NOW()
    """

    for a in items:
        activity_id = str(a.get("activityId") or "").strip()
        if not activity_id:
            continue

        cur.execute(sql, (
            activity_id,
            str(a.get("city") or "") or None,
            str(a.get("activityType") or "") or None,
            str(a.get("apptNumber") or "") or None,
            str(a.get("XA_ORIGIN_BUCKET") or "") or None,
            str(a.get("resourceId") or "") or None,
            str(a.get("status") or "") or None,
            str(a.get("XA_ORG_SYS") or "") or None,
            str(a.get("date") or "") or None,
        ))

        # MySQL: rowcount costuma ser 1 (insert) ou 2 (update) nesse padrão
        if cur.rowcount == 1:
            inserted += 1
        elif cur.rowcount == 2:
            updated += 1

    conn.commit()
    cur.close()
    conn.close()

    flash(f"✅ Importação concluída. Novos: {inserted} | Atualizados: {updated} | Total API: {len(items)}", "success")
    return redirect(url_for("ofs_atividades_base", dateFrom=date_from, dateTo=date_to, resources=resources))
# ===========================
# Logout
# ===========================

@app.route("/logout")
def logout():
    actor = current_actor()

    # AUDIT: logout
    audit_log(
        actor_user_id=actor.get("id"),
        actor_username=actor.get("username"),
        module="auth",
        action="logout",
        entity_type="usuario",
        entity_id=actor.get("id"),
        entity_ref=actor.get("username"),
        summary=f"Logout realizado: {actor.get('username')}",
        meta={"ip": request.remote_addr, "ua": request.user_agent.string},
    )

    usuario_id = session.get("usuario_id")
    if usuario_id:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM usuarios_online WHERE usuario_id = %s", (usuario_id,))
        conn.commit()
        cur.close()
        conn.close()

    session.clear()
    return redirect(url_for("login"))



if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)