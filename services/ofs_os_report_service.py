import json
import os
import time
import threading
import uuid
from datetime import datetime, timedelta
from typing import Dict, List, Tuple

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database.connection import get_connection
from database.audit import audit_log
from ofs.client import OFSClient
from core.utils import xlsx_auto_width


REQUEST_TIMEOUT = 60
API_LIMIT = 2000

RESOURCE_TYPES = {
    "BK": "Bucket",
    "ESTADO": "Estado",
    "GR": "Grupo",
    "LC": "Loja Comercial",
    "TCP": "Técnico Parceiro",
    "TCW": "Técnico Retirada",
    "TCV": "Técnico Vero",
}

RESOURCE_TYPES_OFS = {
    "BK": "Bucket",
    "GR": "Grupo",
    "LC": "Loja Comercial",
    "TCP": "Técnico Parceiro",
    "TCW": "Técnico Retirada",
    "TCV": "Técnico Vero",
}

ESTADO_RESOURCE_IDS = {
    "GO",
    "MG",
    "MS",
    "SP",
    "SUL",
}

STATUS_OPTIONS = [
    "completed",
    "notdone",
    "cancelled",
    "enroute",
    "pending",
    "started",
    "suspended",
]

CLOSURE_FIELDS = [
    "XA_SER_CLO_PRO_ADA",
    "XA_SER_CLO_IMP_ADA",
    "XA_SER_CLO_PRO_NG",
    "XA_SER_CLO_INP_NG",
]

REDES_CLOSURE_FIELDS = [
    "XA_SER_CLO_PRO_HLX",
    "XA_SER_CLO_IMP_HLX",
]

TASK_TYPE_PROPERTY_CODE = "XA_TSK_TYP"
METADATA_ENUM_LIMIT = 100

EXTRA_REPORT_FIELDS_PERMISSION = "relatorios.campos_extras"
FIELD_CHOICES = [
    {
        "key": "resourceId",
        "label": "resourceId",
        "api_fields": ["resourceId"],
        "xlsx_header": "ID do técnico",
    },
    {
        "key": "resourceName",
        "label": "Nome do técnico",
        "api_fields": ["resourceId"],
        "xlsx_header": "Nome do técnico",
    },
    {
        "key": "resourceStatus",
        "label": "Status do recurso",
        "api_fields": ["resourceId"],
        "xlsx_header": "Status do recurso",
    },
    {
        "key": "XA_REQ_CRE_DAT",
        "label": "XA_REQ_CRE_DAT",
        "api_fields": ["XA_REQ_CRE_DAT"],
        "xlsx_header": "Data da criação da OS",
    },
    {
        "key": "timeSlot",
        "label": "timeSlot",
        "api_fields": ["timeSlot"],
        "xlsx_header": "Turno",
    },
    {
        "key": "workZone",
        "label": "workZone",
        "api_fields": ["workZone"],
        "xlsx_header": "workZone",
    },
    {
        "key": "stateProvince",
        "label": "stateProvince",
        "api_fields": ["stateProvince"],
        "xlsx_header": "Estado",
    },
    {
        "key": "XA_SAP_CRT_LDG",
        "label": "XA_SAP_CRT_LDG",
        "api_fields": ["XA_SAP_CRT_LDG"],
        "xlsx_header": "Retorno API SAP",
    },
    {
        "key": "date",
        "label": "date",
        "api_fields": ["date"],
        "xlsx_header": "Data",
    },
    {
        "key": "startTime",
        "label": "startTime (Início da atividade)",
        "api_fields": ["startTime"],
        "xlsx_header": "Início da atividade",
    },
    {
        "key": "endTime",
        "label": "endTime (Fim da atividade)",
        "api_fields": ["endTime"],
        "xlsx_header": "Fim da atividade",
    },
    {
        "key": "XA_TSK_NOT",
        "label": "XA_TSK_NOT (Observações finais)",
        "api_fields": ["XA_TSK_NOT"],
        "xlsx_header": "Observações finais",
    },
    {
        "key": "XA_ORG_SYS",
        "label": "XA_ORG_SYS",
        "api_fields": ["XA_ORG_SYS"],
        "xlsx_header": "Sistema de Origem",
    },
    {
        "key": "status",
        "label": "status",
        "api_fields": ["status"],
        "xlsx_header": "Status",
    },
    {
        "key": "customerName",
        "label": "customerName (somente primeiro nome)",
        "api_fields": ["customerName"],
        "xlsx_header": "Primeiro Nome do Cliente",
    },
    {
        "key": "customerNameFull",
        "label": "customerName (nome completo)",
        "api_fields": ["customerName"],
        "xlsx_header": "Nome completo do cliente",
        "required_perm": EXTRA_REPORT_FIELDS_PERMISSION,
    },
    {
        "key": "city",
        "label": "city",
        "api_fields": ["city"],
        "xlsx_header": "Cidade",
    },
    {
        "key": "fechamento_atividade",
        "label": "Fechamento da atividade",
        "api_fields": CLOSURE_FIELDS,
        "xlsx_header": "Fechamento da atividade",
    },
    {
        "key": "activityType",
        "label": "activityType",
        "api_fields": ["activityType"],
        "xlsx_header": "Tipo de Atividade",
    },
    {
        "key": "XA_TSK_TYP",
        "label": "XA_TSK_TYP",
        "api_fields": ["XA_TSK_TYP"],
        "xlsx_header": "Tipo de tarefa",
    },
    {
        "key": "activityId",
        "label": "activityId",
        "api_fields": ["activityId"],
        "xlsx_header": "ID da atividade",
    },
    {
        "key": "apptNumber",
        "label": "apptNumber",
        "api_fields": ["apptNumber"],
        "xlsx_header": "Código da OS",
    },
    {
        "key": "XA_PLA_CON_CUS",
        "label": "XA_PLA_CON_CUS",
        "api_fields": ["XA_PLA_CON_CUS"],
        "xlsx_header": "Plano contratado",
        "required_perm": EXTRA_REPORT_FIELDS_PERMISSION,
    },
    {
        "key": "customerNumber",
        "label": "customerNumber",
        "api_fields": ["customerNumber"],
        "xlsx_header": "Contrato do cliente",
        "required_perm": EXTRA_REPORT_FIELDS_PERMISSION,
    },
    {
        "key": "XA_PRO_COD_SAP",
        "label": "XA_PRO_COD_SAP",
        "api_fields": ["XA_PRO_COD_SAP"],
        "xlsx_header": "Código SAP do Produto",
        "redes_only": True,
    },
    {
        "key": "XA_RES_API_NG_STA",
        "label": "XA_RES_API_NG_STA",
        "api_fields": ["XA_RES_API_NG_STA"],
        "xlsx_header": "Status da API NG",
        "ofs_os_only": True,
    },
    {
        "key": "XA_RES_API_NG_RESPONSE",
        "label": "XA_RES_API_NG_RESPONSE",
        "api_fields": ["XA_RES_API_NG_RESPONSE"],
        "xlsx_header": "Response API NG",
        "ofs_os_only": True,
    },
]
FIELD_MAP = {f["key"]: f for f in FIELD_CHOICES}
THERMOMETER_API_FIELDS = [
    "activityId",
    "apptNumber",
    "XA_AV_CLI",
    "XA_AV_CLI_CAT",
    "XA_AV_CLI_SUB_CAT",
    "XA_AV_CLI_CON",
    "city",
    "resourceId",
    "timeSlot",
    "activityType",
    "XA_TSK_NOT",
    "date",
    "XA_REQ_CRE_DAT",
    "status",
]

THERMOMETER_HEADERS = [
    "Código da OS",
    "Avaliação do cliente",
    "Categoria da avaliação",
    "Subcategoria da avaliação",
    "Conclusão da avaliação",
    "Cidade",
    "Nome do técnico",
    "Turno",
    "Tipo de atividade",
    "Observações finais",
    "Data",
    "Data da criação da OS",
]


def validate_thermometer_report_payload(payload: dict) -> dict:
    date_from = str(payload.get("dateFrom") or "").strip()
    date_to = str(payload.get("dateTo") or "").strip()

    _validate_dates(date_from, date_to)
    resource_ids = _validate_resources(payload.get("resources") or [])
    activity_types = _validate_activity_types(payload.get("activityTypes") or [])

    allowed_codes = _valid_activity_type_codes_by_category({"customer_home"})
    invalid = [
        code for code in activity_types
        if code not in allowed_codes
    ]

    if invalid:
        raise ValueError(
            "O relatório do termômetro permite apenas tipos de atividade B2C/casa cliente."
        )

    return {
        "date_from": date_from,
        "date_to": date_to,
        "resources": resource_ids,
        "statuses": STATUS_OPTIONS,
        "activity_types": activity_types,
        "fields": THERMOMETER_API_FIELDS,
    }


def _has_customer_rating(item: dict) -> bool:
    value = str(item.get("XA_AV_CLI") or "").strip()
    return value in {"1", "2", "3", "4", "5"}


def _thermometer_conclusion(value):
    value = str(value or "").strip()
    if value == "1":
        return "Concluída"
    return value


def _fetch_thermometer_activities(client: OFSClient, config: dict, base_dir: str, job_id: str, status_payload: dict) -> List[dict]:
    url = f"{client.base_url}/activities/"
    headers = {"Accept": "application/json"}
    resources_param = ",".join(config["resources"])
    status_query = _build_or_equals_query("status", config["statuses"])
    activity_type_query = _build_or_equals_query("activityType", config["activity_types"])
    combined_query = f"{status_query} and {activity_type_query}"

    all_items = []
    seen_activity_ids = set()
    date_list = list(_iter_date_strings(config["date_from"], config["date_to"]))
    total_days = len(date_list)

    for day_index, day in enumerate(date_list, start=1):
        offset = 0
        page = 1

        while True:
            params = [
                ("dateFrom", day),
                ("dateTo", day),
                ("resources", resources_param),
                ("q", combined_query),
                ("fields", ",".join(THERMOMETER_API_FIELDS)),
                ("limit", str(API_LIMIT)),
                ("offset", str(offset)),
            ]

            status_payload.update({
                "status": "running",
                "phase": f"Consultando termômetro - {day} - página {page}",
                "rows_so_far": len(all_items),
                "current_day": day,
                "current_day_index": day_index,
                "total_days": total_days,
                "offset": offset,
                "page": page,
            })
            _write_job_status(base_dir, job_id, status_payload)

            resp = requests.get(
                url,
                headers=headers,
                params=params,
                auth=client.auth,
                timeout=REQUEST_TIMEOUT,
            )
            resp.raise_for_status()

            data = resp.json()
            items = _normalize_items_payload(data)

            if not items:
                break

            for item in items:
                activity_id = str(item.get("activityId") or "").strip()

                if activity_id:
                    if activity_id in seen_activity_ids:
                        continue
                    seen_activity_ids.add(activity_id)

                if _has_customer_rating(item):
                    all_items.append(item)

            has_more = bool(data.get("hasMore")) if isinstance(data, dict) else False
            if not has_more:
                break

            offset += len(items)
            page += 1

    return all_items


def _build_thermometer_xlsx(rows: List[dict], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Termômetro Cliente"
    ws.append(THERMOMETER_HEADERS)

    resource_name_map = _load_resource_name_map()
    activity_type_label_map = _load_activity_type_label_map()

    for item in rows:
        resource_id = str(item.get("resourceId") or "").strip()
        activity_type = str(item.get("activityType") or "").strip()

        ws.append([
            item.get("apptNumber") or "",
            item.get("XA_AV_CLI") or "",
            item.get("XA_AV_CLI_CAT") or "",
            item.get("XA_AV_CLI_SUB_CAT") or "",
            _thermometer_conclusion(item.get("XA_AV_CLI_CON")),
            item.get("city") or "",
            resource_name_map.get(resource_id, "Técnico não encontrado na base"),
            item.get("timeSlot") or "",
            activity_type_label_map.get(activity_type, activity_type),
            item.get("XA_TSK_NOT") or "",
            item.get("date") or "",
            item.get("XA_REQ_CRE_DAT") or "",
        ])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top")

    ws.freeze_panes = "A2"
    xlsx_auto_width(ws)
    wb.save(output_path)


def _run_thermometer_report_job(base_dir: str, job_id: str, actor: dict, config: dict):
    filename = f"relatorio_termometro_cliente_{config['date_from']}_{config['date_to']}_{job_id[:8]}.xlsx"

    status_payload = {
        "status": "running",
        "phase": "Iniciando extração do termômetro",
        "created_at": _now_iso(),
        "finished_at": None,
        "rows_so_far": 0,
        "total_rows": 0,
        "filename": filename,
        "dateFrom": config["date_from"],
        "dateTo": config["date_to"],
        "resources": config["resources"],
        "error": None,
    }

    _write_job_status(base_dir, job_id, status_payload)

    try:
        client = OFSClient()
        rows = _fetch_thermometer_activities(client, config, base_dir, job_id, status_payload)

        status_payload.update({
            "status": "running",
            "phase": "Gerando XLSX",
            "rows_so_far": len(rows),
            "total_rows": len(rows),
        })
        _write_job_status(base_dir, job_id, status_payload)

        output_path = _job_xlsx_path(base_dir, job_id)
        _build_thermometer_xlsx(rows, output_path)

        status_payload.update({
            "status": "completed",
            "phase": "Relatório do termômetro concluído",
            "finished_at": _now_iso(),
            "rows_so_far": len(rows),
            "total_rows": len(rows),
            "download_ready": True,
        })
        _write_job_status(base_dir, job_id, status_payload)

        audit_log(
            actor_user_id=actor.get("id"),
            actor_username=actor.get("username"),
            module="relatorios",
            action="finish_thermometer_report",
            entity_type="report",
            entity_ref=job_id,
            summary="Concluiu relatório do Termômetro do Cliente",
            meta={
                "dateFrom": config["date_from"],
                "dateTo": config["date_to"],
                "resources": config["resources"],
                "total_rows": len(rows),
                "filename": filename,
            },
        )

    except Exception as e:
        status_payload.update({
            "status": "failed",
            "phase": "Falha na extração do termômetro",
            "finished_at": _now_iso(),
            "error": str(e),
        })
        _write_job_status(base_dir, job_id, status_payload)


def start_thermometer_report_job(base_dir: str, actor: dict, config: dict) -> str:
    _ensure_dir(base_dir)
    job_id = uuid.uuid4().hex

    _write_job_status(base_dir, job_id, {
        "status": "queued",
        "phase": "Relatório do termômetro na fila",
        "created_at": _now_iso(),
        "finished_at": None,
        "rows_so_far": 0,
        "total_rows": 0,
        "filename": None,
        "error": None,
    })

    thread = threading.Thread(
        target=_run_thermometer_report_job,
        args=(base_dir, job_id, actor, config),
        daemon=True,
    )
    thread.start()

    return job_id
def _field_allowed_for_user(
    field_config: dict,
    can_view_extra_fields: bool = False,
    report_type: str = "ofs_os",
) -> bool:
    """
    Controla visibilidade/permissão dos campos de relatório.

    report_type:
      - ofs_os: relatório geral de OS
      - redes: relatório específico de Redes
    """
    report_type = str(report_type or "ofs_os").strip().lower()

    if field_config.get("redes_only") and report_type != "redes":
        return False

    if field_config.get("ofs_os_only") and report_type != "ofs_os":
        return False

    required_perm = field_config.get("required_perm")

    if not required_perm:
        return True

    if required_perm == EXTRA_REPORT_FIELDS_PERMISSION:
        return bool(can_view_extra_fields)

    return False


def list_report_field_choices(
    can_view_extra_fields: bool = False,
    report_type: str = "ofs_os",
) -> List[dict]:
    """
    Lista os campos disponíveis para o usuário na tela de relatório.

    Campos com redes_only=True aparecem somente no relatório de Redes.
    Campos com required_perm só aparecem para quem possui a permissão necessária.
    """
    visible_fields = []

    for field in FIELD_CHOICES:
        if _field_allowed_for_user(
            field,
            can_view_extra_fields=can_view_extra_fields,
            report_type=report_type,
        ):
            visible_fields.append(field)

    return visible_fields

def _now_iso():
    return datetime.now().isoformat(timespec="seconds")


def _ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def _job_status_path(base_dir: str, job_id: str) -> str:
    return os.path.join(base_dir, f"{job_id}.json")


def _job_xlsx_path(base_dir: str, job_id: str) -> str:
    return os.path.join(base_dir, f"{job_id}.xlsx")


def _write_job_status(base_dir: str, job_id: str, payload: dict):
    _ensure_dir(base_dir)

    path = _job_status_path(base_dir, job_id)

    payload["job_id"] = job_id
    payload["updated_at"] = _now_iso()

    tmp_path = os.path.join(
        base_dir,
        f"{job_id}.{uuid.uuid4().hex}.json.tmp"
    )

    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    last_error = None

    for attempt in range(12):
        try:
            os.replace(tmp_path, path)
            return

        except PermissionError as e:
            last_error = e
            time.sleep(0.15 * (attempt + 1))

        except OSError as e:
            last_error = e
            time.sleep(0.15 * (attempt + 1))

    try:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
    except Exception:
        pass

    raise last_error


def _load_resource_name_map() -> Dict[str, str]:
    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    try:
        cur.execute("""
            SELECT
                resource_id,
                name
            FROM relatorios_ofs_resources
        """)
        rows = cur.fetchall()

        mapping = {}

        for row in rows:
            resource_id = str(row.get("resource_id") or "").strip()
            name = str(row.get("name") or "").strip()

            if not resource_id:
                continue

            mapping[resource_id] = name or "Técnico não encontrado na base"

        return mapping

    finally:
        cur.close()
        conn.close()

def _load_resource_status_map() -> Dict[str, str]:
    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    try:
        cur.execute("""
            SELECT
                resource_id,
                status
            FROM relatorios_ofs_resources
        """)
        rows = cur.fetchall()

        mapping = {}

        for row in rows:
            resource_id = str(row.get("resource_id") or "").strip()
            status = str(row.get("status") or "").strip().lower()

            if not resource_id:
                continue

            if status == "active":
                mapping[resource_id] = "Ativo"
            elif status == "inactive":
                mapping[resource_id] = "Inativo"
            elif status:
                mapping[resource_id] = status
            else:
                mapping[resource_id] = "Não informado"

        return mapping

    finally:
        cur.close()
        conn.close()

def _load_activity_type_label_map() -> Dict[str, str]:
    mapping = {}

    for item in list_activity_types():
        code = str(item.get("code") or "").strip()
        label = str(item.get("label") or "").strip()

        if not code:
            continue

        mapping[code] = label or code

    return mapping

def _metadata_base_url(client: OFSClient) -> str:
    env_url = (os.getenv("OFS_METADATA_BASE_URL") or "").strip().rstrip("/")
    if env_url:
        return env_url

    core_marker = "/rest/ofscCore/v1"
    metadata_marker = "/rest/ofscMetadata/v1"

    if core_marker in client.base_url:
        return client.base_url.replace(core_marker, metadata_marker).rstrip("/")

    raise RuntimeError(
        "Não foi possível montar a URL metadata. "
        "Defina OFS_METADATA_BASE_URL no .env."
    )


def _safe_text(value) -> str:
    return str(value or "").strip()


def _extract_br_translation(item: dict) -> str:
    label = _safe_text(item.get("label"))

    translations = item.get("translations") or []

    for translation in translations:
        language = _safe_text(translation.get("language")).lower()
        language_iso = _safe_text(translation.get("languageISO")).lower()
        name = _safe_text(translation.get("name"))

        if name and (language == "br" or language_iso == "pt-br"):
            return name

    for translation in translations:
        name = _safe_text(translation.get("name"))
        if name:
            return name

    return label


def _fetch_property_enumeration_page(
    client: OFSClient,
    property_code: str,
    offset: int,
    limit: int = METADATA_ENUM_LIMIT,
) -> dict:
    base_url = _metadata_base_url(client)
    url = f"{base_url}/properties/{property_code}/enumerationList"

    response = requests.get(
        url,
        params={
            "limit": limit,
            "offset": offset,
        },
        headers={"Accept": "application/json"},
        auth=client.auth,
        timeout=REQUEST_TIMEOUT,
    )

    response.raise_for_status()
    return response.json()


def _fetch_property_enumerations(property_code: str) -> List[dict]:
    client = OFSClient()

    items = []
    offset = 0

    while True:
        payload = _fetch_property_enumeration_page(
            client=client,
            property_code=property_code,
            offset=offset,
            limit=METADATA_ENUM_LIMIT,
        )

        page_items = payload.get("items") or []
        has_more = bool(payload.get("hasMore"))

        items.extend(page_items)

        if not has_more or not page_items:
            break

        offset += len(page_items)

    return items


def sync_task_type_map(actor: dict = None) -> dict:
    """
    Atualiza a tabela local ofs_task_type_map com as enumerações da propriedade XA_TSK_TYP.
    """
    actor = actor or {}
    started_at = datetime.now()

    items = _fetch_property_enumerations(TASK_TYPE_PROPERTY_CODE)

    conn = get_connection()
    cur = conn.cursor()

    total_items = 0

    try:
        for item in items:
            label = _safe_text(item.get("label"))

            if not label:
                continue

            name_br = _extract_br_translation(item)
            active = 1 if bool(item.get("active")) else 0

            cur.execute(
                """
                INSERT INTO ofs_task_type_map (
                    property_code,
                    label,
                    name_br,
                    active
                )
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    name_br = VALUES(name_br),
                    active = VALUES(active),
                    updated_at = CURRENT_TIMESTAMP
                """,
                (
                    TASK_TYPE_PROPERTY_CODE,
                    label,
                    name_br,
                    active,
                ),
            )

            total_items += 1

        conn.commit()

    except Exception:
        conn.rollback()
        raise

    finally:
        cur.close()
        conn.close()

    try:
        audit_log(
            actor_user_id=actor.get("id"),
            actor_username=actor.get("username"),
            module="relatorios",
            action="sync_task_type_map",
            entity_type="ofs_property",
            entity_ref=TASK_TYPE_PROPERTY_CODE,
            summary="Atualizou mapa de tipos de tarefa do OFS",
            meta={
                "property_code": TASK_TYPE_PROPERTY_CODE,
                "total_items": total_items,
                "started_at": started_at.strftime("%Y-%m-%d %H:%M:%S"),
                "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            },
        )
    except Exception:
        pass

    return {
        "ok": True,
        "property_code": TASK_TYPE_PROPERTY_CODE,
        "started_at": started_at.strftime("%Y-%m-%d %H:%M:%S"),
        "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total_items": total_items,
    }


def _load_task_type_name_map() -> Dict[str, str]:
    """
    Carrega o mapa local de XA_TSK_TYP:
      label/ID -> name_br
    """
    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    try:
        cur.execute(
            """
            SELECT
                label,
                name_br
            FROM ofs_task_type_map
            WHERE property_code = %s
              AND active = 1
            """,
            (TASK_TYPE_PROPERTY_CODE,),
        )

        rows = cur.fetchall() or []

        mapping = {}

        for row in rows:
            label = _safe_text(row.get("label"))
            name_br = _safe_text(row.get("name_br"))

            if not label:
                continue

            mapping[label] = name_br or label

        return mapping

    finally:
        cur.close()
        conn.close()

def _load_close_reason_name_map() -> Dict[Tuple[str, str], str]:
    """
    Carrega o mapa de motivos de fechamento.

    Importante:
    O mesmo label/ID pode existir em propriedades diferentes com significados diferentes.
    Por isso a chave correta é sempre:
      (property_code, label)
    """
    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    try:
        cur.execute("""
            SELECT
                property_code,
                label,
                name_br
            FROM ofs_close_reason_map
            WHERE active = 1
        """)
        rows = cur.fetchall() or []

        mapping = {}

        for row in rows:
            property_code = str(row.get("property_code") or "").strip()
            label = str(row.get("label") or "").strip()
            name_br = str(row.get("name_br") or "").strip()

            if not property_code or not label:
                continue

            mapping[(property_code, label)] = name_br or label

        return mapping

    finally:
        cur.close()
        conn.close()

def read_job_status(base_dir: str, job_id: str) -> dict:
    path = _job_status_path(base_dir, job_id)

    if not os.path.exists(path):
        return {}

    last_error = None

    for _ in range(5):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)

        except json.JSONDecodeError as e:
            last_error = e
            time.sleep(0.1)

        except PermissionError as e:
            last_error = e
            time.sleep(0.1)

        except OSError as e:
            last_error = e
            time.sleep(0.1)

    print(f"[WARN] Falha ao ler status do job {job_id}: {last_error}")
    return {}
def discard_job(base_dir: str, job_id: str):
    for path in [_job_status_path(base_dir, job_id), _job_xlsx_path(base_dir, job_id)]:
        if os.path.exists(path):
            os.remove(path)


def get_xlsx_path(base_dir: str, job_id: str) -> str:
    return _job_xlsx_path(base_dir, job_id)


def list_resources_grouped() -> Dict[str, List[dict]]:
    grouped = {code: [] for code in RESOURCE_TYPES.keys()}

    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    try:
        cur.execute("""
            SELECT
                resource_id,
                resource_type,
                resource_type_label,
                name,
                status
            FROM relatorios_ofs_resources
            ORDER BY
                resource_type_label ASC,
                CASE WHEN status = 'active' THEN 1 ELSE 2 END,
                name ASC,
                resource_id ASC
        """)
        rows = cur.fetchall()

        for row in rows:
            resource_id = str(row.get("resource_id") or "").strip().upper()
            rtype = str(row.get("resource_type") or "").strip()

            if rtype == "GR" and resource_id in ESTADO_RESOURCE_IDS:
                grouped["ESTADO"].append(row)
                continue

            if rtype in grouped:
                grouped[rtype].append(row)

        return grouped

    finally:
        cur.close()
        conn.close()

def list_activity_types() -> List[dict]:
    """
    Lista tipos de atividade para montar os checkboxes da tela de relatório.

    Exibição para o usuário: label_pt, se existir.
    Valor enviado para a API: code.
    Categoria:
      - customer_home: atividade operacional/casa cliente
      - internal: atividade interna
    """
    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    try:
        cur.execute("""
            SELECT
                code,
                label_pt,
                category,
                include_in_bi,
                is_active
            FROM ofs_activity_type_map
            WHERE is_active = 1
            ORDER BY
                CASE
                    WHEN category = 'internal' THEN 2
                    ELSE 1
                END,
                label_pt ASC,
                code ASC
        """)
        rows = cur.fetchall()

        activity_types = []

        for row in rows:
            code = str(row.get("code") or "").strip()
            if not code:
                continue

            label_pt = (
                row.get("label_pt")
                or row.get("descricao")
                or row.get("description")
                or row.get("label")
                or row.get("name")
                or row.get("nome")
                or code
            )

            category = str(row.get("category") or "customer_home").strip().lower()

            if category not in {"customer_home", "internal","redes"}:
                category = "customer_home"

            activity_types.append({
                "code": code,
                "label": str(label_pt).strip() or code,
                "category": category,
                "is_internal": category == "internal",
            })

        return activity_types

    finally:
        cur.close()
        conn.close()
def _normalize_items_payload(data):
    if isinstance(data, dict):
        return data.get("items") or []
    if isinstance(data, list):
        return data
    return []

def list_ofs_os_activity_types() -> List[dict]:
    """
    Lista tipos usados no relatório geral de OS do OFS.

    Importante:
    Tipos de Redes ficam fora da extração geral para não misturar visões.
    """
    return [
        item for item in list_activity_types()
        if item.get("category") in {"customer_home", "internal"}
    ]


def list_redes_activity_types() -> List[dict]:
    """
    Lista somente tipos de atividade classificados como Redes.
    """
    return [
        item for item in list_activity_types()
        if item.get("category") == "redes"
    ]


def _valid_activity_type_codes_by_category(allowed_categories: set) -> set:
    valid_codes = set()

    for item in list_activity_types():
        category = str(item.get("category") or "").strip().lower()
        code = str(item.get("code") or "").strip()

        if code and category in allowed_categories:
            valid_codes.add(code)

    return valid_codes

def validate_ofs_os_report_payload(payload: dict, can_view_extra_fields: bool = False) -> dict:
    """
    Valida payload do relatório geral de OS.

    Bloqueia tipos de Redes para manter a tela geral separada da visão de Redes.
    """
    config = validate_report_payload(
        payload,
        can_view_extra_fields=can_view_extra_fields,
        report_type="ofs_os",
    )

    config["report_type"] = "ofs_os"

    allowed_codes = _valid_activity_type_codes_by_category({"customer_home", "internal"})
    invalid = [
        code for code in config["activity_types"]
        if code not in allowed_codes
    ]

    if invalid:
        raise ValueError(
            "O relatório geral de OS não permite tipos de atividade de Redes. "
            "Use o relatório específico de Redes."
        )

    return config

def validate_redes_report_payload(payload: dict, can_view_extra_fields: bool = False) -> dict:
    """
    Valida payload do relatório de Redes.

    Mesmo que alguém altere o payload pelo navegador, somente activityTypes com
    category='redes' serão aceitos.
    """
    config = validate_report_payload(
        payload,
        can_view_extra_fields=can_view_extra_fields,
        report_type="redes",
    )

    config["report_type"] = "redes"

    allowed_codes = _valid_activity_type_codes_by_category({"redes"})
    invalid = [
        code for code in config["activity_types"]
        if code not in allowed_codes
    ]

    if invalid:
        raise ValueError(
            "O relatório de Redes permite apenas tipos de atividade de Redes."
        )

    if not config["activity_types"]:
        raise ValueError("Nenhum tipo de atividade de Redes foi selecionado.")

    return config

def sync_resources_from_ofs(actor: dict) -> Tuple[int, int]:
    client = OFSClient()

    url = f"{client.base_url}/resources"
    headers = {"Accept": "application/json"}

    limit = 100
    offset = 0

    all_rows = []

    raw_total = 0
    active_total = 0
    inactive_total = 0
    allowed_type_total = 0
    ignored_type = {}

    while True:
        params = {
            "fields": "resourceId,status,resourceType,name",
            "limit": limit,
            "offset": offset,
        }

        resp = requests.get(
            url,
            headers=headers,
            params=params,
            auth=client.auth,
            timeout=REQUEST_TIMEOUT,
        )
        resp.raise_for_status()

        data = resp.json()
        items = _normalize_items_payload(data)

        if not items:
            break

        raw_total += len(items)

        for item in items:
            resource_id = str(item.get("resourceId") or "").strip()
            status = str(item.get("status") or "").strip().lower()
            resource_type = str(item.get("resourceType") or "").strip()
            name = str(item.get("name") or "").strip()

            if not resource_id:
                continue

            if not status:
                status = "unknown"

            if status == "active":
                active_total += 1
            else:
                inactive_total += 1

            if resource_type not in RESOURCE_TYPES_OFS:
                ignored_type[resource_type or "-"] = ignored_type.get(resource_type or "-", 0) + 1
                continue

            allowed_type_total += 1

            all_rows.append({
                "resource_id": resource_id,
                "resource_type": resource_type,
                "resource_type_label": RESOURCE_TYPES_OFS[resource_type],
                "name": name or None,
                "status": status,
            })

        if len(items) < limit:
            break

        offset += limit

    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("DELETE FROM relatorios_ofs_resources")

        sql = """
            INSERT INTO relatorios_ofs_resources
            (
                resource_id,
                resource_type,
                resource_type_label,
                name,
                status
            )
            VALUES (%s, %s, %s, %s, %s)
        """

        for row in all_rows:
            cur.execute(sql, (
                row["resource_id"],
                row["resource_type"],
                row["resource_type_label"],
                row["name"],
                row["status"],
            ))

        conn.commit()

    except Exception:
        conn.rollback()
        raise

    finally:
        cur.close()
        conn.close()

    audit_log(
        actor_user_id=actor.get("id"),
        actor_username=actor.get("username"),
        module="relatorios",
        action="sync_ofs_resources",
        entity_type="ofs_resources",
        summary="Atualizou lista de recursos OFS para relatórios",
        meta={
            "raw_total_from_api": raw_total,
            "active_total": active_total,
            "inactive_total": inactive_total,
            "allowed_type_total": allowed_type_total,
            "inserted_total": len(all_rows),
            "ignored_type": ignored_type,
            "resource_types_allowed": list(RESOURCE_TYPES_OFS.keys()),
            "limit": limit,
            "last_offset": offset,
        },
    )

    return len(all_rows), raw_total
def _validate_dates(date_from: str, date_to: str):
    try:
        dt_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        dt_to = datetime.strptime(date_to, "%Y-%m-%d").date()
    except Exception:
        raise ValueError("Informe um período válido.")

    if dt_to < dt_from:
        raise ValueError("A data final não pode ser menor que a data inicial.")
def _valid_activity_type_codes() -> set:
    return {item["code"] for item in list_activity_types()}

def _iter_date_strings(date_from: str, date_to: str) -> List[str]:
    start = datetime.strptime(date_from, "%Y-%m-%d").date()
    end = datetime.strptime(date_to, "%Y-%m-%d").date()

    current = start

    while current <= end:
        yield current.strftime("%Y-%m-%d")
        current += timedelta(days=1)
def _validate_activity_types(activity_types: list) -> List[str]:
    cleaned = []
    valid_codes = _valid_activity_type_codes()

    for activity_type in activity_types:
        value = str(activity_type or "").strip()

        if value in valid_codes and value not in cleaned:
            cleaned.append(value)

    if not cleaned:
        raise ValueError("Selecione pelo menos um tipo de atividade.")

    return cleaned

def _validate_statuses(statuses: list) -> List[str]:
    cleaned = []
    allowed = set(STATUS_OPTIONS)

    for status in statuses:
        value = str(status or "").strip()
        if value in allowed and value not in cleaned:
            cleaned.append(value)

    if not cleaned:
        raise ValueError("Selecione pelo menos um status.")

    return cleaned


def _validate_fields(
    fields: list,
    can_view_extra_fields: bool = False,
    report_type: str = "ofs_os",
) -> List[str]:
    cleaned = []
    allowed = set(FIELD_MAP.keys())

    for field in fields:
        value = str(field or "").strip()
        if value in allowed and value not in cleaned:
            cleaned.append(value)

    if not cleaned:
        raise ValueError("Selecione pelo menos um campo para extração.")

    denied_fields = []

    for field_key in cleaned:
        field_config = FIELD_MAP[field_key]

        if not _field_allowed_for_user(
            field_config,
            can_view_extra_fields=can_view_extra_fields,
            report_type=report_type,
        ):
            denied_fields.append(
                field_config.get("xlsx_header")
                or field_config.get("label")
                or field_key
            )

    if denied_fields:
        raise ValueError(
            "Você não tem permissão para extrair os seguintes campos neste relatório: "
            + ", ".join(denied_fields)
        )

    return cleaned

def _validate_resources(resource_ids: list) -> List[str]:
    selected = []
    for rid in resource_ids:
        value = str(rid or "").strip()
        if value and value not in selected:
            selected.append(value)

    if not selected:
        raise ValueError("Selecione pelo menos um recurso.")

    conn = get_connection()
    cur = conn.cursor()

    try:
        placeholders = ",".join(["%s"] * len(selected))
        cur.execute(f"""
            SELECT resource_id
            FROM relatorios_ofs_resources
            WHERE resource_id IN ({placeholders})
        """, selected)

        valid = {row[0] for row in cur.fetchall()}

    finally:
        cur.close()
        conn.close()

    invalid = [rid for rid in selected if rid not in valid]
    if invalid:
        raise ValueError("Um ou mais recursos selecionados não existem na lista local. Atualize a lista de recursos.")

    return selected

def validate_report_payload(
    payload: dict,
    can_view_extra_fields: bool = False,
    report_type: str = "ofs_os",
) -> dict:
    report_type = str(report_type or "ofs_os").strip().lower()

    date_from = str(payload.get("dateFrom") or "").strip()
    date_to = str(payload.get("dateTo") or "").strip()

    _validate_dates(date_from, date_to)

    statuses = _validate_statuses(payload.get("statuses") or [])
    activity_types = _validate_activity_types(payload.get("activityTypes") or [])
    selected_fields = _validate_fields(
        payload.get("fields") or [],
        can_view_extra_fields=can_view_extra_fields,
        report_type=report_type,
    )
    resource_ids = _validate_resources(payload.get("resources") or [])

    return {
        "report_type": report_type,
        "date_from": date_from,
        "date_to": date_to,
        "statuses": statuses,
        "activity_types": activity_types,
        "fields": selected_fields,
        "resources": resource_ids,
    }

def _closure_fields_for_report(report_type: str = "ofs_os") -> List[str]:
    report_type = str(report_type or "ofs_os").strip().lower()

    if report_type == "redes":
        return REDES_CLOSURE_FIELDS

    return CLOSURE_FIELDS
def _api_fields_for_selected(
    selected_fields: List[str],
    report_type: str = "ofs_os",
) -> List[str]:
    api_fields = []

    for key in selected_fields:
        if key == "fechamento_atividade":
            fields = _closure_fields_for_report(report_type)
        else:
            cfg = FIELD_MAP[key]
            fields = cfg["api_fields"]

        for api_field in fields:
            if api_field not in api_fields:
                api_fields.append(api_field)

    return api_fields

def _first_name(value):
    value = str(value or "").strip()
    if not value:
        return ""
    return value.split()[0]


def _first_filled(item: dict, keys: list):
    for key in keys:
        value = item.get(key)
        if value is None:
            continue

        value_str = str(value).strip()
        if value_str:
            return value_str

    return ""


def _row_value(
    item: dict,
    field_key: str,
    resource_name_map: Dict[str, str] = None,
    resource_status_map: Dict[str, str] = None,
    activity_type_label_map: Dict[str, str] = None,
    close_reason_name_map: Dict[Tuple[str, str], str] = None,
    task_type_name_map: Dict[str, str] = None,
    report_type: str = "ofs_os",
):
    if field_key == "customerName":
        return _first_name(item.get("customerName"))
    if field_key == "customerNameFull":
        return str(item.get("customerName") or "").strip()
    if field_key == "fechamento_atividade":
        for closure_field in _closure_fields_for_report(report_type):
            raw_value = item.get(closure_field)

            if raw_value is None:
                continue

            label = str(raw_value).strip()

            if not label:
                continue

            if close_reason_name_map:
                translated = close_reason_name_map.get((closure_field, label))
                if translated:
                    return translated

            return label

        return ""

    if field_key == "resourceName":
        resource_id = str(item.get("resourceId") or "").strip()
        if not resource_id:
            return "Técnico não encontrado na base"

        if resource_name_map and resource_id in resource_name_map:
            return resource_name_map[resource_id] or "Técnico não encontrado na base"

        return "Técnico não encontrado na base"

    if field_key == "resourceStatus":
        resource_id = str(item.get("resourceId") or "").strip()

        if not resource_id:
            return "Não encontrado"

        if resource_status_map and resource_id in resource_status_map:
            return resource_status_map[resource_id]

        return "Não encontrado"

    if field_key == "activityType":
        activity_code = str(item.get("activityType") or "").strip()

        if not activity_code:
            return ""

        if activity_type_label_map and activity_code in activity_type_label_map:
            return activity_type_label_map[activity_code]

        return activity_code
    if field_key == TASK_TYPE_PROPERTY_CODE:
        raw_value = str(item.get(TASK_TYPE_PROPERTY_CODE) or "").strip()

        if not raw_value:
            return ""

        if task_type_name_map and raw_value in task_type_name_map:
            return task_type_name_map[raw_value]

        return raw_value
    value = item.get(field_key)

    if value is None:
        return ""

    return value

def _build_or_equals_query(field_name: str, values: List[str]) -> str:
    cleaned = []

    for value in values:
        value = str(value or "").strip()
        if not value:
            continue

        value = value.replace("'", "\\'")
        cleaned.append(value)

    if not cleaned:
        raise ValueError(f"Nenhum valor informado para o filtro {field_name}.")

    if len(cleaned) == 1:
        return f"{field_name}=='{cleaned[0]}'"

    parts = [f"{field_name}=='{value}'" for value in cleaned]
    return "(" + " OR ".join(parts) + ")"

def _fetch_activities(client: OFSClient, config: dict, base_dir: str, job_id: str, status_payload: dict) -> List[dict]:
    url = f"{client.base_url}/activities/"
    headers = {"Accept": "application/json"}

    api_fields = _api_fields_for_selected(
        config["fields"],
        report_type=config.get("report_type", "ofs_os"),
    )

    status_payload["api_fields"] = api_fields
    for required_field in ("activityId", "status", "activityType", "date"):
        if required_field not in api_fields:
            api_fields.append(required_field)

    resources_param = ",".join(config["resources"])

    status_query = _build_or_equals_query("status", config["statuses"])
    activity_type_query = _build_or_equals_query("activityType", config["activity_types"])

    # IMPORTANTE:
    # O OFS deve receber apenas UM parâmetro q contendo a expressão completa.
    combined_query = f"{status_query} and {activity_type_query}"

    all_items = []
    seen_activity_ids = set()

    daily_counts = {}
    daily_raw_counts = {}
    duplicate_activity_ids = 0
    total_raw_rows = 0
    total_pages_processed = 0

    date_list = list(_iter_date_strings(config["date_from"], config["date_to"]))
    total_days = len(date_list)

    for day_index, day in enumerate(date_list, start=1):
        offset = 0
        page = 1
        day_count = 0
        day_raw_count = 0

        while True:
            params = [
                ("dateFrom", day),
                ("dateTo", day),
                ("resources", resources_param),
                ("q", combined_query),
                ("fields", ",".join(api_fields)),
                ("limit", str(API_LIMIT)),
                ("offset", str(offset)),
            ]

            status_payload.update({
                "status": "running",
                "phase": f"Consultando OFS - {day} - página {page}, offset {offset}",
                "rows_so_far": len(all_items),
                "raw_rows_so_far": total_raw_rows,
                "q": combined_query,
                "current_day": day,
                "current_day_index": day_index,
                "total_days": total_days,
                "offset": offset,
                "page": page,
                "total_pages_processed": total_pages_processed,
                "daily_counts": daily_counts,
                "daily_raw_counts": daily_raw_counts,
                "duplicate_activity_ids": duplicate_activity_ids,
            })
            _write_job_status(base_dir, job_id, status_payload)

            resp = requests.get(
                url,
                headers=headers,
                params=params,
                auth=client.auth,
                timeout=REQUEST_TIMEOUT,
            )
            resp.raise_for_status()

            data = resp.json()
            items = _normalize_items_payload(data)

            returned_count = len(items)
            day_raw_count += returned_count
            total_raw_rows += returned_count
            total_pages_processed += 1

            for item in items:
                activity_id = str(item.get("activityId") or "").strip()

                if activity_id:
                    if activity_id in seen_activity_ids:
                        duplicate_activity_ids += 1
                        continue

                    seen_activity_ids.add(activity_id)

                all_items.append(item)
                day_count += 1

            daily_counts[day] = day_count
            daily_raw_counts[day] = day_raw_count

            status_payload.update({
                "status": "running",
                "phase": f"Processando OFS - {day} - página {page}",
                "rows_so_far": len(all_items),
                "raw_rows_so_far": total_raw_rows,
                "current_day": day,
                "current_day_index": day_index,
                "total_days": total_days,
                "offset": offset,
                "page": page,
                "total_pages_processed": total_pages_processed,
                "daily_counts": daily_counts,
                "daily_raw_counts": daily_raw_counts,
                "duplicate_activity_ids": duplicate_activity_ids,
            })
            _write_job_status(base_dir, job_id, status_payload)

            has_more = bool(data.get("hasMore")) if isinstance(data, dict) else False

            if not has_more:
                break

            if returned_count <= 0:
                raise RuntimeError(
                    f"A API informou hasMore=true para o dia {day}, mas não retornou itens. "
                    "A consulta pode estar pesada demais ou excedendo o tempo limite do OFS."
                )

            offset += returned_count
            page += 1

    status_payload.update({
        "status": "running",
        "phase": "Consulta OFS concluída. Preparando XLSX.",
        "rows_so_far": len(all_items),
        "total_rows": len(all_items),
        "raw_rows_so_far": total_raw_rows,
        "total_raw_rows": total_raw_rows,
        "total_pages_processed": total_pages_processed,
        "daily_counts": daily_counts,
        "daily_raw_counts": daily_raw_counts,
        "duplicate_activity_ids": duplicate_activity_ids,
    })
    _write_job_status(base_dir, job_id, status_payload)

    return all_items

def _build_activity_report_summary(rows: List[dict]) -> dict:
    """
    Monta resumo visual para relatórios operacionais.

    Usado inicialmente no relatório de Redes.
    """
    resource_name_map = _load_resource_name_map()
    activity_type_label_map = _load_activity_type_label_map()

    by_status = {}
    by_resource = {}
    by_activity_type = {}

    for item in rows:
        status = str(item.get("status") or "Não informado").strip() or "Não informado"
        resource_id = str(item.get("resourceId") or "").strip()
        activity_type_code = str(item.get("activityType") or "").strip()

        by_status[status] = by_status.get(status, 0) + 1

        activity_type_label = activity_type_label_map.get(activity_type_code, activity_type_code or "Não informado")
        by_activity_type[activity_type_label] = by_activity_type.get(activity_type_label, 0) + 1

        resource_key = resource_id or "sem_resource"
        resource_name = resource_name_map.get(resource_id, "Técnico não encontrado na base")

        if resource_key not in by_resource:
            by_resource[resource_key] = {
                "resource_id": resource_id or "-",
                "resource_name": resource_name,
                "total": 0,
                "by_status": {},
            }

        by_resource[resource_key]["total"] += 1
        by_resource[resource_key]["by_status"][status] = (
            by_resource[resource_key]["by_status"].get(status, 0) + 1
        )

    by_resource_list = list(by_resource.values())
    by_resource_list.sort(
        key=lambda item: (
            int(item.get("total") or 0),
            str(item.get("resource_name") or "")
        ),
        reverse=True,
    )

    by_activity_type_list = [
        {
            "label": label,
            "total": total,
        }
        for label, total in by_activity_type.items()
    ]
    by_activity_type_list.sort(key=lambda item: item["total"], reverse=True)

    by_status_list = [
        {
            "status": status,
            "total": total,
        }
        for status, total in by_status.items()
    ]
    by_status_list.sort(key=lambda item: item["total"], reverse=True)

    return {
        "total": len(rows),
        "by_status": by_status,
        "by_status_list": by_status_list,
        "by_resource": by_resource_list,
        "by_activity_type": by_activity_type_list,
    }

def _build_xlsx(
    rows: List[dict],
    selected_fields: List[str],
    output_path: str,
    report_type: str = "ofs_os",
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório OS OFS"

    headers = [FIELD_MAP[key]["xlsx_header"] for key in selected_fields]
    ws.append(headers)

    resource_name_map = _load_resource_name_map()
    resource_status_map = (
        _load_resource_status_map()
        if "resourceStatus" in selected_fields
        else {}
    )
    activity_type_label_map = (
        _load_activity_type_label_map()
        if "activityType" in selected_fields
        else {}
    )
    close_reason_name_map = (
        _load_close_reason_name_map()
        if "fechamento_atividade" in selected_fields
        else {}
    )
    task_type_name_map = (
        _load_task_type_name_map()
        if TASK_TYPE_PROPERTY_CODE in selected_fields
        else {}
    )
    for item in rows:
        ws.append([
            _row_value(
                item,
                key,
                resource_name_map=resource_name_map,
                resource_status_map=resource_status_map,
                activity_type_label_map=activity_type_label_map,
                close_reason_name_map=close_reason_name_map,
                task_type_name_map=task_type_name_map,
                report_type=report_type,
            )
            for key in selected_fields
        ])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top")

    ws.freeze_panes = "A2"
    xlsx_auto_width(ws)

    wb.save(output_path)
def _run_report_job(base_dir: str, job_id: str, actor: dict, config: dict):
    filename = f"relatorio_os_ofs_{config['date_from']}_{config['date_to']}_{job_id[:8]}.xlsx"

    status_payload = {
        "status": "running",
        "phase": "Iniciando extração",
        "created_at": _now_iso(),
        "finished_at": None,
        "rows_so_far": 0,
        "total_rows": 0,
        "filename": filename,
        "dateFrom": config["date_from"],
        "dateTo": config["date_to"],
        "statuses": config["statuses"],
        "resources": config["resources"],
        "activity_types": config["activity_types"],
        "fields": config["fields"],
        "error": None,
    }

    _write_job_status(base_dir, job_id, status_payload)

    try:
        audit_log(
            actor_user_id=actor.get("id"),
            actor_username=actor.get("username"),
            module="relatorios",
            action="start_ofs_os_report",
            entity_type="report",
            entity_ref=job_id,
            summary="Iniciou extração do Relatório de OS do OFS",
            meta={
                "dateFrom": config["date_from"],
                "dateTo": config["date_to"],
                "statuses": config["statuses"],
                "resources": config["resources"],
                "activity_types": config["activity_types"],
                "fields": config["fields"],
            },
        )

        client = OFSClient()
        config["report_type"] = "ofs_os"
        rows = _fetch_activities(client, config, base_dir, job_id, status_payload)

        status_payload.update({
            "status": "running",
            "phase": "Gerando XLSX",
            "rows_so_far": len(rows),
            "total_rows": len(rows),
            "total_raw_rows": status_payload.get("total_raw_rows"),
            "duplicate_activity_ids": status_payload.get("duplicate_activity_ids"),
            "total_pages_processed": status_payload.get("total_pages_processed"),
            "daily_counts": status_payload.get("daily_counts"),
            "daily_raw_counts": status_payload.get("daily_raw_counts"),
        })
        _write_job_status(base_dir, job_id, status_payload)

        output_path = _job_xlsx_path(base_dir, job_id)
        _build_xlsx(rows, config["fields"], output_path, report_type="ofs_os")

        status_payload.update({
            "status": "completed",
            "phase": "Extração concluída",
            "finished_at": _now_iso(),
            "rows_so_far": len(rows),
            "total_rows": len(rows),
            "download_ready": True,
        })
        _write_job_status(base_dir, job_id, status_payload)

        audit_log(
            actor_user_id=actor.get("id"),
            actor_username=actor.get("username"),
            module="relatorios",
            action="finish_ofs_os_report",
            entity_type="report",
            entity_ref=job_id,
            summary="Concluiu extração do Relatório de OS do OFS",
            meta={
                "dateFrom": config["date_from"],
                "dateTo": config["date_to"],
                "statuses": config["statuses"],
                "resources": config["resources"],
                "activity_types": config["activity_types"],
                "fields": config["fields"],
                "total_rows": len(rows),
                "total_raw_rows": status_payload.get("total_raw_rows"),
                "duplicate_activity_ids": status_payload.get("duplicate_activity_ids"),
                "total_pages_processed": status_payload.get("total_pages_processed"),
                "daily_counts": status_payload.get("daily_counts"),
                "daily_raw_counts": status_payload.get("daily_raw_counts"),
                "filename": filename,
            },
        )

    except Exception as e:
        status_payload.update({
            "status": "failed",
            "phase": "Falha na extração",
            "finished_at": _now_iso(),
            "error": str(e),
        })
        _write_job_status(base_dir, job_id, status_payload)

        audit_log(
            actor_user_id=actor.get("id"),
            actor_username=actor.get("username"),
            module="relatorios",
            action="fail_ofs_os_report",
            entity_type="report",
            entity_ref=job_id,
            summary="Falha na extração do Relatório de OS do OFS",
            meta={
                "dateFrom": config["date_from"],
                "dateTo": config["date_to"],
                "statuses": config["statuses"],
                "resources": config["resources"],
                "activity_types": config["activity_types"],
                "fields": config["fields"],
                "error": str(e),
            },
        )

def _run_redes_report_job(base_dir: str, job_id: str, actor: dict, config: dict):
    filename = f"relatorio_redes_{config['date_from']}_{config['date_to']}_{job_id[:8]}.xlsx"

    status_payload = {
        "status": "running",
        "phase": "Iniciando extração de Redes",
        "created_at": _now_iso(),
        "finished_at": None,
        "rows_so_far": 0,
        "total_rows": 0,
        "filename": filename,
        "dateFrom": config["date_from"],
        "dateTo": config["date_to"],
        "statuses": config["statuses"],
        "resources": config["resources"],
        "activity_types": config["activity_types"],
        "fields": config["fields"],
        "summary": None,
        "error": None,
    }

    _write_job_status(base_dir, job_id, status_payload)

    try:
        audit_log(
            actor_user_id=actor.get("id"),
            actor_username=actor.get("username"),
            module="relatorios",
            action="start_redes_report",
            entity_type="report",
            entity_ref=job_id,
            summary="Iniciou extração do Relatório de Redes",
            meta={
                "dateFrom": config["date_from"],
                "dateTo": config["date_to"],
                "statuses": config["statuses"],
                "resources": config["resources"],
                "activity_types": config["activity_types"],
                "fields": config["fields"],
            },
        )

        client = OFSClient()
        config["report_type"] = "redes"
        rows = _fetch_activities(client, config, base_dir, job_id, status_payload)

        status_payload.update({
            "status": "running",
            "phase": "Gerando resumo de Redes",
            "rows_so_far": len(rows),
            "total_rows": len(rows),
            "total_raw_rows": status_payload.get("total_raw_rows"),
            "duplicate_activity_ids": status_payload.get("duplicate_activity_ids"),
            "total_pages_processed": status_payload.get("total_pages_processed"),
            "daily_counts": status_payload.get("daily_counts"),
            "daily_raw_counts": status_payload.get("daily_raw_counts"),
        })
        _write_job_status(base_dir, job_id, status_payload)

        summary_payload = _build_activity_report_summary(rows)

        status_payload.update({
            "status": "running",
            "phase": "Gerando XLSX",
            "summary": summary_payload,
        })
        _write_job_status(base_dir, job_id, status_payload)

        output_path = _job_xlsx_path(base_dir, job_id)
        _build_xlsx(rows, config["fields"], output_path, report_type="redes")

        status_payload.update({
            "status": "completed",
            "phase": "Extração de Redes concluída",
            "finished_at": _now_iso(),
            "rows_so_far": len(rows),
            "total_rows": len(rows),
            "summary": summary_payload,
            "download_ready": True,
        })
        _write_job_status(base_dir, job_id, status_payload)

        audit_log(
            actor_user_id=actor.get("id"),
            actor_username=actor.get("username"),
            module="relatorios",
            action="finish_redes_report",
            entity_type="report",
            entity_ref=job_id,
            summary="Concluiu extração do Relatório de Redes",
            meta={
                "dateFrom": config["date_from"],
                "dateTo": config["date_to"],
                "statuses": config["statuses"],
                "resources": config["resources"],
                "activity_types": config["activity_types"],
                "fields": config["fields"],
                "total_rows": len(rows),
                "total_raw_rows": status_payload.get("total_raw_rows"),
                "duplicate_activity_ids": status_payload.get("duplicate_activity_ids"),
                "total_pages_processed": status_payload.get("total_pages_processed"),
                "daily_counts": status_payload.get("daily_counts"),
                "daily_raw_counts": status_payload.get("daily_raw_counts"),
                "filename": filename,
                "summary": summary_payload,
            },
        )

    except Exception as e:
        status_payload.update({
            "status": "failed",
            "phase": "Falha na extração de Redes",
            "finished_at": _now_iso(),
            "error": str(e),
        })
        _write_job_status(base_dir, job_id, status_payload)

        audit_log(
            actor_user_id=actor.get("id"),
            actor_username=actor.get("username"),
            module="relatorios",
            action="fail_redes_report",
            entity_type="report",
            entity_ref=job_id,
            summary="Falha na extração do Relatório de Redes",
            meta={
                "dateFrom": config["date_from"],
                "dateTo": config["date_to"],
                "statuses": config["statuses"],
                "resources": config["resources"],
                "activity_types": config["activity_types"],
                "fields": config["fields"],
                "error": str(e),
            },
        )


def start_redes_report_job(base_dir: str, actor: dict, config: dict) -> str:
    _ensure_dir(base_dir)

    job_id = uuid.uuid4().hex

    initial_payload = {
        "status": "queued",
        "phase": "Extração de Redes na fila",
        "created_at": _now_iso(),
        "finished_at": None,
        "rows_so_far": 0,
        "total_rows": 0,
        "filename": None,
        "summary": None,
        "error": None,
    }
    _write_job_status(base_dir, job_id, initial_payload)

    thread = threading.Thread(
        target=_run_redes_report_job,
        args=(base_dir, job_id, actor, config),
        daemon=True,
    )
    thread.start()

    return job_id

def start_report_job(base_dir: str, actor: dict, config: dict) -> str:
    _ensure_dir(base_dir)

    job_id = uuid.uuid4().hex

    initial_payload = {
        "status": "queued",
        "phase": "Extração na fila",
        "created_at": _now_iso(),
        "finished_at": None,
        "rows_so_far": 0,
        "total_rows": 0,
        "filename": None,
        "error": None,
    }
    _write_job_status(base_dir, job_id, initial_payload)

    thread = threading.Thread(
        target=_run_report_job,
        args=(base_dir, job_id, actor, config),
        daemon=True,
    )
    thread.start()

    return job_id

RESOURCE_SYNC_LOCK_STALE_SECONDS = 6 * 60 * 60


def _resource_sync_lock_path(base_dir: str) -> str:
    return os.path.join(base_dir, "resource_sync.lock")


def _read_resource_sync_lock(base_dir: str) -> dict:
    path = _resource_sync_lock_path(base_dir)
    if not os.path.exists(path):
        return {}

    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _remove_resource_sync_lock(base_dir: str):
    path = _resource_sync_lock_path(base_dir)
    if os.path.exists(path):
        os.remove(path)


def _lock_is_stale(lock_payload: dict) -> bool:
    created_ts = float(lock_payload.get("created_ts") or 0)
    if not created_ts:
        return True

    return (time.time() - created_ts) > RESOURCE_SYNC_LOCK_STALE_SECONDS


def _try_create_resource_sync_lock(base_dir: str, job_id: str, actor: dict):
    _ensure_dir(base_dir)

    lock_path = _resource_sync_lock_path(base_dir)
    current_lock = _read_resource_sync_lock(base_dir)

    if current_lock and _lock_is_stale(current_lock):
        _remove_resource_sync_lock(base_dir)

    payload = {
        "job_id": job_id,
        "actor_user_id": actor.get("id"),
        "actor_username": actor.get("username"),
        "created_at": _now_iso(),
        "created_ts": time.time(),
    }

    try:
        fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

        return True, None

    except FileExistsError:
        return False, _read_resource_sync_lock(base_dir)


def _extract_total_from_resource_payload(data: dict):
    if not isinstance(data, dict):
        return None

    for key in ("totalResults", "total", "totalCount", "count"):
        value = data.get(key)
        try:
            value_int = int(value)
            if value_int > 0:
                return value_int
        except Exception:
            pass

    return None


def _resource_percent(raw_total: int, total_expected):
    if total_expected:
        return min(95, int((raw_total / total_expected) * 100))

    # Fallback quando a API não informa total.
    # Não é percentual real, mas evita a barra parada até finalizar.
    return min(95, max(5, int(raw_total / 30)))


def _sync_resources_from_ofs_with_progress(base_dir: str, job_id: str, actor: dict, status_payload: dict):
    client = OFSClient()

    url = f"{client.base_url}/resources"
    headers = {"Accept": "application/json"}

    limit = 100
    offset = 0
    page = 1
    max_pages = 500

    all_rows = []

    raw_total = 0
    active_total = 0
    inactive_total = 0
    allowed_type_total = 0
    ignored_type = {}
    total_expected = None

    while True:
        params = {
            "fields": "resourceId,status,resourceType,name",
            "limit": limit,
            "offset": offset,
        }

        status_payload.update({
            "status": "running",
            "phase": f"Consultando recursos no OFS - página {page}",
            "percent": _resource_percent(raw_total, total_expected),
            "raw_total_from_api": raw_total,
            "inserted_so_far": len(all_rows),
            "page": page,
            "offset": offset,
            "total_expected": total_expected,
        })
        _write_job_status(base_dir, job_id, status_payload)

        resp = requests.get(
            url,
            headers=headers,
            params=params,
            auth=client.auth,
            timeout=REQUEST_TIMEOUT,
        )
        resp.raise_for_status()

        data = resp.json()

        if total_expected is None:
            total_expected = _extract_total_from_resource_payload(data)

        items = _normalize_items_payload(data)

        if not items:
            break

        raw_total += len(items)

        for item in items:
            resource_id = str(item.get("resourceId") or "").strip()
            status = str(item.get("status") or "").strip().lower()
            resource_type = str(item.get("resourceType") or "").strip()
            name = str(item.get("name") or "").strip()

            if not resource_id:
                continue

            if not status:
                status = "unknown"

            if status == "active":
                active_total += 1
            else:
                inactive_total += 1

            if resource_type not in RESOURCE_TYPES_OFS:
                ignored_type[resource_type or "-"] = ignored_type.get(resource_type or "-", 0) + 1
                continue

            allowed_type_total += 1

            all_rows.append({
                "resource_id": resource_id,
                "resource_type": resource_type,
                "resource_type_label": RESOURCE_TYPES_OFS[resource_type],
                "name": name or None,
                "status": status,
            })

        status_payload.update({
            "status": "running",
            "phase": f"Processando recursos - página {page}",
            "percent": _resource_percent(raw_total, total_expected),
            "raw_total_from_api": raw_total,
            "inserted_so_far": len(all_rows),
            "page": page,
            "offset": offset,
            "total_expected": total_expected,
            "active_total": active_total,
            "inactive_total": inactive_total,
        })
        _write_job_status(base_dir, job_id, status_payload)

        if len(items) < limit:
            break

        offset += limit
        page += 1

        if page > max_pages:
            raise RuntimeError("Limite máximo de páginas atingido ao consultar recursos OFS.")

    status_payload.update({
        "status": "running",
        "phase": "Atualizando tabela de recursos no banco",
        "percent": 96,
        "raw_total_from_api": raw_total,
        "inserted_so_far": len(all_rows),
        "active_total": active_total,
        "inactive_total": inactive_total,
    })
    _write_job_status(base_dir, job_id, status_payload)

    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("DELETE FROM relatorios_ofs_resources")

        sql = """
            INSERT INTO relatorios_ofs_resources
            (
                resource_id,
                resource_type,
                resource_type_label,
                name,
                status
            )
            VALUES (%s, %s, %s, %s, %s)
        """

        values = [
            (
                row["resource_id"],
                row["resource_type"],
                row["resource_type_label"],
                row["name"],
                row["status"],
            )
            for row in all_rows
        ]

        if values:
            cur.executemany(sql, values)

        conn.commit()

    except Exception:
        conn.rollback()
        raise

    finally:
        cur.close()
        conn.close()

    return {
        "raw_total_from_api": raw_total,
        "active_total": active_total,
        "inactive_total": inactive_total,
        "allowed_type_total": allowed_type_total,
        "inserted_total": len(all_rows),
        "ignored_type": ignored_type,
        "resource_types_allowed": list(RESOURCE_TYPES_OFS.keys()),
        "limit": limit,
        "last_offset": offset,
        "total_expected": total_expected,
    }
def _run_resource_sync_job(base_dir: str, job_id: str, actor: dict):
    status_payload = {
        "status": "running",
        "phase": "Iniciando atualização da lista de recursos",
        "created_at": _now_iso(),
        "finished_at": None,
        "percent": 0,
        "raw_total_from_api": 0,
        "inserted_so_far": 0,
        "inserted_total": 0,
        "error": None,
    }

    _write_job_status(base_dir, job_id, status_payload)

    try:
        audit_log(
            actor_user_id=actor.get("id"),
            actor_username=actor.get("username"),
            module="relatorios",
            action="start_sync_ofs_resources",
            entity_type="ofs_resources",
            entity_ref=job_id,
            summary="Iniciou atualização da lista de recursos OFS em segundo plano",
            meta={
                "resource_types": list(RESOURCE_TYPES.keys()),
            },
        )

        result = _sync_resources_from_ofs_with_progress(base_dir, job_id, actor, status_payload)

        status_payload.update({
            "status": "completed",
            "phase": "Lista de recursos atualizada com sucesso",
            "finished_at": _now_iso(),
            "percent": 100,
            "raw_total_from_api": result.get("raw_total_from_api", 0),
            "inserted_total": result.get("inserted_total", 0),
            "inserted_so_far": result.get("inserted_total", 0),
            "result": result,
        })
        _write_job_status(base_dir, job_id, status_payload)

        audit_log(
            actor_user_id=actor.get("id"),
            actor_username=actor.get("username"),
            module="relatorios",
            action="finish_sync_ofs_resources",
            entity_type="ofs_resources",
            entity_ref=job_id,
            summary="Concluiu atualização da lista de recursos OFS",
            meta=result,
        )

    except Exception as e:
        status_payload.update({
            "status": "failed",
            "phase": "Falha ao atualizar lista de recursos",
            "finished_at": _now_iso(),
            "error": str(e),
        })
        _write_job_status(base_dir, job_id, status_payload)

        audit_log(
            actor_user_id=actor.get("id"),
            actor_username=actor.get("username"),
            module="relatorios",
            action="fail_sync_ofs_resources",
            entity_type="ofs_resources",
            entity_ref=job_id,
            summary="Falha ao atualizar lista de recursos OFS",
            meta={
                "error": str(e),
            },
        )

    finally:
        _remove_resource_sync_lock(base_dir)


def start_resource_sync_job(base_dir: str, actor: dict):
    _ensure_dir(base_dir)

    job_id = uuid.uuid4().hex

    locked, active_lock = _try_create_resource_sync_lock(base_dir, job_id, actor)

    if not locked:
        return None, active_lock

    initial_payload = {
        "status": "queued",
        "phase": "Atualização de recursos na fila",
        "created_at": _now_iso(),
        "finished_at": None,
        "percent": 0,
        "raw_total_from_api": 0,
        "inserted_so_far": 0,
        "inserted_total": 0,
        "error": None,
    }
    _write_job_status(base_dir, job_id, initial_payload)

    thread = threading.Thread(
        target=_run_resource_sync_job,
        args=(base_dir, job_id, actor),
        daemon=True,
    )
    thread.start()

    return job_id, None


def read_resource_sync_job_status(base_dir: str, job_id: str) -> dict:
    return read_job_status(base_dir, job_id)